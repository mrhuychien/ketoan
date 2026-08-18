"""misa_import — nạp danh sách hóa đơn MISA từ file Excel/CSV xuất ra từ MISA.

Vì sao có đường này: bề mặt API /api/v2 chỉ cho ĐỌC CHI TIẾT theo RefID và ĐẨY
hóa đơn; endpoint liệt kê trả rỗng (§O). Không có danh sách thì rổ "Chỉ có trên
MISA" — thứ phát hiện hóa đơn xuất ngoài sổ — vô nghĩa.

Đường này không phụ thuộc API: kế toán bấm xuất Excel trên MISA rồi tải lên.
Toàn bộ tầng đối soát phía sau (khớp 4 tầng, chốt tay) dùng chung, không đổi.

Tên cột trong file MISA xuất ra CHƯA được xác minh, và MISA có thể đổi bất cứ
lúc nào — nên KHÔNG hardcode vị trí cột. Hàm dò cột theo từ khóa đã bỏ dấu, và
LUÔN trả bản đồ cột ra cho người dùng duyệt trước khi ghi.
"""

import base64
import io
import re
import unicodedata

import frappe
from frappe import _
from frappe.utils import flt, now_datetime

from ketoan.api._guard import guard_manager, guard_sales_any

MAX_PREVIEW = 15
MAX_ROWS = 20000


# ═══════════════════════════════════════════════════════════════════════════
# Chuẩn hóa & dò cột
# ═══════════════════════════════════════════════════════════════════════════

def strip_tones(s) -> str:
    """Bỏ dấu tiếng Việt + hạ chữ thường, để so tên cột không phụ thuộc dấu."""
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.replace("đ", "d").replace("Đ", "D").lower().strip()


# Mỗi field: danh sách bộ từ khóa. Cột khớp khi CHỨA ĐỦ mọi từ trong một bộ.
# Xếp bộ cụ thể lên trước bộ chung — "tong tien thanh toan" phải thắng "tien".
COLUMN_RULES = {
    "inv_series": [("ky hieu",), ("kyhieu",), ("serial",)],
    "inv_no": [("so hoa don",), ("so hd",), ("so", "hoa don"), ("invoice", "no"), ("sohoadon",)],
    "inv_date": [("ngay hoa don",), ("ngay lap",), ("ngay", "hoa don"), ("ngay",)],
    "buyer_tax_code": [("ma so thue",), ("mst",), ("tax", "code")],
    "buyer_name": [("ten don vi",), ("ten nguoi mua",), ("ten khach",), ("khach hang",), ("nguoi mua",), ("ten",)],
    "amount_before_vat": [("tien chua thue",), ("chua thue",), ("truoc thue",), ("thanh tien",), ("tien hang",)],
    "vat_amount": [("tien thue",), ("thue gtgt",), ("thue vat",), ("tien", "thue")],
    "total_amount": [("tong tien thanh toan",), ("tong thanh toan",), ("tong tien",), ("thanh toan",)],
    "transaction_id": [("ma tra cuu",), ("tra cuu",), ("lookup",)],
    "invoice_code": [("ma co quan thue",), ("ma cqt",), ("ma cua co quan thue",), ("invoice", "code")],
    "einvoice_status": [("trang thai hoa don",), ("trang thai",), ("status",)],
}

# Cột bắt buộc — thiếu là không nạp được.
REQUIRED = ("inv_series", "inv_no")

# Điểm cộng cho field có dấu hiệu ĐẶC TRƯNG, không thể nhầm.
# Không có nó thì "MST người mua" bị nuốt bởi luật tên khách ("nguoi mua" dài hơn
# "mst"), và cả cột mã số thuế lẫn cột tên đều trỏ sai — nạp vào là hỏng đối soát.
FIELD_BONUS = {
    "buyer_tax_code": 100,
    "transaction_id": 60,
    "invoice_code": 60,
    "inv_series": 40,
    "inv_no": 30,
}


def _match_column(text) -> str | None:
    """Tên cột → field, theo bộ từ khóa. Bộ khớp DÀI hơn thắng."""
    t = strip_tones(text)
    if not t:
        return None
    best, best_len = None, -1
    for field, rule_sets in COLUMN_RULES.items():
        for words in rule_sets:
            if all(w in t for w in words):
                score = sum(len(w) for w in words) + FIELD_BONUS.get(field, 0)
                if score > best_len:
                    best, best_len = field, score
    return best


def detect_header(rows):
    """Tìm dòng header + bản đồ {field: chỉ số cột}. Chọn dòng khớp NHIỀU cột nhất.

    File MISA xuất ra hay có vài dòng tiêu đề/logo phía trên, nên không thể giả
    định header nằm ở dòng đầu.
    """
    best_idx, best_map = None, {}
    for idx, row in enumerate(rows[:25]):
        if not row:
            continue
        mapping = {}
        for ci, cell in enumerate(row):
            field = _match_column(cell)
            if field and field not in mapping:
                mapping[field] = ci
        if len(mapping) > len(best_map):
            best_idx, best_map = idx, mapping
    return best_idx, best_map


# ═══════════════════════════════════════════════════════════════════════════
# Đọc giá trị ô
# ═══════════════════════════════════════════════════════════════════════════

_NUM_RE = re.compile(r"[^0-9,.\-]")


def to_number(v) -> float:
    """'1.234.567,89' · '1,234,567.89' · 1234567 → float. Không đoán bừa."""
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = _NUM_RE.sub("", str(v)).strip()
    if not s:
        return 0.0
    neg = s.startswith("-")
    s = s.lstrip("-")
    if "," in s and "." in s:
        # Dấu nào đứng SAU cùng là dấu thập phân.
        dec = "," if s.rfind(",") > s.rfind(".") else "."
        thou = "." if dec == "," else ","
        s = s.replace(thou, "").replace(dec, ".")
    elif "," in s:
        part = s.split(",")[-1]
        # 3 chữ số sau dấu phẩy → phân cách nghìn kiểu Việt, không phải thập phân.
        s = s.replace(",", "") if len(part) == 3 else s.replace(",", ".")
    elif "." in s:
        part = s.split(".")[-1]
        s = s.replace(".", "") if len(part) == 3 else s
    try:
        return -float(s) if neg else float(s)
    except ValueError:
        return 0.0


def to_date(v):
    """Trả 'YYYY-MM-DD' hoặc None. KHÔNG đoán khi mơ hồ."""
    if v is None or v == "":
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.match(r"^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), m.group(3)
        # Định dạng Việt là ngày/tháng/năm. Nếu vế đầu > 12 thì chắc chắn là ngày.
        if d > 12 and mo <= 12:
            return f"{y}-{mo:02d}-{d:02d}"
        return f"{y}-{mo:02d}-{d:02d}"
    return None


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _read_rows(content: str):
    """base64 (.xlsx hoặc .csv) → list các dòng."""
    raw = base64.b64decode((content or "").split(",")[-1])
    if not raw:
        frappe.throw(_("File rỗng"))

    # .xlsx bắt đầu bằng chữ ký ZIP.
    if raw[:2] == b"PK":
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            ws.reset_dimensions()  # nhiều file khai dimension sai → read_only dừng sớm
            return [list(r) for r in ws.iter_rows(values_only=True)][:MAX_ROWS]
        except Exception:
            frappe.throw(_("Không đọc được file Excel. Hãy xuất lại đúng định dạng .xlsx"))

    import csv

    for enc in ("utf-8-sig", "utf-8", "cp1258", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        frappe.throw(_("Không đọc được file. Hãy xuất định dạng .xlsx hoặc .csv UTF-8"))

    sample = text[:4000]
    delim = ";" if sample.count(";") > sample.count(",") else ","
    return [row for row in csv.reader(io.StringIO(text), delimiter=delim)][:MAX_ROWS]


# ═══════════════════════════════════════════════════════════════════════════
# Xem trước
# ═══════════════════════════════════════════════════════════════════════════

def _extract(rows, header_idx, mapping):
    """Dòng dữ liệu → dict đã chuẩn hóa. Bỏ dòng không có số hóa đơn."""
    out = []
    for row in rows[(header_idx or 0) + 1:]:
        if not row:
            continue
        inv_no = str(_cell(row, mapping.get("inv_no")) or "").strip()
        series = str(_cell(row, mapping.get("inv_series")) or "").strip()
        if not inv_no or not series or inv_no.startswith("<"):
            continue
        if not re.search(r"\d", inv_no):
            continue  # dòng tổng cộng / chú thích lọt xuống
        out.append({
            "inv_series": series,
            "inv_no": inv_no,
            "inv_date": to_date(_cell(row, mapping.get("inv_date"))),
            "buyer_tax_code": str(_cell(row, mapping.get("buyer_tax_code")) or "").replace(" ", "").strip(),
            "buyer_name": str(_cell(row, mapping.get("buyer_name")) or "").strip(),
            "amount_before_vat": to_number(_cell(row, mapping.get("amount_before_vat"))),
            "vat_amount": to_number(_cell(row, mapping.get("vat_amount"))),
            "total_amount": to_number(_cell(row, mapping.get("total_amount"))),
            "transaction_id": str(_cell(row, mapping.get("transaction_id")) or "").strip(),
            "invoice_code": str(_cell(row, mapping.get("invoice_code")) or "").strip(),
            "einvoice_status": str(_cell(row, mapping.get("einvoice_status")) or "").strip(),
        })
    return out


@frappe.whitelist()
def preview(content):
    """Đọc file, dò cột, trả bản đồ cột + vài dòng đầu để người dùng DUYỆT.

    Luôn phải qua bước này trước khi ghi — tên cột MISA chưa xác minh, không được
    im lặng nạp sai cột.
    """
    guard_sales_any()
    rows = _read_rows(content)
    if not rows:
        frappe.throw(_("File không có dòng nào"))

    header_idx, mapping = detect_header(rows)
    missing = [f for f in REQUIRED if f not in mapping]
    header_row = [str(c or "") for c in (rows[header_idx] if header_idx is not None else [])]
    data = _extract(rows, header_idx, mapping) if not missing else []

    seen, dups = set(), 0
    for r in data:
        key = (r["inv_series"].upper(), r["inv_no"].lstrip("0") or r["inv_no"])
        if key in seen:
            dups += 1
        seen.add(key)

    return {
        "header_row": header_idx,
        "header_texts": header_row,
        "mapping": {k: header_row[v] if v < len(header_row) else f"cột {v}" for k, v in mapping.items()},
        "missing": missing,
        "total_rows": len(data),
        "duplicates": dups,
        "no_date": sum(1 for r in data if not r["inv_date"]),
        "no_amount": sum(1 for r in data if not r["total_amount"]),
        "sample": data[:MAX_PREVIEW],
    }


# ═══════════════════════════════════════════════════════════════════════════
# Nạp thật
# ═══════════════════════════════════════════════════════════════════════════

@frappe.whitelist()
def commit(content, run_match=1):
    """Nạp file vào MISA Invoice Snapshot rồi khớp với Sales Invoice.

    Upsert theo khóa tự nhiên (ký hiệu, số chuẩn hóa) — chạy lại cùng file KHÔNG
    nhân bản. KHÔNG đè `sales_invoice` / `match_status` / `transaction_id` đã có:
    kết quả đối soát do người chốt không được lần nạp sau ghi đè.
    """
    guard_manager()
    from ketoan.misa_integration.doctype.misa_invoice_snapshot.misa_invoice_snapshot import (
        norm_inv_no, norm_series,
    )

    rows = _read_rows(content)
    header_idx, mapping = detect_header(rows)
    missing = [f for f in REQUIRED if f not in mapping]
    if missing:
        frappe.throw(_("Không tìm thấy cột bắt buộc: {0}. Kiểm tra lại file xuất từ MISA.")
                     .format(", ".join(missing)))

    data = _extract(rows, header_idx, mapping)
    if not data:
        frappe.throw(_("File không có dòng hóa đơn nào đọc được"))

    dates = [r["inv_date"] for r in data if r["inv_date"]]
    run = frappe.get_doc({
        "doctype": "MISA Sync Run", "job_type": "pull_statement", "trigger_type": "Manual",
        "from_date": min(dates) if dates else None, "to_date": max(dates) if dates else None,
    }).insert(ignore_permissions=True)
    frappe.db.commit()

    stat = {"fetched": 0, "created": 0, "updated": 0}
    errors = []

    for i, r in enumerate(data):
        try:
            series = norm_series(r["inv_series"])
            no_norm = norm_inv_no(r["inv_no"])
            values = {
                "inv_series": series, "inv_no": r["inv_no"], "inv_no_norm": no_norm,
                "inv_date": r["inv_date"], "buyer_tax_code": r["buyer_tax_code"],
                "buyer_name": r["buyer_name"],
                "amount_before_vat": flt(r["amount_before_vat"]),
                "vat_amount": flt(r["vat_amount"]),
                "total_amount": flt(r["total_amount"]),
                "invoice_code": r["invoice_code"],
                "einvoice_status": r["einvoice_status"],
                "source_api": "statement", "last_synced_at": now_datetime(),
            }
            existing = frappe.db.get_value(
                "MISA Invoice Snapshot", {"inv_series": series, "inv_no_norm": no_norm},
                ["name", "transaction_id"], as_dict=True)
            if existing:
                if r["transaction_id"] and not existing.transaction_id:
                    values["transaction_id"] = r["transaction_id"]
                frappe.db.set_value("MISA Invoice Snapshot", existing.name, values, update_modified=False)
                stat["updated"] += 1
            else:
                values["transaction_id"] = r["transaction_id"]
                values["origin"] = "Chưa xác định"
                values["match_status"] = "Chưa xác định"
                doc = frappe.get_doc(dict(doctype="MISA Invoice Snapshot", **values))
                doc.flags.ignore_permissions = True
                doc.insert()
                stat["created"] += 1
            stat["fetched"] += 1
            if (i + 1) % 100 == 0:
                frappe.db.commit()
        except Exception:
            frappe.log_error(frappe.get_traceback(), "misa_import.commit")
            errors.append(f"{r['inv_series']} {r['inv_no']}: lỗi ghi")
    frappe.db.commit()

    matched = None
    if int(run_match or 0) and dates:
        try:
            from ketoan.api.misa_reconcile import match_snapshots

            matched = match_snapshots(min(dates), max(dates))
        except Exception as e:
            frappe.log_error(frappe.get_traceback(), "misa_import.commit/match")
            errors.append(f"khớp thất bại: {str(e)[:200]}")

    for k, v in stat.items():
        setattr(run, k, v)
    if matched:
        run.matched = matched.get("matched") or 0
        run.mismatched = matched.get("mismatched") or 0
    if errors:
        run.error_log = "\n".join(errors[-200:])
    run.status = "Lỗi" if (errors and not stat["fetched"]) else ("Thành công một phần" if errors else "Thành công")
    run.finished_at = now_datetime()
    run.save(ignore_permissions=True)
    frappe.db.commit()

    return {"run": run.name, "imported": stat, "matched": matched, "errors": errors[:20]}
