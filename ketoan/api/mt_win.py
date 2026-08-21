# -*- coding: utf-8 -*-
"""mt_win — HỒ SƠ THANH TOÁN nộp cho WinCommerce (§2.2 SOP).

Win chỉ xử lý thanh toán khi nhận đủ bảng kê + file PDF hóa đơn ĐẶT ĐÚNG TÊN.
Sai tên file là hồ sơ bị trả về và cả đợt thanh toán trượt kỳ — nên tên file
được SINH RA, không để gõ tay:

    20260817_2007766_01_PF
    └─YYYYMMDD─┘ └NCC─┘ └NN┘└PF

Ba việc module này làm:

  1. Gom hóa đơn WinCommerce CHƯA nộp hồ sơ nào -> đề xuất danh sách.
  2. Lập hồ sơ, sinh tên file PDF chuẩn.
  3. Xuất file Excel ĐÚNG khuôn mẫu Win gửi, để nộp thẳng.

MỘT HÓA ĐƠN CHỈ NỘP MỘT LẦN. Cả tầng đề xuất lẫn `validate()` của DocType đều
loại hóa đơn đã nằm trong hồ sơ khác — nộp trùng là Win trả hồ sơ.
"""

import hashlib
import io

import frappe
from frappe import _
from frappe.utils import cint, cstr, flt, getdate, today

from ketoan.api._guard import guard_manager, guard_mt, is_chief
from ketoan.mt.doctype.mt_win_dossier.mt_win_dossier import (
    STATUS_DRAFT,
    STATUS_SUBMITTED,
    build_prefix,
)
from ketoan.misa_integration.doctype.misa_invoice_snapshot.misa_invoice_snapshot import norm_text

SOURCE_DT = "MT Win Dossier"
WIN_CHAIN = "WinCommerce"

# Nhãn cột của mẫu Win, ĐÚNG THỨ TỰ và ĐÚNG CHỮ đọc từ file thật
# `Mẫu bảng kê ghi nhận hồ sơ thanh toán Winmart.xlsx` (header ở dòng 2).
# Đổi một chữ ở đây là Win không nhận file.
WIN_COLUMNS = (
    "STT", "Code", "PO VCM", "Ký hiệu HĐ", "Số hóa đơn", "Ngày hóa đơn",
    "Số Tiền trước VAT", "VAT", "Tổng tiền thanh toán", "Tên File PDF",
)

# Trần số hóa đơn trong một hồ sơ. Mẫu thật 19 dòng; 500 là dư rất xa mà vẫn
# chặn được ca chọn nhầm cả năm hóa đơn vào một hồ sơ.
MAX_LINES = 500


def _require_tables():
    for dt in (SOURCE_DT, "MT Win Dossier Line"):
        if not frappe.db.table_exists(dt):
            frappe.throw(_(
                "Chức năng hồ sơ WinCommerce chưa được cài trên site này (thiếu bảng {0}). "
                "Quản trị chạy: bench --site TÊN_SITE migrate"
            ).format(dt))


def _company(company=None):
    from ketoan.api.mt import _company as _mt_company
    return _mt_company(company)


def _po_column():
    """Cột số PO của WinCommerce trên Sales Invoice.

    `custom_po_` là field CỦA SITE (không do app này tạo) — `misa_push` đã đọc
    nó để ghép vào tên người mua khi đẩy hóa đơn Winmart. Site nào chưa có thì
    hồ sơ vẫn lập được, chỉ trống cột PO và có cảnh báo.
    """
    return "custom_po_" if frappe.db.has_column("Sales Invoice", "custom_po_") else None


def _win_customers(company):
    """Khách hàng thuộc chuỗi WinCommerce.

    Dùng CHUNG `mt.chain_customers` với mọi màn hình khác. Bản cũ chỉ đọc field
    khai `Customer.custom_mt_chain`, nên khách đã có bảng kê WinCommerce mà kế
    toán chưa kịp khai field thì hiện là "WinCommerce" ở màn công nợ nhưng BIẾN
    MẤT khỏi danh sách gom hồ sơ Win — và cái mất chính là cái làm hồ sơ nộp tiền.
    """
    from ketoan.api.mt import chain_customers

    return chain_customers(WIN_CHAIN)


def _candidates(company, from_date, to_date, customer=None):
    """Hóa đơn WinCommerce ĐÃ GHI SỔ, có số hóa đơn, CHƯA nằm trong hồ sơ nào."""
    customers = [customer] if customer else _win_customers(company)
    if not customers:
        return [], ("Chưa có khách hàng nào được gán chuỗi WinCommerce. Vào màn "
                    "'Sửa chuỗi của khách' gán trước.")
    if not frappe.db.has_column("Sales Invoice", "custom_misa_inv_no"):
        return [], ("Sales Invoice chưa có field số hóa đơn MISA — chạy "
                    "bench migrate rồi làm lại.")

    po = _po_column()
    rows = frappe.db.sql("""
        SELECT si.name, si.customer, si.posting_date,
               si.custom_misa_inv_series AS inv_series,
               si.custom_misa_inv_no     AS inv_no,
               si.custom_misa_inv_date   AS inv_date,
               si.base_net_total, si.grand_total,
               {po} AS po_vcm
        FROM `tabSales Invoice` si
        WHERE si.company = %(company)s AND si.docstatus = 1
          AND si.customer IN %(customers)s
          AND si.posting_date BETWEEN %(fd)s AND %(td)s
          AND IFNULL(si.custom_misa_inv_no, '') != ''
          AND NOT EXISTS (
              SELECT 1 FROM `tabMT Win Dossier Line` l
              WHERE l.parenttype = 'MT Win Dossier' AND l.sales_invoice = si.name)
        ORDER BY si.custom_misa_inv_no
    """.format(po=("si." + po) if po else "NULL"),
        {"company": company, "customers": tuple(customers),
         "fd": from_date, "td": to_date}, as_dict=True)
    return rows, None


def _build_plan(company, from_date, to_date, customer, submit_date, dossier_no, vendor_code):
    rows, blocker = _candidates(company, from_date, to_date, customer)
    warnings = []
    if blocker:
        warnings.append(blocker)
    if len(rows) > MAX_LINES:
        frappe.throw(_(
            "Chọn được {0} hóa đơn — vượt trần {1}. Thu hẹp khoảng ngày; một hồ sơ Win "
            "thật chỉ vài chục hóa đơn."
        ).format(len(rows), MAX_LINES))

    prefix = build_prefix(submit_date, vendor_code, dossier_no)
    if not _po_column():
        warnings.append("Site chưa có field `custom_po_` trên Sales Invoice — cột "
                        "'PO VCM' sẽ TRỐNG. Win có thể trả hồ sơ vì thiếu số PO.")

    lines, no_po = [], 0
    for i, r in enumerate(rows, start=1):
        # Tiền TRƯỚC VAT lấy `base_net_total` — số ròng của hóa đơn. Suy ngược
        # từ grand_total bằng cách chia thuế suất là bịa số khi hóa đơn có nhiều
        # thuế suất hoặc có phí vận chuyển.
        base = flt(r.base_net_total)
        vat = flt(r.grand_total) - base
        if not norm_text(r.po_vcm):
            no_po += 1
        lines.append({
            "stt": i,
            "po_vcm": norm_text(r.po_vcm),
            "inv_series": norm_text(r.inv_series),
            "inv_no": norm_text(r.inv_no),
            "inv_date": cstr(r.inv_date or r.posting_date),
            "amount_before_vat": round(base, 2),
            "vat_amount": round(vat, 2),
            "total_amount": round(flt(r.grand_total), 2),
            "pdf_name": prefix,
            "sales_invoice": r.name,
            "customer": r.customer,
        })
    if no_po:
        warnings.append("%d/%d hóa đơn KHÔNG có số PO VCM. Win đối chiếu theo PO — "
                        "điền `custom_po_` trên hóa đơn trước khi nộp." % (no_po, len(lines)))

    return {
        "file_prefix": prefix,
        "vendor_code": norm_text(vendor_code),
        "submit_date": cstr(submit_date),
        "dossier_no": cint(dossier_no) or 1,
        "customer": customer or (lines[0]["customer"] if lines else None),
        "period_from": cstr(from_date), "period_to": cstr(to_date),
        "lines": lines,
        "total_before_vat": round(sum(l["amount_before_vat"] for l in lines), 2),
        "total_vat": round(sum(l["vat_amount"] for l in lines), 2),
        "total_amount": round(sum(l["total_amount"] for l in lines), 2),
        "warnings": warnings,
    }


def _plan_hash(plan):
    h = hashlib.sha1()
    h.update("W|{}|{}\n".format(plan["file_prefix"], plan["customer"] or "").encode())
    for l in plan["lines"]:
        h.update("L|{}|{}|{:.2f}\n".format(
            l["sales_invoice"], l["inv_no"], flt(l["total_amount"])).encode())
    return h.hexdigest()


# ═══════════════════════════════════════════════════════════════════════════
# Whitelisted
# ═══════════════════════════════════════════════════════════════════════════

@frappe.whitelist()
def preview_dossier(from_date=None, to_date=None, customer=None, submit_date=None,
                    dossier_no=1, vendor_code=None, company=None):
    """XEM TRƯỚC hồ sơ: hóa đơn nào vào, tên file PDF là gì. KHÔNG ghi gì."""
    guard_mt()
    _require_tables()
    company = _company(company)

    submit_date = cstr(submit_date or today())
    vendor_code = norm_text(vendor_code) or _last_vendor_code(company)
    if not vendor_code:
        frappe.throw(_(
            "Chưa biết mã NCC của mình tại WinCommerce. Điền vào ô 'Mã NCC' (mẫu thật là "
            "2007766) — tên file PDF cần mã này, sai tên là Win trả hồ sơ."
        ))
    plan = _build_plan(company, from_date or today(), to_date or today(),
                       customer, submit_date, dossier_no, vendor_code)
    return {
        "plan_hash": _plan_hash(plan),
        "n_lines": len(plan["lines"]),
        "sample": plan["lines"][:60],
        **{k: v for k, v in plan.items() if k != "lines"},
        "columns": list(WIN_COLUMNS),
        "can_commit": bool(plan["lines"]) and is_chief(),
        "note": _(
            "File PDF hóa đơn nộp kèm PHẢI đặt tên đúng '{0}'. Sai tên là Win trả hồ sơ "
            "và cả đợt trượt kỳ thanh toán."
        ).format(plan["file_prefix"]),
    }


def _last_vendor_code(company):
    """Mã NCC dùng lần gần nhất — đỡ phải gõ lại mỗi kỳ."""
    v = frappe.db.get_value(SOURCE_DT, {"company": company}, "vendor_code",
                            order_by="creation desc")
    if v:
        return norm_text(v)
    # Chưa có hồ sơ nào -> thử lấy từ master điểm siêu thị của chuỗi Win.
    return norm_text(frappe.db.get_value("MT Store", {"chain": WIN_CHAIN}, "vendor_code")
                     if frappe.db.table_exists("MT Store") else "")


@frappe.whitelist()
def commit_dossier(from_date=None, to_date=None, customer=None, submit_date=None,
                   dossier_no=1, vendor_code=None, expected_hash=None, company=None):
    """Lập hồ sơ (trạng thái Nháp). Đòi vân tay của bản xem trước."""
    guard_manager()
    _require_tables()
    company = _company(company)

    submit_date = cstr(submit_date or today())
    vendor_code = norm_text(vendor_code) or _last_vendor_code(company)
    plan = _build_plan(company, from_date or today(), to_date or today(),
                       customer, submit_date, dossier_no, vendor_code)
    if not plan["lines"]:
        frappe.throw(_("Không có hóa đơn nào để nộp trong khoảng này"))
    if not expected_hash:
        frappe.throw(_("Phải xem trước rồi mới lập hồ sơ được"))
    if _plan_hash(plan) != expected_hash:
        frappe.throw(_(
            "Dữ liệu đã đổi kể từ lúc xem trước (có hóa đơn mới ghi sổ, hoặc một hóa đơn "
            "vừa được nộp ở hồ sơ khác). Xem lại rồi lập — không ghi gì cả."
        ))

    doc = frappe.new_doc(SOURCE_DT)
    doc.company = company
    doc.customer = plan["customer"]
    doc.vendor_code = plan["vendor_code"]
    doc.submit_date = getdate(plan["submit_date"])
    doc.dossier_no = plan["dossier_no"]
    doc.period_from = getdate(plan["period_from"])
    doc.period_to = getdate(plan["period_to"])
    doc.status = STATUS_DRAFT
    for l in plan["lines"]:
        doc.append("lines", {k: l[k] for k in (
            "stt", "po_vcm", "inv_series", "inv_no", "inv_date", "amount_before_vat",
            "vat_amount", "total_amount", "pdf_name", "sales_invoice")})
    doc.insert()
    frappe.db.commit()
    return {
        "name": doc.name, "file_prefix": doc.file_prefix,
        "n_lines": len(doc.lines), "total_amount": flt(doc.total_amount),
        "warnings": plan["warnings"],
        "message": _("Đã lập hồ sơ {0} với {1} hóa đơn. Xuất Excel rồi nộp cho Win.")
                   .format(doc.file_prefix, len(doc.lines)),
    }


@frappe.whitelist()
def list_dossiers(from_date=None, to_date=None, status=None, search=None,
                  page=1, page_size=20, company=None):
    """Danh sách hồ sơ đã lập."""
    guard_mt()
    _require_tables()
    company = _company(company)

    page = max(1, cint(page) or 1)
    page_size = min(100, max(1, cint(page_size) or 20))
    where = ["d.company = %(company)s"]
    params = {"company": company}
    if from_date and to_date:
        where.append("d.submit_date BETWEEN %(fd)s AND %(td)s")
        params["fd"], params["td"] = from_date, to_date
    if status:
        where.append("d.status = %(status)s")
        params["status"] = status
    if search:
        where.append("(d.name LIKE %(q)s OR IFNULL(d.file_prefix,'') LIKE %(q)s "
                     "OR IFNULL(d.customer,'') LIKE %(q)s)")
        params["q"] = "%" + cstr(search).strip() + "%"
    clause = " AND ".join(where)

    head = frappe.db.sql("SELECT COUNT(*) AS n, IFNULL(SUM(d.total_amount), 0) AS amount "
                         "FROM `tabMT Win Dossier` d WHERE " + clause, params, as_dict=True)[0]
    params["limit"] = page_size
    params["offset"] = (page - 1) * page_size
    rows = frappe.db.sql("""
        SELECT d.name, d.file_prefix, d.submit_date, d.dossier_no, d.vendor_code,
               d.customer, cus.customer_name, d.status, d.period_from, d.period_to,
               d.total_before_vat, d.total_vat, d.total_amount,
               (SELECT COUNT(*) FROM `tabMT Win Dossier Line` l
                 WHERE l.parent = d.name AND l.parenttype = 'MT Win Dossier') AS n_lines
        FROM `tabMT Win Dossier` d
        LEFT JOIN `tabCustomer` cus ON cus.name = d.customer
        WHERE {clause}
        ORDER BY d.submit_date DESC, d.creation DESC
        LIMIT %(limit)s OFFSET %(offset)s
    """.format(clause=clause), params, as_dict=True)

    total = cint(head.n)
    return {
        "rows": rows, "total": total, "total_amount": flt(head.amount),
        "page": page, "page_size": page_size,
        "pages": (total + page_size - 1) // page_size if total else 1,
        "statuses": [STATUS_DRAFT, STATUS_SUBMITTED],
        "can_manage": is_chief(),
        "default_vendor_code": _last_vendor_code(company),
    }


@frappe.whitelist()
def get_dossier(name, company=None):
    guard_mt()
    _require_tables()
    company = _company(company)
    doc = frappe.get_doc(SOURCE_DT, name)
    if cstr(doc.company) != cstr(company):
        frappe.throw(_("Hồ sơ {0} thuộc công ty khác").format(name))
    return {"doc": doc.as_dict(), "lines": [l.as_dict() for l in doc.lines or []],
            "columns": list(WIN_COLUMNS), "can_manage": is_chief()}


@frappe.whitelist()
def mark_submitted(name, company=None):
    """Đánh dấu ĐÃ NỘP. Sau đó hồ sơ không xóa được nữa."""
    guard_manager()
    _require_tables()
    company = _company(company)
    doc = frappe.get_doc(SOURCE_DT, name)
    if cstr(doc.company) != cstr(company):
        frappe.throw(_("Hồ sơ {0} thuộc công ty khác").format(name))
    doc.status = STATUS_SUBMITTED
    doc.save()
    frappe.db.commit()
    return {"name": doc.name, "status": doc.status,
            "message": _("Đã đánh dấu hồ sơ {0} là ĐÃ NỘP.").format(doc.file_prefix)}


@frappe.whitelist()
def export_dossier(name, company=None):
    """Xuất bảng kê hồ sơ ra .xlsx ĐÚNG KHUÔN mẫu Win gửi (tải về).

    Đúng 10 cột, đúng chữ, header ở DÒNG 2 — y như file mẫu. Đổi một chữ là Win
    không nhận file, nên khuôn nằm ở hằng số `WIN_COLUMNS` chứ không rải trong mã.
    """
    guard_mt()
    _require_tables()
    company = _company(company)

    doc = frappe.get_doc(SOURCE_DT, name)
    if cstr(doc.company) != cstr(company):
        frappe.throw(_("Hồ sơ {0} thuộc công ty khác").format(name))

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, Side
    except ImportError:
        frappe.throw(_("Site thiếu thư viện openpyxl — không xuất Excel được"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    thin = Side(style="thin")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header ở DÒNG 2, đúng như mẫu Win (dòng 1 để trống).
    for c, label in enumerate(WIN_COLUMNS, start=1):
        cell = ws.cell(row=2, column=c, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = box

    for i, l in enumerate(doc.lines or [], start=3):
        vals = (cint(l.stt), doc.vendor_code, l.po_vcm or "", l.inv_series or "",
                l.inv_no or "", getdate(l.inv_date) if l.inv_date else None,
                flt(l.amount_before_vat), flt(l.vat_amount), flt(l.total_amount),
                l.pdf_name or doc.file_prefix)
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = box
            if c == 6 and v is not None:
                cell.number_format = "DD/MM/YYYY"
            elif c in (7, 8, 9):
                cell.number_format = "#,##0"

    for c, w in enumerate((6, 12, 16, 12, 13, 13, 18, 14, 20, 26), start=1):
        ws.column_dimensions[ws.cell(row=2, column=c).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    frappe.local.response.filename = "%s.xlsx" % doc.file_prefix
    frappe.local.response.filecontent = buf.getvalue()
    frappe.local.response.type = "download"
