"""Whitelisted methods — Nhập sổ quỹ từ file sao kê ngân hàng (Excel).

- parse_statement: đọc file .xlsx (cột: Số tham chiếu / Ngày / Ghi nợ / Ghi có /
  Số dư / Nội dung), trả danh sách giao dịch + cờ trùng (đã nhập trước đó).
- import_transactions: tạo Journal Entry (voucher_type "Journal Entry") cho các
  dòng đã chọn và GHI SỔ (submit) luôn — Ghi có (tiền vào) → Nợ TK ngân hàng /
  Có TK đối ứng; Ghi nợ (tiền ra) → ngược lại. Ghi chú người dùng tự nhập đưa
  vào field remark, kèm marker chống trùng [BANKIMP-<key>].

TK ngân hàng (112) do người dùng chọn. Guard quỹ.
"""

import base64
import hashlib
import io
import json
import re
from datetime import datetime

import frappe
from frappe import _
from frappe.utils import flt

from ketoan.api._guard import guard_cash, resolve_company
from ketoan.utils import je_remark_field

_MARK_RE = re.compile(r"\[BANKIMP-([0-9a-f]{16})\]")


def _load_rules(company: str) -> list:
    """Quy tắc map đang bật cho company (hoặc dùng chung), ưu tiên cao + từ khóa dài trước."""
    rules = frappe.get_all(
        "Ketoan Bank Map Rule",
        filters={"enabled": 1},
        or_filters=[["company", "=", company], ["company", "is", "not set"]],
        fields=["name", "keyword", "counter_account", "direction", "party_type", "party", "priority"],
    )
    rules = [r for r in rules if (r.keyword or "").strip()]
    rules.sort(key=lambda r: (-(r.priority or 0), -len(r.keyword or "")))
    return rules


def _match_rule(content: str, direction: str, rules: list):
    """Trả quy tắc khớp đầu tiên (theo từ khóa + chiều tiền) hoặc None."""
    low = (content or "").lower()
    for r in rules:
        d = r.direction or "Bất kỳ"
        if d == "Tiền vào" and direction != "in":
            continue
        if d == "Tiền ra" and direction != "out":
            continue
        if (r.keyword or "").lower() in low:
            return r
    return None


def _num(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace(" ", "").strip()
    try:
        return float(s or 0)
    except ValueError:
        return 0.0


def _parse_date(v):
    """Trả (iso_date, raw_str). Hỗ trợ datetime hoặc 'dd/mm/yyyy[ HH:MM:SS]'."""
    if isinstance(v, datetime):
        return v.date().isoformat(), v.isoformat(sep=" ")
    s = str(v or "").strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat(), s
        except ValueError:
            continue
    return None, s


def _row_key(ref, date_iso, debit, credit, content) -> str:
    raw = f"{ref}|{date_iso}|{debit}|{credit}|{(content or '')[:60]}"
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return ""
    return row[idx]


def _detect_columns(all_rows):
    """Dò dòng header + chỉ số cột theo tên. Fallback layout chuẩn (0..5)."""
    default = {"ref": 0, "date": 1, "debit": 2, "credit": 3, "content": 5}
    for idx, row in enumerate(all_rows[:15]):
        if not row:
            continue
        texts = [str(c or "").strip().lower() for c in row]
        joined = " | ".join(texts)
        if not (("ghi" in joined and ("nợ" in joined or "no" in joined)) or "tham chiếu" in joined or "nội dung" in joined):
            continue
        cols = dict(default)
        found = False
        for ci, t in enumerate(texts):
            if "tham chiếu" in t:
                cols["ref"] = ci; found = True
            elif "ngày" in t or "ngay" in t:
                cols["date"] = ci; found = True
            elif "ghi nợ" in t or t in ("nợ", "no", "debit", "ghi no"):
                cols["debit"] = ci; found = True
            elif "ghi có" in t or t in ("có", "co", "credit", "ghi co"):
                cols["credit"] = ci; found = True
            elif "nội dung" in t or "diễn giải" in t or "mô tả" in t or "noi dung" in t:
                cols["content"] = ci; found = True
        if found:
            return cols, idx
    return default, 0


def _suggest_bank_account(company: str) -> str | None:
    """Đoán TK ngân hàng 112 (account_number/name bắt đầu 112)."""
    rows = frappe.get_all(
        "Account",
        filters={"company": company, "account_type": "Bank", "is_group": 0, "disabled": 0},
        fields=["name", "account_number"],
        order_by="account_number, name",
    )
    for a in rows:
        if (a.account_number or "").startswith("112") or a.name.startswith("112"):
            return a.name
    return rows[0].name if rows else None


@frappe.whitelist()
def get_import_options(company: str | None = None) -> dict:
    """TK ngân hàng (Bank) + TK đối ứng + TK ngân hàng đề xuất (112).

    counter_accounts kèm account_number và XẾP THEO TẦN SUẤT DÙNG (số dòng
    Journal Entry Account 12 tháng gần nhất) — client hiện top trước, gõ để tìm.
    """
    guard_cash()
    company = resolve_company(company)
    bank_accounts = frappe.get_all(
        "Account",
        filters={"company": company, "account_type": ["in", ["Bank", "Cash"]], "is_group": 0, "disabled": 0},
        fields=["name", "account_name", "account_number", "account_type"],
        order_by="account_type, name",
    )
    counter_accounts = frappe.db.sql(
        """
        SELECT a.name, a.account_name, a.account_number, a.root_type, a.account_type,
               COALESCE(u.cnt, 0) AS usage_count
        FROM `tabAccount` a
        LEFT JOIN (
            SELECT account, COUNT(*) AS cnt
            FROM `tabJournal Entry Account`
            WHERE docstatus < 2 AND creation >= DATE_SUB(NOW(), INTERVAL 365 DAY)
            GROUP BY account
        ) u ON u.account = a.name
        WHERE a.company = %(company)s AND a.is_group = 0 AND a.disabled = 0
        ORDER BY COALESCE(u.cnt, 0) DESC, a.account_number ASC, a.name ASC
        LIMIT 1000
        """,
        {"company": company},
        as_dict=True,
    )
    return {
        "company": company,
        "bank_accounts": bank_accounts,
        "counter_accounts": counter_accounts,
        "suggested_bank": _suggest_bank_account(company),
    }


@frappe.whitelist()
def search_party(party_type: str, txt: str = "", limit: int = 15) -> list:
    """Tìm đối tượng cho TK phải thu/phải trả (giống chọn Party khi bút toán trên
    Desk): Customer/Supplier đang hoạt động, hay dùng (JE 12 tháng) lên trước."""
    guard_cash()
    if party_type not in ("Customer", "Supplier"):
        frappe.throw(_("Loại đối tượng không hợp lệ"))
    label = "customer_name" if party_type == "Customer" else "supplier_name"
    limit = max(1, min(int(limit or 15), 50))
    like = "%" + (txt or "").strip() + "%"
    return frappe.db.sql(
        f"""
        SELECT p.name, p.`{label}` AS label, COALESCE(u.cnt, 0) AS usage_count
        FROM `tab{party_type}` p
        LEFT JOIN (
            SELECT party, COUNT(*) AS cnt
            FROM `tabJournal Entry Account`
            WHERE party_type = %(pt)s AND docstatus < 2
              AND creation >= DATE_SUB(NOW(), INTERVAL 365 DAY)
            GROUP BY party
        ) u ON u.party = p.name
        WHERE IFNULL(p.disabled, 0) = 0
          AND (p.name LIKE %(like)s OR p.`{label}` LIKE %(like)s)
        ORDER BY COALESCE(u.cnt, 0) DESC, p.`{label}` ASC
        LIMIT {limit}
        """,
        {"pt": party_type, "like": like},
        as_dict=True,
    )


@frappe.whitelist()
def parse_statement(content: str, company: str | None = None) -> dict:
    """Đọc file sao kê .xlsx (base64) → danh sách giao dịch + cờ trùng."""
    guard_cash()
    company = resolve_company(company)

    raw = base64.b64decode((content or "").split(",")[-1])
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        # Nhiều file ngân hàng khai báo dimension sai (chỉ A1) → read_only dừng sau 1 dòng.
        ws.reset_dimensions()
    except Exception:
        frappe.throw(_("Không đọc được file Excel. Hãy xuất đúng định dạng .xlsx"))

    all_rows = [r for r in ws.iter_rows(values_only=True)]

    # Dò dòng header + vị trí cột (chịu được file có dòng tiêu đề thừa / đổi thứ tự cột).
    cols, header_idx = _detect_columns(all_rows)

    txns = []
    for i in range(header_idx + 1, len(all_rows)):
        row = all_rows[i]
        if not row:
            continue
        ref = str(_cell(row, cols["ref"]) or "").strip()
        date_iso, date_raw = _parse_date(_cell(row, cols["date"]))
        debit = _num(_cell(row, cols["debit"]))   # Ghi nợ = tiền RA khỏi ngân hàng
        credit = _num(_cell(row, cols["credit"]))  # Ghi có = tiền VÀO ngân hàng
        content_txt = str(_cell(row, cols["content"]) or "").strip()
        if not date_iso or (debit == 0 and credit == 0):
            continue
        key = _row_key(ref, date_iso, debit, credit, content_txt)
        txns.append({
            "key": key,
            "ref": ref,
            "date": date_iso,
            "datetime": date_raw,
            "debit": debit,
            "credit": credit,
            "direction": "in" if credit > 0 else "out",
            "content": content_txt,
        })

    if not txns:
        return {"company": company, "transactions": [], "duplicates": 0}

    # Lọc trùng: dò marker [BANKIMP-key] trong JE đã có (theo khoảng ngày của file).
    dates = [t["date"] for t in txns]
    field = je_remark_field()
    existing = set()
    cands = frappe.get_all(
        "Journal Entry",
        filters={field: ["like", "%[BANKIMP-%"], "docstatus": ["<", 2],
                 "company": company, "posting_date": ["between", [min(dates), max(dates)]]},
        fields=["name", field],
        limit=5000,
    )
    for c in cands:
        for m in _MARK_RE.findall(c.get(field) or ""):
            existing.add(m)

    dup_count = 0
    for t in txns:
        t["duplicate"] = t["key"] in existing
        if t["duplicate"]:
            dup_count += 1

    # Gợi ý TK đối ứng theo quy tắc map đã lưu — bỏ rule trỏ TK không còn dùng
    # được với company này (đã disable / group / company khác).
    rules = _load_rules(company)
    _acc_ok = {}
    def _usable(acc):
        if acc not in _acc_ok:
            row = frappe.db.get_value("Account", acc, ["company", "is_group", "disabled"], as_dict=True)
            _acc_ok[acc] = bool(row and row.company == company and not row.is_group and not row.disabled)
        return _acc_ok[acc]
    rules = [r for r in rules if _usable(r.counter_account)]
    suggested = 0
    for t in txns:
        r = _match_rule(t["content"], t["direction"], rules)
        t["suggested_counter"] = r.counter_account if r else None
        t["suggested_party_type"] = (r.party_type or None) if r else None
        t["suggested_party"] = (r.party or None) if r else None
        t["suggested_rule"] = (r.keyword or None) if r else None
        if r and not t["duplicate"]:
            suggested += 1

    return {
        "company": company, "transactions": txns,
        "duplicates": dup_count, "total": len(txns), "suggested": suggested,
    }


@frappe.whitelist()
def import_transactions(rows, bank_account: str, company: str | None = None) -> dict:
    """Tạo Journal Entry cho các giao dịch đã chọn và SUBMIT (ghi sổ) luôn.

    rows: JSON list {key, date, debit, credit, content, counter_account,
    party?, remark?}. remark (ghi chú tự nhập) ghi vào field remark của JE;
    submit lỗi thì vẫn giữ bản nháp và báo lại lý do.
    """
    guard_cash()
    company = resolve_company(company)
    if isinstance(rows, str):
        rows = json.loads(rows)
    if not rows:
        frappe.throw(_("Chưa chọn giao dịch nào"))
    if not bank_account:
        frappe.throw(_("Chưa chọn tài khoản ngân hàng"))

    bank = frappe.db.get_value("Account", bank_account, ["company", "is_group", "disabled"], as_dict=True)
    if not bank or bank.company != company or bank.is_group or bank.disabled:
        frappe.throw(_("Tài khoản ngân hàng không hợp lệ"))
    if len(rows) > 500:
        frappe.throw(_("Tối đa 500 giao dịch mỗi lần nhập"))

    field = je_remark_field()
    # Khóa trùng hiện có trong khoảng ngày.
    dates = [r.get("date") for r in rows if r.get("date")]
    existing = set()
    if dates:
        for c in frappe.get_all(
            "Journal Entry",
            filters={field: ["like", "%[BANKIMP-%"], "docstatus": ["<", 2],
                     "company": company, "posting_date": ["between", [min(dates), max(dates)]]},
            fields=[field], limit=5000,
        ):
            existing.update(_MARK_RE.findall(c.get(field) or ""))

    created, skipped = [], []
    for r in rows:
        key = r.get("key") or _row_key(r.get("ref", ""), r.get("date"), r.get("debit"), r.get("credit"), r.get("content"))
        if key in existing:
            skipped.append({"key": key, "reason": "Đã nhập trước đó"})
            continue
        counter = r.get("counter_account")
        if not counter:
            skipped.append({"key": key, "reason": "Chưa chọn TK đối ứng"})
            continue
        debit = _num(r.get("debit"))
        credit = _num(r.get("credit"))
        amount = credit if credit > 0 else debit
        if amount <= 0:
            skipped.append({"key": key, "reason": "Số tiền = 0"})
            continue

        cacc = frappe.db.get_value("Account", counter, ["company", "is_group", "account_type"], as_dict=True)
        if not cacc or cacc.company != company or cacc.is_group:
            skipped.append({"key": key, "reason": "TK đối ứng không hợp lệ"})
            continue
        party = (r.get("party") or "").strip() or None
        if cacc.account_type in ("Receivable", "Payable") and not party:
            skipped.append({"key": key, "reason": "TK đối ứng cần đối tượng (KH/NCC)"})
            continue
        party_type = None
        if party:
            party_type = "Customer" if cacc.account_type == "Receivable" else "Supplier"

        je = frappe.new_doc("Journal Entry")
        je.voucher_type = "Journal Entry"
        je.posting_date = r.get("date")
        je.company = company
        # remark = ghi chú TỰ NHẬP (kèm marker chống trùng — dedup dò trên field này);
        # user_remark = NỘI DUNG sao kê nguyên bản.
        note = (r.get("remark") or "").strip()
        content = (r.get("content") or "").strip()
        if field == "remark":
            je.set("remark", f"[BANKIMP-{key}]" + (f" {note[:240]}" if note else ""))
            je.set("user_remark", content[:240])
        else:
            # Site không có field remark riêng → gộp cả vào user_remark.
            je.set("user_remark", (f"[BANKIMP-{key}] " + (f"{note} | " if note else "") + content)[:300])

        bank_line = {"account": bank_account}
        counter_line = {"account": counter}
        if party:
            counter_line["party_type"] = party_type
            counter_line["party"] = party
        if credit > 0:
            # Tiền VÀO ngân hàng: Nợ ngân hàng / Có đối ứng
            bank_line["debit_in_account_currency"] = amount
            counter_line["credit_in_account_currency"] = amount
        else:
            # Tiền RA khỏi ngân hàng: Nợ đối ứng / Có ngân hàng
            counter_line["debit_in_account_currency"] = amount
            bank_line["credit_in_account_currency"] = amount
        je.append("accounts", counter_line if credit <= 0 else bank_line)
        je.append("accounts", bank_line if credit <= 0 else counter_line)

        try:
            je.insert()
        except Exception as e:
            skipped.append({"key": key, "reason": str(e)[:120]})
            continue
        existing.add(key)
        row_out = {"key": key, "name": je.name, "route": f"/desk/journal-entry/{je.name}"}
        try:
            je.submit()
            row_out["docstatus"] = 1
        except Exception as e:
            # Đã tạo nháp nhưng ghi sổ lỗi (kỳ khóa, thiếu quyền submit...) — giữ nháp để xử lý trong Desk.
            row_out["docstatus"] = 0
            row_out["note"] = str(e)[:120]
        created.append(row_out)

    submitted = sum(1 for c in created if c.get("docstatus") == 1)
    return {
        "created": created, "skipped": skipped,
        "count": len(created), "submitted": submitted,
        "draft_count": len(created) - submitted,
    }


# ═══════════════════════════════════════════════════════════════════════════
# Quản lý quy tắc map nội dung → TK đối ứng
# ═══════════════════════════════════════════════════════════════════════════

@frappe.whitelist()
def get_rules(company: str | None = None) -> list:
    """Danh sách quy tắc map (cho company hoặc dùng chung)."""
    guard_cash()
    company = resolve_company(company)
    return frappe.get_all(
        "Ketoan Bank Map Rule",
        or_filters=[["company", "=", company], ["company", "is", "not set"]],
        fields=["name", "keyword", "counter_account", "direction", "party_type", "party", "priority", "enabled"],
        order_by="priority desc, keyword asc",
    )


@frappe.whitelist()
def save_rule(keyword: str, counter_account: str, direction: str = "Bất kỳ",
              party_type: str | None = None, party: str | None = None,
              priority: int = 0, company: str | None = None, name: str | None = None) -> dict:
    """Tạo/cập nhật 1 quy tắc map."""
    guard_cash()
    company = resolve_company(company)
    keyword = (keyword or "").strip()
    if not keyword:
        frappe.throw(_("Thiếu từ khóa"))
    if not counter_account or not frappe.db.exists("Account", counter_account):
        frappe.throw(_("TK đối ứng không hợp lệ"))
    if direction not in ("Bất kỳ", "Tiền vào", "Tiền ra"):
        direction = "Bất kỳ"
    party = (party or "").strip() or None
    party_type = party_type or None
    if party and not party_type:
        frappe.throw(_("Đã chọn đối tượng thì phải chọn loại đối tượng"))

    doc = frappe.get_doc("Ketoan Bank Map Rule", name) if name else frappe.new_doc("Ketoan Bank Map Rule")
    doc.update({
        "keyword": keyword,
        "counter_account": counter_account,
        "direction": direction,
        "party_type": party_type,
        "party": party,
        "priority": int(priority or 0),
        "enabled": 1,
        "company": company,
    })
    doc.save()
    return {"name": doc.name, "keyword": doc.keyword}


@frappe.whitelist()
def delete_rule(name: str) -> dict:
    """Xóa 1 quy tắc map."""
    guard_cash()
    if name and frappe.db.exists("Ketoan Bank Map Rule", name):
        frappe.delete_doc("Ketoan Bank Map Rule", name, ignore_permissions=False)
    return {"deleted": name}
