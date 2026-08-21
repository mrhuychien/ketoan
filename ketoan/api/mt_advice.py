# -*- coding: utf-8 -*-
"""mt_advice — tầng ĐỌC bảng kê thanh toán của các chuỗi siêu thị (kênh MT).

CẢ TÁM chuỗi đã có tầng đọc: WinCommerce · Central Retail · LOTTE · Emart ·
Saigon Co.op · AEON · Fuji · Mega Market — xem `PARSERS` ở cuối module.

VÌ SAO có module riêng, không dùng lại `misa_import._rows()`:

1. `misa_import._rows()` chỉ đọc `.xlsx` (openpyxl) và `.csv`. **2/5 chuỗi gửi
   `.xls` BIFF thật** (LOTTE, Emart) — openpyxl ném lỗi. Không đọc được `.xls`
   là kênh MT dùng không được.
2. `misa_import._rows()` chỉ trả **sheet đầu tiên**. WinCommerce có 9 sheet (bảng
   dữ liệu bị cắt rời qua Table 5/7/8) và Co.op có 8 sheet (mỗi sheet là MỘT lần
   thanh toán riêng). Đọc sheet đầu là bỏ sót 71% tiền của WinCommerce và 64%
   của Co.op.
3. MISA có MỘT khuôn cột; MT có **bảy khuôn khác hẳn nhau**, ba quy ước dấu khác
   nhau, và mỗi chuỗi có kiểu dòng rác riêng. Dò cột theo từ khóa chung như MISA
   sẽ gán sai cột tiền.

Module này THUẦN ĐỌC: không ghi database, không tạo/sửa/hủy chứng từ kế toán.
Con người xem kết quả, đối chiếu số kiểm tra, rồi mới quyết định ghi nhận.

═══════════════════════════════════════════════════════════════════════════════
NGUYÊN TẮC XUYÊN SUỐT (vi phạm là sai tiền, không phải sai style)
═══════════════════════════════════════════════════════════════════════════════

* **Không phân loại dòng bằng DẤU của số tiền.** Có chuỗi để hàng hóa dương, có
  chuỗi để hàng hóa âm, và AEON để hàng TRẢ dương trong bảng nhưng in âm ở khối
  tổng của chính nó, và trong CÙNG một chuỗi vẫn có ngoại lệ (Central Retail
  có dòng `K1` dương là trả hàng; LOTTE có dòng NET OFF cả dương lẫn âm). Luôn
  phân loại bằng cột loại chứng từ / ký hiệu hóa đơn của từng chuỗi.
* **Không đoán tên cột, không đoán ký hiệu hóa đơn.** Emart và Fuji không in ký
  hiệu ⇒ `inv_series=None` và luôn `Cần review`. Đoán ký hiệu là gán nhầm hóa đơn.
* **Không tự suy ra tiền trước thuế / tiền thuế.** 6/7 chuỗi chỉ có MỘT cột tiền
  (đã gồm thuế). Chia 1.1 hay 1.08 để "suy ra" là bịa số.
* **Dòng tổng trong file là SỐ KIỂM TRA, không phải tiền phát sinh.** Cộng nhầm
  là nhân đôi tiền. Ở đây chúng được tách sang `declared_totals` + `checks` để
  đối chiếu, không bao giờ nằm trong `rows`.
* **Số kiểm tra do file in ra là chân lý, không tự tính lại.** Co.op làm tròn
  chiết khấu ở cấp DÒNG và làm tròn tiền thanh toán ở cấp NHÓM độc lập nhau nên
  `trị giá − chiết khấu ≠ thanh toán` ở 17/374 nhóm (lệch ±1đ). Tự tính lại là
  tạo báo động giả cho kế toán.

Về kiểu số: mọi số tiền trong 5 file đều là **VND nguyên đồng**. Số nguyên dưới
2^53 biểu diễn CHÍNH XÁC bằng float64, nên cộng dồn hàng trăm dòng không sinh sai
số — dùng float ở đây an toàn. Chỗ dễ mất chữ số là lúc BÓC CHUỖI (dấu chấm là
phân cách nghìn kiểu VN chứ không phải thập phân), nên `to_number()` xử lý riêng
phần đó rất kỹ.

Kết quả trả về của `read_payment_advice()`:

    {
      "chain":            "LOTTE"            # nhãn đúng Select của MT Payment Advice
      "chain_key":        "lotte"            # khóa ASCII nội bộ
      "advice_no":        "2000141337"|None  # số chứng từ thanh toán, nếu file có
      "payment_dates":    ["2026-07-10", ...]# MỘT file có thể có NHIỀU ngày (§H)
      "declared_totals":  {...}              # SỐ KIỂM TRA đọc từ chính file
      "computed_totals":  {...}              # tổng do parser cộng từ các dòng
      "checks":           [{label, declared, computed, diff, ok}, ...]
      "reconciled":       True/False         # mọi check khớp?
      "groups":           [...]              # mỗi lần thanh toán riêng trong file
      "rows":             [...]              # DÒNG TIỀN, không có dòng tổng/rác
      "warnings":         [...]              # cảnh báo cho người duyệt
    }
"""

import base64
import binascii
import datetime
import io
import re
import unicodedata

import frappe
from frappe import _

from ketoan.api._guard import guard_mt
from ketoan.misa_integration.doctype.misa_invoice_snapshot.misa_invoice_snapshot import (
    norm_inv_no,
    norm_series,
)
from ketoan.mt.doctype.mt_payment_advice_line.mt_payment_advice_line import norm_series_mt

# Giới hạn phòng file hỏng/khổng lồ làm treo worker.
MAX_ROWS_PER_SHEET = 50000
MAX_SHEETS = 200
# Bảng kê rộng nhất trong 5 file là 15 cột (LOTTE). 200 đã dư rất xa; cao hơn nữa
# nghĩa là file khai khống vùng dữ liệu để bơm bộ nhớ, không phải bảng kê thật.
MAX_COLS_PER_SHEET = 200

# TRẦN Ô CỦA CẢ WORKBOOK — thứ DUY NHẤT đo đúng cái tốn kém.
#
# Ba trần trên đều là trần THEO TỪNG SHEET, nhân với nhau ra ngân sách 2 tỉ ô
# (50.000 x 200 x 200). Đo thật trên file dựng riêng để thử:
#     1.426 byte,   1 sheet  khai 50.000 dòng x 100 cột -> 20,6 giây · 1.073 MB
#     2.007 byte,   3 sheet  y hệt                      -> 83,9 giây · 3.171 MB
#     3.997 byte,  10 sheet  y hệt                      -> quá 120 giây, phải kill
#    58.342 byte, 200 sheet  y hệt (VẪN dưới MAX_SHEETS) -> ngoại suy ~70 GB
# Tức là một file vài KB, hợp lệ dưới MỌI trần cũ, vẫn giết được worker — và chỉ
# cần role 'Ke Toan MT' gọi được `preview`. Trần dung lượng 12MB vô nghĩa vì
# .xlsx là file nén: khai khống vùng dữ liệu tốn vài chục byte.
#
# 5 file thật lớn nhất: Co.op 817 dòng x 13 cột x 9 sheet ~ 9.500 ô. Để 2 triệu
# vẫn dư hơn 200 lần mà chặn đứng mọi ca bơm bộ nhớ.
MAX_CELLS_TOTAL = 2_000_000

# Khóa ASCII nội bộ -> nhãn đúng option của field `chain` trên MT Payment Advice.
# Fieldname/khóa ASCII, label tiếng Việt — theo ràng buộc của dự án.
CHAIN_LABEL = {
    "wincommerce": "WinCommerce",
    "central_retail": "Central Retail",
    "lotte": "LOTTE",
    "emart": "Emart",
    "coop": "Saigon Co.op",
    "aeon": "AEON",
    "fuji": "Fuji",
    # Có nhãn để `_resolve_chain_key` nhận ra tên chuỗi và báo đúng lý do "chưa
    # có parser", thay vì báo "không nhận ra chuỗi" làm kế toán tưởng file hỏng.
    "mega_market": "Mega Market",
}

# row_kind ASCII -> nhãn đúng option của field `row_kind` trên MT Payment Advice Line.
ROW_KIND_LABEL = {
    "thanh_toan": "Thanh toán",
    "chiet_khau": "Chiết khấu",
    "phi": "Phí",
    "ghi_giam": "Ghi giảm",
    "khac": "Khác",
}

# Sai số cho phép khi đối chiếu với số kiểm tra trong file. Tiền VND nguyên đồng
# nên về nguyên tắc phải lệch ĐÚNG 0; 0.5 chỉ để chống rác dấu phẩy động khi file
# lỡ xuất số thực. Lệch 1 đồng vẫn là LỆCH, phải báo.
MONEY_EPS = 0.5


# ═══════════════════════════════════════════════════════════════════════════
# 1. Hàm chung — chuẩn hóa chuỗi / số / ngày
# ═══════════════════════════════════════════════════════════════════════════

def strip_tones(s) -> str:
    """Bỏ dấu tiếng Việt + hạ chữ thường, để so nhãn cột không phụ thuộc dấu.

    Cần thật: Central Retail có 63 dòng ghi 'Hàng hóa quầy 0410' và 3 dòng ghi
    'Hang hoa quay 0410' — cùng nội dung, khác dấu, trong CÙNG một file.
    """
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.replace("đ", "d").replace("Đ", "D").lower().strip()


def _txt(v) -> str:
    """Ô bất kỳ -> chuỗi đã trim. None/rỗng -> ''.

    VÌ SAO không dùng thẳng str(): openpyxl/xlrd trả float cho ô số, `str(3996.0)`
    ra '3996.0' và khớp số hóa đơn sẽ trượt sạch. Ngoài ra file Emart/Co.op có
    ký tự \\xa0 (non-breaking space) thật, `strip()` thường không sạch.
    """
    if v is None:
        return ""
    if isinstance(v, bool):
        return "1" if v else "0"
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else repr(v)
    if isinstance(v, int):
        return str(v)
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()
    return str(v).replace("\xa0", " ").strip()


_NUM_OK_RE = re.compile(r"^[0-9.,]+$")


def to_number(v):
    """Ô tiền -> float, hoặc None nếu KHÔNG đọc được. Không bao giờ đoán 0.

    VÌ SAO trả None chứ không 0.0 như `misa_import.to_number`: ở đây 0 là một số
    tiền hợp lệ (WinCommerce có 36/36 dòng chiết khấu = 0). Nuốt lỗi thành 0 sẽ
    làm một dòng đọc hỏng trông y hệt một dòng chiết khấu bằng 0.

    Các dạng thật đã gặp trong 5 file:
      1234567            (số thuần — Central Retail, Emart, Co.op, WinCommerce)
      '12,365,000'       (LOTTE: phẩy là phân cách NGHÌN)
      '-274,950'         (LOTTE: dấu trừ đứng trước)
      '70.880.508'       (WinCommerce: CHẤM là phân cách NGHÌN kiểu VN)
      '******245.795.904*' (WinCommerce Table 9: số tổng bị bọc dấu sao)
      '(1.234)' / '1.234-' (dạng kế toán — chưa gặp thật, giữ nhánh để phòng)
    Nhầm quy ước dấu chấm là sai 1000 lần, nên phần này tách bạch rất kỹ.
    """
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)

    s = str(v).replace("\xa0", "").replace(" ", "").strip()
    # WinCommerce bọc số tổng bằng dấu sao: '******245.795.904*'
    s = s.strip("*")
    if not s:
        return None

    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg, s = True, s[1:-1]
    if s.endswith("-"):
        neg, s = True, s[:-1]
    if s.startswith("-"):
        neg, s = True, s[1:]
    elif s.startswith("+"):
        s = s[1:]
    if not s or not _NUM_OK_RE.match(s):
        return None

    if "," in s and "." in s:
        # Dấu đứng SAU cùng là dấu thập phân, dấu còn lại là phân cách nghìn.
        dec = "," if s.rfind(",") > s.rfind(".") else "."
        thou = "." if dec == "," else ","
        s = s.replace(thou, "").replace(dec, ".")
    elif "," in s:
        parts = s.split(",")
        # '12,365,000' -> mọi cụm sau dấu phẩy đều đúng 3 chữ số = phân cách nghìn.
        s = s.replace(",", "") if all(len(p) == 3 for p in parts[1:]) else s.replace(",", ".")
    elif "." in s:
        parts = s.split(".")
        # '70.880.508' -> phân cách nghìn. '17.75' -> thập phân, giữ nguyên.
        if all(len(p) == 3 for p in parts[1:]):
            s = s.replace(".", "")

    try:
        x = float(s)
    except ValueError:
        return None
    return -x if neg else x


# Năm 4 chữ số đứng TRƯỚC nên không mơ hồ ngày/tháng -> nhận cả '-', '/', '.'.
# Không đụng tới _DATE_DMY_RE: ở đó năm đứng SAU, đổi thứ tự là sai mọi ngày <= 12.
_DATE_ISO_RE = re.compile(r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})")
_DATE_DMY_RE = re.compile(r"^(\d{1,2})[-/.](\d{1,2})[-/.](\d{4})$")
_DATE_YMD8_RE = re.compile(r"^(\d{4})(\d{2})(\d{2})$")


def to_date(v):
    """Ô ngày -> 'YYYY-MM-DD', hoặc None. KHÔNG đoán khi mơ hồ.

    Bốn dạng thật trong 5 file:
      datetime thật      (WinCommerce, Central Retail, Co.op — openpyxl trả sẵn)
      '20260523'         (LOTTE — chuỗi YYYYMMDD, KHÔNG phải serial Excel)
      '31-08-2025'       (Emart, trong bảng)
      '15/09/2025'       (Emart, ở khối header — CÙNG file, KHÁC dấu phân cách)
      '25.06.2026'       (WinCommerce, trong khối text nhiều dòng)

    VÌ SAO từ chối số trần (float/int chưa qua _txt): số 45000 có thể là serial
    ngày Excel mà cũng có thể là số tiền. `xlrd.xldate_as_tuple` trên ô TEXT của
    LOTTE cho ra ngày 1970 — sai kỳ thanh toán. Thà trả None để người xử lý.
    Thứ tự luôn NGÀY trước THÁNG sau; đọc nhầm thành mm-dd là sai mọi ngày <= 12.
    """
    if v is None or v == "":
        return None
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()

    s = _txt(v)
    if not s:
        return None

    m = _DATE_ISO_RE.match(s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    else:
        m = _DATE_DMY_RE.match(s)
        if m:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        else:
            m = _DATE_YMD8_RE.match(s)
            if not m:
                return None
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return datetime.date(y, mo, d).isoformat()
    except ValueError:
        return None


def norm_series_loose(s) -> str:
    """Ký hiệu hóa đơn -> dạng so khớp: bỏ chữ số dạng hóa đơn ở ĐẦU.

    '1C26THG' và 'C26THG' là MỘT ký hiệu (§E hợp đồng). Đã quan sát thật:
      * Central Retail: 4 dòng '1C26THG|...' lẫn giữa 180 dòng 'C26THG|...'
      * LOTTE: 35 dòng '1C26THG' + 10 dòng 'C26THG' cùng dải số hóa đơn
      * Co.op: 'K26TEK-80' thiếu hẳn chữ số đầu
    Không bỏ chữ số đầu là 10-14 hóa đơn không tìm được Sales Invoice.

    DÙNG LẠI `norm_series_mt` của MT Payment Advice Line, KHÔNG cài lại logic:
    bản cũ ở đây thiếu nhánh 'ký hiệu toàn chữ số thì trả nguyên bản', nên với ký
    hiệu ghi nhầm thành '126' bản xem trước hiển thị 'không có ký hiệu' (rỗng)
    trong khi tầng khớp lại thấy '126' và đi nhánh khớp theo khóa. Kế toán soi
    preview rồi kết luận sai về việc dòng đó đã được khớp bằng cách nào. Chuẩn
    hóa ký hiệu phải có ĐÚNG MỘT định nghĩa cho cả hai tầng.
    """
    return norm_series_mt(s)


# Dấu phân cách ký hiệu|số đã gặp thật. Truyền `seps` tường minh theo từng chuỗi
# thay vì dùng chung regex 'ký tự không phải chữ số': WinCommerce dùng '#',
# Central Retail dùng '|'. Gộp bừa là nuốt nhầm dấu khác trong chuỗi mô tả.
REF_SEPS_WINCOMMERCE = "#"
REF_SEPS_CENTRAL_RETAIL = "|"


def split_invoice_ref(raw, seps):
    """'1C26THG#1730' -> ('1C26THG', '1730'); 'C26THG|4675' -> ('C26THG', '4675').

    Tách tại dấu phân cách ĐẦU TIÊN thuộc `seps`. Không có dấu nào thì trả
    ký hiệu RỖNG và đưa cả chuỗi vào số — TUYỆT ĐỐI không đoán ký hiệu.
    Có thật: Central Retail có dòng Reference = 'CK T07.2026' (không có '|');
    `ref.split('|')[1]` ở đó nổ IndexError.
    """
    t = _txt(raw)
    if not t:
        return "", ""
    pos = min((t.find(c) for c in seps if c in t), default=-1)
    if pos < 0:
        return "", t
    return t[:pos].strip(), t[pos + 1:].strip()


def _abs(x):
    return None if x is None else abs(x)


def _sum(rows, key, kinds=None):
    """Cộng một khóa số trên các dòng. None coi như bỏ qua, không coi là 0."""
    tot = 0.0
    for r in rows:
        if kinds is not None and r["row_kind"] not in kinds:
            continue
        v = r.get(key)
        if v is not None:
            tot += v
    return tot


def _check(label, declared, computed):
    """Một mục đối chiếu declared (file in ra) vs computed (parser cộng)."""
    ok = declared is not None and abs(declared - computed) <= MONEY_EPS
    return {
        "label": label,
        "declared": declared,
        "computed": computed,
        "diff": None if declared is None else round(declared - computed, 2),
        "ok": bool(ok),
    }


def _assert_rows_frozen(rows, n_at_totals, where):
    """Cấm thêm dòng tiền SAU khi parser đã cộng tổng và dựng số kiểm tra.

    VÌ SAO cần chốt máy móc này: `parse_sheets` cộng LẠI tổng từ `rows` sau khi
    parser trả về, còn `checks` thì parser đã dựng xong từ trước. Một dòng thêm
    muộn sẽ vào tổng của DocType mà KHÔNG vào bất kỳ số kiểm tra nào — kế toán
    thấy `reconciled=True` trong khi tiền đã lệch. Đã đo bằng đột biến: chèn
    thêm dòng phí sau mục '── Đối chiếu ──' của `parse_aeon` thì 16/16 số kiểm
    tra vẫn báo khớp.

    Đây là lỗi LẬP TRÌNH, không phải lỗi dữ liệu, nên throw chứ không warning.
    """
    if len(rows) != n_at_totals:
        frappe.throw(_("Lỗi lập trình ở {0}: {1} dòng tiền được thêm SAU khi đã dựng "
                       "số kiểm tra — tổng sẽ lệch mà không phép kiểm nào phát hiện.")
                     .format(where, len(rows) - n_at_totals))


def _line(**kw):
    """Tạo một dòng theo BỘ KHÓA CHUẨN dùng chung cho mọi chuỗi.

    `total_amount` / `amount_before_vat` / `vat_amount` là ĐỘ LỚN (trị tuyệt đối)
    để còn so trực tiếp với `grand_total` của Sales Invoice; chiều tiền đã nằm ở
    `row_kind`. `signed_amount` giữ NGUYÊN DẤU của file — đây mới là số dùng để
    đối chiếu với dòng tổng do chuỗi in ra, đổi dấu ở đó là mất chốt kiểm tra.
    """
    series = kw.get("inv_series") or ""
    inv_no = kw.get("inv_no") or ""
    kind = kw["row_kind"]
    signed = kw.get("signed_amount")
    return {
        "row_kind": kind,
        "row_kind_label": ROW_KIND_LABEL[kind],
        "row_subtype": kw.get("row_subtype") or "",
        "inv_series": series or None,
        # ký hiệu đã bỏ chữ số dạng hóa đơn ở đầu — dùng để KHỚP, không để hiển thị
        "inv_series_norm": norm_series_loose(series) or None,
        "inv_no": inv_no or None,
        "inv_no_norm": norm_inv_no(inv_no) or None,
        "inv_date": kw.get("inv_date"),
        "store_code": kw.get("store_code") or None,
        "store_name": kw.get("store_name") or None,
        "doc_no": kw.get("doc_no") or None,
        "description": kw.get("description") or None,
        "amount_before_vat": _abs(kw.get("amount_before_vat")),
        "vat_amount": _abs(kw.get("vat_amount")),
        "total_amount": _abs(signed),
        "signed_amount": signed,
        "payment_date": kw.get("payment_date"),
        # Cờ này bắt buộc phải chảy tới tận `match_confidence` = 'Cần review'.
        "needs_review": bool(kw.get("needs_review")),
        "source_sheet": kw.get("source_sheet") or None,
        "source_row": kw.get("source_row"),
    }


# ═══════════════════════════════════════════════════════════════════════════
# 2. Đọc file — .xlsx (openpyxl) VÀ .xls (xlrd)
# ═══════════════════════════════════════════════════════════════════════════

_XLS_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"  # Composite Document File V2 (BIFF)


def _check_sheet_count(n):
    """Quá nhiều sheet thì DỪNG, không lặng lẽ cắt bớt.

    `sheets[:MAX_SHEETS]` bỏ phần dôi ra mà không nói gì — mỗi sheet của các chuỗi
    là MỘT TRANG bảng kê (§J.1), bỏ sót sheet là bỏ sót tiền trong khi số kiểm tra
    ở sheet đầu vẫn khớp. Bảng kê thật nhiều sheet nhất mới có 9 sheet.
    """
    if n > MAX_SHEETS:
        frappe.throw(_("File có {0} sheet (trần {1}) — không phải bảng kê thật.")
                     .format(n, MAX_SHEETS))


def decode_upload(content) -> bytes:
    """base64 (có/không tiền tố data URI) -> bytes.

    VÌ SAO bọc try/except: `FileReader.readAsDataURL` bị cắt giữa chừng (mạng rớt
    lúc tải file lớn) cho chuỗi sai padding; `b64decode` khi đó ném `binascii.Error`
    TRẦN, chạy TRƯỚC mọi try/except của `read_sheets` nên kế toán nhận traceback 500
    thay vì một câu tiếng Việt bảo chọn lại file.
    """
    try:
        raw = base64.b64decode((content or "").split(",")[-1])
    except (binascii.Error, ValueError):
        frappe.throw(_("Nội dung file tải lên không hợp lệ, hãy chọn lại file"))
    if not raw:
        frappe.throw(_("File rỗng"))
    # Trần dung lượng đặt ở ĐÂY, không ở từng whitelisted method: `preview()` và
    # `detect()` của module này trước đó không hề gọi `_check_size`, nên bất kỳ ai
    # có role 'Ke Toan MT' cũng nạp được file 100MB và giết worker của phòng kế
    # toán. Đặt trong `decode_upload` thì MỌI đường vào đều đi qua đúng một chốt.
    # Import trễ để không tạo vòng import với `ketoan.api.mt` (mt cũng nạp mt_advice).
    from ketoan.api.mt import MAX_UPLOAD_BYTES

    if len(raw) > MAX_UPLOAD_BYTES:
        frappe.throw(_("File quá lớn ({0} MB). Tách bớt rồi tải lại.")
                     .format(round(len(raw) / 1024 / 1024, 1)))
    return raw


def _sheets_xlsx(raw, allow_wide=False):
    """.xlsx -> [(tên sheet, lưới giá trị)].

    KHÔNG dùng read_only: Co.op cần đọc theo tọa độ tuyệt đối (I5/I6/I7, M13-M15)
    và cả 5 file đều nhỏ (<110KB). data_only=True để lấy giá trị đã cache thay vì
    công thức — nếu file do máy tạo mà chưa cache thì ô ra None và parser sẽ báo
    lệch số kiểm tra chứ không âm thầm đọc sai.
    """
    from itertools import islice

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), data_only=True)
    _check_sheet_count(len(wb.worksheets))

    # CỘNG DỒN Ô TOÀN WORKBOOK TRƯỚC KHI ĐỌC BẤT KỲ SHEET NÀO.
    # Kiểm từng sheet rời không đủ: mỗi sheet lọt trần vẫn nhân lên 200 lần.
    # `max_row`/`max_column` đọc từ khai báo dimension nên rẻ, chưa vật chất hóa ô nào.
    total_cells = sum(min(ws.max_row or 0, MAX_ROWS_PER_SHEET + 1)
                      * min(ws.max_column or 0, MAX_COLS_PER_SHEET + 1)
                      for ws in wb.worksheets)
    if total_cells > MAX_CELLS_TOTAL:
        wb.close()
        frappe.throw(_("File khai {0} ô trên {1} sheet — vượt trần {2}. Không phải bảng kê thật.")
                     .format(f"{total_cells:,}", len(wb.worksheets), f"{MAX_CELLS_TOTAL:,}"))

    out = []
    for ws in wb.worksheets:
        # CHẶN TRƯỚC KHI ĐỌC, không cắt sau khi đã đọc.
        # VÌ SAO: một file .xlsx 1,4KB chỉ cần khai hai ô r="1" và r="1048576" là
        # `ws.max_row` = 1.048.576; `[list(r) for r in ws.iter_rows()][:50000]` sẽ
        # SINH ĐỦ hơn một triệu tuple (~26GB) rồi mới cắt còn 50.000 -> worker bị
        # OOM-kill. Trần dung lượng 12MB không đỡ được vì file chỉ 1,4KB.
        wide = (ws.max_column or 0) > MAX_COLS_PER_SHEET
        if (ws.max_row or 0) > MAX_ROWS_PER_SHEET or (wide and not allow_wide):
            frappe.throw(_("Sheet {0} quá lớn ({1} dòng x {2} cột) — không phải bảng kê thật.")
                         .format(ws.title, ws.max_row, ws.max_column))
        ncols = min(ws.max_column or 0, MAX_COLS_PER_SHEET)
        grid = [list(r) for r in islice(
            ws.iter_rows(values_only=True, max_col=ncols or None), MAX_ROWS_PER_SHEET)]
        out.append((ws.title, grid))
    return out


def _sheets_xls(raw):
    """.xls BIFF -> [(tên sheet, lưới giá trị)]. Bắt buộc dùng xlrd (2/5 chuỗi).

    Giữ NGUYÊN kiểu ô: LOTTE có Store CD = '01019' là TEXT, ép sang số là mất số 0
    ở đầu và hỏng khớp mã siêu thị. Ô ngày kiểu XL_CELL_DATE mới đổi sang datetime
    — ngày dạng TEXT ('20260523', '31-08-2025') để nguyên cho `to_date` xử lý.
    """
    import xlrd

    book = xlrd.open_workbook(file_contents=raw)
    dm = book.datemode
    _check_sheet_count(book.nsheets)

    # Trần ô toàn workbook, cùng lý do như nhánh .xlsx: trần theo từng sheet vẫn
    # cho nhân lên 200 lần. Với .xls thì xlrd đã nạp sẵn nên chốt này không cứu
    # được bộ nhớ lúc mở, nhưng vẫn chặn vòng lặp dựng lưới bên dưới — chỗ tốn kém.
    total_cells = sum(sh.nrows * sh.ncols for sh in book.sheets())
    if total_cells > MAX_CELLS_TOTAL:
        frappe.throw(_("File khai {0} ô trên {1} sheet — vượt trần {2}. Không phải bảng kê thật.")
                     .format(f"{total_cells:,}", book.nsheets, f"{MAX_CELLS_TOTAL:,}"))

    out = []
    for sh in book.sheets():
        # DỪNG thay vì cắt ngầm: cắt ở dòng 50.000 mà không nói gì thì phần tiền
        # phía sau biến mất trong im lặng — đúng kiểu lỗi mà số kiểm tra ở đầu file
        # vẫn khớp. Bảng kê thật lớn nhất mới 817 dòng.
        if sh.nrows > MAX_ROWS_PER_SHEET or sh.ncols > MAX_COLS_PER_SHEET:
            frappe.throw(_("Sheet {0} quá lớn ({1} dòng x {2} cột) — không phải bảng kê thật.")
                         .format(sh.name, sh.nrows, sh.ncols))
        grid = []
        for r in range(min(sh.nrows, MAX_ROWS_PER_SHEET)):
            row = []
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                t, v = cell.ctype, cell.value
                if t == xlrd.XL_CELL_TEXT:
                    s = v.replace("\xa0", " ").strip()
                    row.append(s or None)
                elif t == xlrd.XL_CELL_NUMBER:
                    row.append(v)
                elif t == xlrd.XL_CELL_DATE:
                    y, mo, d, hh, mi, ss = xlrd.xldate_as_tuple(v, dm)
                    row.append(datetime.datetime(y, mo, d, hh, mi, ss) if y else None)
                elif t == xlrd.XL_CELL_BOOLEAN:
                    row.append(bool(v))
                else:
                    row.append(None)  # EMPTY / BLANK / ERROR
            grid.append(row)
        out.append((sh.name, grid))
    return out


def read_sheets(content, allow_wide=False):
    """base64 -> [(tên sheet, lưới)]. Tự nhận .xlsx / .xls theo chữ ký file.

    Nhận diện theo BYTE ĐẦU chứ không theo đuôi tên file: kế toán hay đổi tên file
    hoặc "Save As" nhầm định dạng, và đọc .xls bằng openpyxl thì lỗi khó hiểu.

    `allow_wide=True`: sheet KHAI nhiều hơn `MAX_COLS_PER_SHEET` cột thì CẮT còn
    đúng trần đó thay vì từ chối cả file.

    VÌ SAO cần và vì sao vẫn an toàn: file theo dõi công nợ Emart khai 16.375 cột
    trong khi cột có dữ liệu xa nhất là 17 — bề rộng đó là rác định dạng, và file
    chỉ 187KB. Chốt chặn ban đầu sinh ra để chống OOM, mà OOM đến từ việc VẬT CHẤT
    HÓA ô: đọc với `max_col` đã cắt thì bộ nhớ bị chặn bởi (số dòng × trần cột) bất
    kể sheet khai bao nhiêu. Trần DÒNG vẫn giữ nguyên, không nới.

    Mặc định là `False` để đường nạp bảng kê thanh toán KHÔNG đổi hành vi.
    """
    raw = decode_upload(content)
    if raw[:2] == b"PK":
        try:
            return _sheets_xlsx(raw, allow_wide=allow_wide)
        except frappe.ValidationError:
            # Đây là frappe.throw của chính tầng đọc (vd 'Sheet X quá lớn') — thông
            # báo đã nói rõ hỏng ở đâu, nuốt nó thành câu chung là bịt mắt kế toán.
            raise
        except Exception:
            frappe.log_error(frappe.get_traceback(), "mt_advice.read_sheets/xlsx")
            frappe.throw(_("Không đọc được file .xlsx. Hãy kiểm tra lại file gốc."))
    if raw[:8] == _XLS_MAGIC:
        try:
            return _sheets_xls(raw)
        except frappe.ValidationError:
            raise
        except Exception:
            frappe.log_error(frappe.get_traceback(), "mt_advice.read_sheets/xls")
            frappe.throw(_("Không đọc được file .xls. Hãy kiểm tra lại file gốc."))
    frappe.throw(_("Định dạng không hỗ trợ. Kênh MT chỉ nhận file Excel .xlsx hoặc .xls"))


def _g(grid, r, c):
    """Ô theo tọa độ 1-based như người dùng nhìn trong Excel. Ngoài biên -> None."""
    if r < 1 or c < 1 or r > len(grid):
        return None
    row = grid[r - 1]
    return row[c - 1] if c <= len(row) else None


def _row_texts(row):
    return [_txt(c) for c in row]


def _is_blank(row):
    return all(_txt(c) == "" for c in row)


def _sheet_has_content(grid) -> bool:
    """Sheet còn ô nào không rỗng không?

    VÌ SAO cần: §J.1 — file của các chuỗi là bản in PDF convert sang Excel, mỗi
    TRANG IN thành một sheet, nên bảng dữ liệu bị cắt rời. Bỏ qua một sheet còn
    dữ liệu là BỎ SÓT TIỀN, mà số kiểm tra lại nằm ngay sheet đầu nên `reconciled`
    vẫn báo khớp. Hàm này để phân biệt "sheet trống thật" (bỏ qua được) với
    "sheet còn dữ liệu mà tầng đọc chưa hiểu" (phải dừng, không được đọc tiếp).
    """
    return any(not _is_blank(row) for row in grid)


def _declared_sum(parts):
    """Gộp số kiểm tra nhặt được từ NHIỀU sheet của cùng một file.

    Trả None khi chưa gặp giá trị nào HOẶC có giá trị đọc không ra: "không đọc
    được số kiểm tra" phải làm `reconciled=False`, TUYỆT ĐỐI không được coi là 0
    — coi là 0 thì một file thiếu chốt kiểm tra sẽ trông y hệt một file khớp 0đ.
    """
    if not parts or any(p is None for p in parts):
        return None
    return sum(parts)


# ═══════════════════════════════════════════════════════════════════════════
# 3. Nhận diện chuỗi
# ═══════════════════════════════════════════════════════════════════════════

# Mỗi chuỗi: danh sách "bộ dấu hiệu"; khớp khi CHỨA ĐỦ mọi từ trong một bộ.
# Dấu hiệu lấy từ NHÃN CỘT thật của file, đã bỏ dấu tiếng Việt.
# Dấu hiệu phải ĐẶC TRƯNG cho đúng một chuỗi, không được là nhãn phổ thông.
# Đã đo thật: bộ ('so chung tu','so hoa don','chiet khau') từng dùng cho
# WinCommerce cũng khớp luôn file Co.op (Co.op có 'Số chứng từ : 26303039' ở meta,
# 'SỐ HÓA ĐƠN LH' và 'CHIẾT KHẤU' ở header) ⇒ hai chuỗi cùng khớp ⇒ trả None ⇒
# file Co.op không nhận diện được. Nay chỉ giữ nhãn CHỈ WinCommerce mới có.
_CHAIN_SIGNS = {
    "lotte": [("deduct name", "deduct cause")],
    "emart": [("amount in doc. curr", "document type")],
    "central_retail": [("clearing doc", "doc.type", "reference")],
    "wincommerce": [
        ("so doi soat", "so hoa don"),          # cột 'Số đối soát' — chỉ WinCommerce có
        ("chung tu thanh toan", "so hoa don"),  # bảng Table 9
        ("so du mang sang trang sau",),         # chân trang bản in
    ],
    "coop": [("so hoa don lh",), ("dien giai", "tri gia")],
    # AEON: tiêu đề r1 của MỌI sheet. Không chuỗi nào khác in câu này.
    "aeon": [("payment statement to suppliers",)],
    # Fuji: nhãn hai tầng của khối 1. 'theo hdtc' là nhãn RIÊNG của Fuji.
    "fuji": [("theo hdtc",), ("gia tri tien theo pnk",)],
    # Mega Market: header tiếng Anh ở dòng 1. Đã đối chiếu với CẢ 7 file bảng kê
    # mẫu của các chuỗi khác — không file nào chứa bộ ba nhãn này.
    "mega_market": [("store no", "supplier code", "invoice no")],
}


def detect_chain(sheets):
    """Đoán chuỗi từ dấu hiệu tiêu đề. KHÔNG chắc thì trả None cho người chọn tay.

    Chấp nhận cả `[(tên sheet, lưới)]` lẫn một lưới trần.

    VÌ SAO trả None thay vì chọn đại cái gần nhất: parser sai chuỗi không "đọc
    thiếu", nó đọc SAI CỘT TIỀN và vẫn ra một con số trông hợp lý. Người dùng chọn
    tay mất 2 giây; ghi nhận nhầm chuỗi thì phải dò lại cả bảng kê.
    """
    if sheets and not isinstance(sheets[0], (tuple, list)):
        sheets = [("", sheets)]
    elif sheets and isinstance(sheets[0], list):
        sheets = [("", sheets)]

    blob_parts = []
    for _name, grid in sheets:
        # Dấu hiệu nằm ở vùng tiêu đề; Co.op có header tận r18-r19 nên quét 40 dòng.
        for row in grid[:40]:
            blob_parts.extend(strip_tones(c) for c in row if c is not None)
    blob = " | ".join(p for p in blob_parts if p)

    hits = [key for key, sign_sets in _CHAIN_SIGNS.items()
            if any(all(w in blob for w in words) for words in sign_sets)]
    if len(hits) == 1:
        return hits[0]
    return None  # 0 dấu hiệu, hoặc nhiều chuỗi cùng khớp -> để người chọn


# ═══════════════════════════════════════════════════════════════════════════
# 4a. WinCommerce  (.xlsx, 9 sheet 'Table N' — bản in PDF convert sang Excel)
# ═══════════════════════════════════════════════════════════════════════════

# Nhãn cột thật đọc từ file, không suy đoán.
_WC_COL = {
    "Số chứng từ": "doc_no",
    "Số hóa đơn": "invoice_raw",
    "Ngày hóa đơn": "inv_date",
    "Số đối soát": "recon_no",
    "Chiết khấu": "discount",
    "Số tiền": "amount",
}
_WC_HEADER_MARKERS = ("Số hóa đơn", "Số tiền")


def _wc_is_star(v) -> bool:
    """Dòng '**********' xen giữa mỗi dòng dữ liệu — chỉ là đường kẻ của bản in.

    VÌ SAO kiểm 'tập ký tự == {*}' chứ không so bằng chuỗi 10 sao: bản in còn có
    dòng 41 sao (Table 4). Cách này cũng KHÔNG dính '******245.795.904*' của
    Table 9 vì chuỗi đó có chữ số — đúng như mong muốn.
    """
    t = _txt(v)
    return len(t) >= 3 and set(t) == {"*"}


def _wc_find_header(grid):
    """Trả (chỉ số dòng header 1-based, {khóa: chỉ số cột 0-based}) hoặc (None, None).

    VÌ SAO dò header theo NHÃN thay vì hardcode cột: Table 7 có THÊM một cột A rỗng
    (merged A1:A42) nên mọi cột dịch phải 1 ô so với Table 5 và Table 8 ('Số tiền'
    ở index 6 thay vì 5). Hardcode chỉ số cột là đọc lệch cột tiền của 21/36 dòng.
    """
    for i, row in enumerate(grid):
        labels = {_txt(c): j for j, c in enumerate(row) if _txt(c)}
        if all(m in labels for m in _WC_HEADER_MARKERS):
            return i + 1, {_WC_COL[k]: j for k, j in labels.items() if k in _WC_COL}
    return None, None


def _wc_payment_meta(sheets):
    """Ngày thanh toán + số chứng từ thanh toán + số tiền tổng của lần thanh toán.

    HAI nguồn ngày và chúng phải khớp nhau:
      (1) Table 2 r2 — khối text nhiều dòng, nhãn 'Ngày thanh toán' ở một dòng và
          giá trị '25.06.2026' ở dòng NGAY DƯỚI (dd.mm.yyyy, dấu chấm).
      (2) Table 9 r2 cột 'Ngày' — datetime thật.
    Ưu tiên (2) vì là kiểu ngày thật, không phụ thuộc bóc chuỗi.
    """
    date_text, date_real, pay_doc, pay_amount = None, None, "", None
    for _name, grid in sheets:
        for row in grid:
            for c in row:
                t = _txt(c)
                if "Ngày thanh toán" in t and date_text is None:
                    lines = [x.strip() for x in t.split("\n")]
                    for k, ln in enumerate(lines):
                        if ln.startswith("Ngày thanh toán") and k + 1 < len(lines):
                            date_text = to_date(lines[k + 1])
                            break
        # Bảng Table 9 đọc theo NHÃN, không theo chỉ số cột cứng (row[0]/[1]/[3]).
        # VÌ SAO: §J.2 — Table 7 của chính file này đã có THÊM một cột A rỗng (merged
        # A1:A42) làm mọi cột dịch phải 1 ô. Cùng khiếm khuyết đó rơi vào Table 9 thì
        # `advice_no` về None; mất advice_no là mất KHÓA CHỐNG TRÙNG (`_existing_advice`
        # phải quay sang lọc theo tên file), kế toán tải lại cùng bảng kê với tên file
        # khác là tạo bản ghi thứ hai — 245.795.904đ bị đếm hai lần.
        for i, row in enumerate(grid):
            labels = {t: j for j, t in enumerate(_row_texts(row)) if t}
            if "Chứng từ thanh toán" in labels and i + 1 < len(grid):
                nxt = grid[i + 1]

                def at(key, _labels=labels, _nxt=nxt):
                    j = _labels.get(key)
                    return _nxt[j] if j is not None and j < len(_nxt) else None

                pay_doc = _txt(at("Chứng từ thanh toán"))
                date_real = to_date(at("Ngày")) or date_real
                # Ô này là CHUỖI '******245.795.904*' — to_number đã bóc sao + dấu chấm nghìn.
                pay_amount = to_number(at("Số tiền"))
    return (date_real or date_text), date_text, date_real, pay_doc, pay_amount


def parse_wincommerce(sheets):
    """Bảng kê WinCommerce: MỖI TRANG IN là MỘT SHEET, bảng dữ liệu bị cắt rời.

    Phải quét MỌI sheet và dò header — chỉ đọc Table 5 là mất 174.915.396đ (71%).
    Table 6 KHÔNG phải tiêu đề mục 'Chiết khấu' mà là CHÂN TRANG 'Số dư mang sang
    trang sau'; bằng chứng số học: 70.880.508 = đúng tổng Table 5. Nếu tin gợi ý
    ban đầu mà gán Table 7 = chiết khấu thì 134.593.596đ bị ghi sai loại và
    40.321.800đ của Table 8 bị bỏ quên hoàn toàn.
    """
    warnings = []
    pay_date, d_text, d_real, pay_doc, pay_amount = _wc_payment_meta(sheets)
    if d_text and d_real and d_text != d_real:
        warnings.append(
            "Ngày thanh toán ở khối text (%s) khác với bảng Chứng từ thanh toán (%s)."
            % (d_text, d_real))

    rows, carry_forwards = [], []
    declared_total, declared_discount = None, None

    for name, grid in sheets:
        hi, cols = _wc_find_header(grid)

        if hi is None:
            # Sheet không có bảng dữ liệu: khối thông tin bên mua/bên bán, lời chào,
            # hoặc chân trang chuyển số. Chỉ nhặt SỐ KIỂM TRA, không sinh dòng tiền.
            for row in grid:
                txt = " / ".join(t for t in _row_texts(row) if t)
                if "Số dư mang sang trang sau" in txt:
                    # SỐ KIỂM TRA cộng dồn hết trang trước, KHÔNG phải tiền mới.
                    # Cộng vào là nhân đôi tiền của trang 1.
                    m = re.search(r"Số dư mang sang trang sau\s+([\d.,]+)\s+([\d.,]+)", txt)
                    if m:
                        carry_forwards.append(to_number(m.group(2)))
            continue

        def cell(row, key):
            j = cols.get(key)
            return row[j] if j is not None and j < len(row) else None

        for j in range(hi, len(grid)):   # hi là 1-based -> grid[hi] là dòng ngay sau header
            row = grid[j]
            rno = j + 1
            if _is_blank(row):
                continue
            if _wc_is_star(cell(row, "inv_date")) or any(_wc_is_star(c) for c in row):
                continue  # đường kẻ của bản in

            first = _txt(cell(row, "doc_no")) if cols.get("doc_no") is not None else _txt(row[0])
            if first.lower().startswith("tổng cộng"):
                # SỐ KIỂM TRA cuối bảng — giữ để đối chiếu, tuyệt đối không nạp.
                declared_total = to_number(cell(row, "amount"))
                declared_discount = to_number(cell(row, "discount"))
                continue

            inv_raw = _txt(cell(row, "invoice_raw"))
            amount = to_number(cell(row, "amount"))
            if not inv_raw:
                # Ví dụ Table 7 r44: mảnh tiêu đề chân trang lặp lại, chỉ có chữ
                # 'Chiết khấu' và 'Số tiền', không số chứng từ, không tiền. Nếu lọc
                # 'dòng nào có chữ Chiết khấu thì là chiết khấu' sẽ đẻ ra dòng ma.
                continue
            if amount is None:
                # CÓ số hóa đơn mà KHÔNG đọc được tiền là dòng tiền bị mất, không phải
                # rác chân trang. Số kiểm tra sẽ báo lệch, nhưng nếu không nói rõ SHEET
                # và DÒNG nào thì kế toán phải dò tay 1 dòng giữa 9 sheet.
                warnings.append("Sheet %s dòng %d: có số hóa đơn %s nhưng KHÔNG đọc được cột "
                                "'Số tiền' (%r) — dòng này đã bị bỏ, tổng sẽ lệch."
                                % (name, rno, inv_raw, _txt(cell(row, "amount"))))
                continue

            series, no = split_invoice_ref(inv_raw, REF_SEPS_WINCOMMERCE)
            disc = to_number(cell(row, "discount")) or 0.0
            recon = _txt(cell(row, "recon_no"))
            rows.append(_line(
                row_kind="thanh_toan", row_subtype="hang_hoa",
                inv_series=series, inv_no=no,
                inv_date=to_date(cell(row, "inv_date")),
                doc_no=_txt(cell(row, "doc_no")),
                # File chỉ có MỘT cột tiền (đã gồm thuế) -> không suy VAT.
                amount_before_vat=None, vat_amount=None,
                signed_amount=amount,
                payment_date=pay_date,
                description="Số đối soát %s" % recon if recon else None,
                # Ký hiệu rỗng nghĩa là cột 'Số hóa đơn' không có '#' -> không khớp mù được.
                needs_review=not series,
                source_sheet=name, source_row=rno,
            ))
            if disc:
                # Chiết khấu WinCommerce nằm ở CỘT RIÊNG trên chính dòng hóa đơn.
                # Tách thành dòng ghi giảm để tổng chiết khấu của phiếu cộng được.
                # CHƯA XÁC MINH: file 25.06.2026 có chiết khấu = 0 ở cả 36/36 dòng.
                rows.append(_line(
                    row_kind="chiet_khau", row_subtype="chiet_khau_cot_rieng",
                    inv_series=series, inv_no=no,
                    inv_date=to_date(cell(row, "inv_date")),
                    doc_no=_txt(cell(row, "doc_no")),
                    signed_amount=disc, payment_date=pay_date,
                    description="Chiết khấu trên dòng hóa đơn",
                    needs_review=True,  # nhánh chưa từng chạy trên dữ liệu thật
                    source_sheet=name, source_row=rno,
                ))

    computed_payment = _sum(rows, "signed_amount", {"thanh_toan"})
    computed_discount = _sum(rows, "signed_amount", {"chiet_khau"})

    checks = [_check("Tổng cộng (cuối bảng)", declared_total, computed_payment)]
    if pay_amount is not None:
        checks.append(_check("Chứng từ thanh toán (Table 9)", pay_amount, computed_payment))
    if declared_discount is not None:
        checks.append(_check("Chiết khấu (dòng Tổng cộng)", declared_discount, computed_discount))
    # Chân trang 'Số dư mang sang trang sau' = tổng luỹ kế các trang TRƯỚC, không
    # phải tiền mới. Chỉ báo ra để người xem, không đưa vào phép đối chiếu tổng.
    if carry_forwards:
        warnings.append("Chân trang 'Số dư mang sang trang sau': %s (số kiểm tra luỹ kế, không cộng vào tổng)."
                        % ", ".join("{:,.0f}".format(x) for x in carry_forwards if x is not None))

    groups = [{
        "key": pay_doc or None, "advice_no": pay_doc or None, "payment_date": pay_date,
        "n_rows": len(rows), "computed_payment": computed_payment,
        "declared_payment": declared_total,
    }]

    return {
        "advice_no": pay_doc or None,
        "payment_dates": [pay_date] if pay_date else [],
        "declared_totals": {
            "total_payment": declared_total,
            "total_discount": declared_discount,
            "payment_document_amount": pay_amount,
            "carry_forward": carry_forwards,
        },
        "rows": rows, "checks": checks, "groups": groups, "warnings": warnings,
    }


# ═══════════════════════════════════════════════════════════════════════════
# 4b. Central Retail (EB/GO!)  (.xlsx 1 sheet, do SAP xuất)
# ═══════════════════════════════════════════════════════════════════════════

# Doc.Type -> loại dòng. BẮT BUỘC phân loại bằng cột này, TUYỆT ĐỐI không bằng
# dấu của Amount: chính file này có 2 dòng Doc.Type=K1 mang số DƯƠNG (trả hàng).
_CR_DOCTYPE_KIND = {
    "K1": "thanh_toan",   # hàng hóa — hóa đơn bán ra của mình (âm trong file)
    "D1": "phi",          # phí dịch vụ / phí hỗ trợ (dương)
    "KS": "chiet_khau",   # chiết khấu (dương)
}
# Ký hiệu chứng từ do BÊN CHUỖI phát hành, không phải hóa đơn bán ra của mình.
# Hai chuỗi này chỉ khác 1 ký tự và cùng tiền tố 'K..T' -> so 'K26T' là gộp nhầm.
#
# VÌ SAO là REGEX chứ không phải chuỗi 'K26TRT': hai số ở giữa là NĂM của ký hiệu
# hóa đơn điện tử. Đóng cứng '26' thì sang 01/2027 ký hiệu thành 'K27TRT', dòng
# TRẢ HÀNG rơi khỏi nhánh này và bị Doc.Type='K1' xếp thành 'thanh_toan' — sai
# HAI chiều (mất ghi giảm + cộng khống tiền thu), trong khi CẢ BA số kiểm tra
# (Overall Result và từng Result) vẫn khớp 0đ vì tổng NET không đổi. Không một
# cảnh báo nào bật, đúng ngày 01/01/2027. `\d?` cho ký hiệu thiếu số mẫu đầu (§E).
_CR_RETURN_RE = re.compile(r"^\d?K\d{2}TRT")   # phiếu trả hàng (Doc.Type vẫn K1 nhưng số DƯƠNG)
_CR_FEE_RE = re.compile(r"^\d?K\d{2}TEB")      # hóa đơn phí Central Retail phát hành


def parse_central_retail(sheets):
    """Bảng kê Central Retail: 1 sheet, header ở dòng 1, 12 cột do SAP xuất.

    BẪY TIỀN NẶNG NHẤT: 'Pmnt Date' KHÔNG phải ngày thanh toán mà là ngày ĐẾN HẠN
    của từng hóa đơn (24 giá trị khác nhau, có cả ngày SAU ngày trả thực tế). Ngày
    trả thật là 'Clearing Date'. Lấy nhầm là ghi sai kỳ và tạo bút toán ở tháng
    chưa tới.
    """
    warnings = []
    rows = []
    # Số kiểm tra gom theo TỪNG SHEET rồi mới cộng: file có thể bị cắt thành nhiều
    # trang in, mỗi trang một dòng tổng. `_declared_sum` trả None nếu có mảnh nào
    # đọc không ra, để check báo lệch thay vì âm thầm coi mảnh đó bằng 0.
    overall_parts = []
    by_doc_parts = {}
    n_table = 0
    skipped_sheets = []

    # §J.1 — CHỐT CHẶN: quét MỌI sheet, dò header TRÊN TỪNG SHEET.
    # VÌ SAO không đọc `sheets[0]`: file các chuỗi gửi là bản in PDF convert sang
    # Excel, mỗi TRANG IN thành một sheet nên bảng bị cắt rời (WinCommerce 9 sheet,
    # Co.op 9 sheet). Bỏ sót sheet là bỏ sót TIỀN, mà dòng tổng lại nằm ngay sheet
    # đầu nên `reconciled` vẫn báo True — sai tiền mà màn hình báo khớp 100%.
    # Dò header lại trên từng sheet (không tái dùng map cột của sheet trước) vì
    # các trang có thể LỆCH CỘT — §J.2, Table 7 của WinCommerce thừa một cột A.
    for name, grid in sheets:
        hdr_idx, hdr = None, []
        for i, row in enumerate(grid[:20]):
            cells = _row_texts(row)
            if "Reference" in cells and "Doc.Type" in cells and "Amount" in cells:
                hdr_idx, hdr = i, cells
                break
        if hdr_idx is None:
            # Sheet không có bảng: trống thật thì bỏ qua, còn dữ liệu thì KHÔNG
            # được đọc tiếp (xử lý sau vòng lặp) — đọc thiếu là mất tiền.
            if _sheet_has_content(grid):
                skipped_sheets.append(name)
            continue
        n_table += 1

        col = {n: (hdr.index(n) if n in hdr else None) for n in (
            "Account", "Clearing Doc.", "Terms of Pmnt", "Assignment", "Reference",
            "Doc.Type", "Doc. Date", "Entry Date", "Pmnt Date", "Text", "Clearing Date", "Amount")}
        for must in ("Reference", "Doc.Type", "Amount", "Clearing Date"):
            if col[must] is None:
                frappe.throw(_("Sheet {0}: file Central Retail thiếu cột bắt buộc: {1}")
                             .format(name, must))

        def cell(row, nm, _col=col):
            i = _col[nm]
            return None if i is None else (row[i] if i < len(row) else None)

        for n, row in enumerate(grid[hdr_idx + 1:], start=hdr_idx + 2):
            if _is_blank(row):
                continue
            account = _txt(cell(row, "Account"))
            terms = _txt(cell(row, "Terms of Pmnt"))
            doctype = _txt(cell(row, "Doc.Type"))
            amount = to_number(cell(row, "Amount"))
            ref = _txt(cell(row, "Reference"))
            assignment = _txt(cell(row, "Assignment"))
            text = re.sub(r"\s+", " ", _txt(cell(row, "Text")))  # Text phí có HAI dấu cách liền

            # --- DÒNG TỔNG: tách ra làm số kiểm tra, KHÔNG nạp ---
            if account == "Overall Result":
                overall_parts.append(amount)
                continue
            if terms == "Result":
                # MỖI Clearing Doc. có MỘT dòng Result — file thật có HAI, không phải một.
                # Bỏ theo vị trí dòng cứng (r3) sẽ cộng nhầm -94.485.536 vào chi tiết.
                by_doc_parts.setdefault(_txt(cell(row, "Clearing Doc.")), []).append(amount)
                continue

            series, no = split_invoice_ref(ref, REF_SEPS_CENTRAL_RETAIL)
            kind = _CR_DOCTYPE_KIND.get(doctype)
            needs_review = False

            if kind is None:
                # Doc.Type lạ: KHÔNG đoán và cũng KHÔNG nuốt. Đưa vào 'Khác' để người
                # thấy tiền, kèm cảnh báo — chôn nó đi là lệch tổng mà không ai biết vì sao.
                kind, sub, needs_review = "khac", "doctype_la:%s" % (doctype or "rong"), True
                warnings.append("Sheet %s dòng %d: Doc.Type lạ %r, đã xếp vào 'Khác' để người xử lý."
                                % (name, n, doctype))
            elif kind == "thanh_toan" and _CR_RETURN_RE.match(series.upper()):
                # Trả hàng: Doc.Type vẫn K1 nhưng là GHI GIẢM và số DƯƠNG. Phân loại chỉ
                # bằng Doc.Type sẽ cộng 5.119.605 vào tiền hàng (sai 2 lần con số này).
                kind, sub = "ghi_giam", "tra_hang"
            elif kind == "phi":
                sub = "phi_dich_vu" if _CR_FEE_RE.match(series.upper()) else "phi"
            elif kind == "chiet_khau":
                # CẢNH BÁO: dòng KS có thể mang ký hiệu hóa đơn bán ra của MÌNH
                # ('1C26THG|5656') mà KHÔNG phải thanh toán hóa đơn đó. Chỉ dòng
                # row_kind='thanh_toan' mới được phép khớp sang Sales Invoice.
                sub = "chiet_khau"
            else:
                sub = "hang_hoa"

            rows.append(_line(
                row_kind=kind, row_subtype=sub,
                inv_series=series, inv_no=no,
                inv_date=to_date(cell(row, "Doc. Date")),
                # File KHÔNG có mã siêu thị. Assignment chỉ là TÊN siêu thị ở dòng K1;
                # ở dòng D1/KS nó là mã tài khoản ('3003172.9999.') -> không phải cửa hàng.
                store_code=None,
                store_name=assignment if kind in ("thanh_toan", "ghi_giam") else None,
                doc_no=_txt(cell(row, "Clearing Doc.")),
                amount_before_vat=None, vat_amount=None,   # file KHÔNG tách VAT
                signed_amount=amount,
                # 'Clearing Date' mới là ngày chuỗi thực trả. KHÔNG dùng 'Pmnt Date'.
                payment_date=to_date(cell(row, "Clearing Date")),
                description=text or None,
                needs_review=needs_review or not series,
                source_sheet=name, source_row=n,
            ))

    if not n_table:
        frappe.throw(_("Không tìm thấy dòng tiêu đề Central Retail (cần Reference/Doc.Type/Amount)"))
    if skipped_sheets:
        # THROW chứ không warning: khuôn đã xác minh của Central Retail là MỘT bảng.
        # Sheet lạ còn dữ liệu nghĩa là tầng đọc chưa hiểu hết file — ghi nhận tiếp
        # là ghi nhận thiếu tiền trong khi mọi số kiểm tra vẫn khớp.
        frappe.throw(_("File Central Retail còn sheet có dữ liệu mà không nhận ra bảng: {0}. "
                       "Tầng đọc chưa xác minh khuôn này — KHÔNG ghi nhận để tránh bỏ sót tiền.")
                     .format(", ".join(skipped_sheets)))
    if n_table > 1:
        warnings.append("File Central Retail có %d sheet chứa bảng dữ liệu; đã đọc GỘP cả %d sheet. "
                        "Khuôn đã xác minh là 1 sheet — hãy soi kỹ số kiểm tra." % (n_table, n_table))

    declared_overall = _declared_sum(overall_parts)
    declared_by_doc = {k: _declared_sum(v) for k, v in by_doc_parts.items()}

    computed_all = _sum(rows, "signed_amount")
    checks = [_check("Overall Result", declared_overall, computed_all)]
    for doc, declared in declared_by_doc.items():
        got = _sum([r for r in rows if (r["doc_no"] or "") == doc], "signed_amount")
        checks.append(_check("Result %s" % doc, declared, got))

    dates = sorted({r["payment_date"] for r in rows if r["payment_date"]})
    if len(dates) > 1:
        warnings.append("File có %d ngày Clearing Date khác nhau — phải tách thành nhiều lần thanh toán."
                        % len(dates))

    groups = []
    for doc in sorted(declared_by_doc) or sorted({r["doc_no"] for r in rows if r["doc_no"]}):
        sub = [r for r in rows if (r["doc_no"] or "") == doc]
        groups.append({
            "key": doc, "advice_no": doc,
            "payment_date": (sub[0]["payment_date"] if sub else None),
            "n_rows": len(sub),
            # `computed_payment` phải CÙNG NGHĨA với `declared_payment` thì người đọc
            # mới so được. Dòng 'Result' của Central Retail là SỐ RÒNG của cả Clearing
            # Doc. (hàng hóa − phí − chiết khấu − trả hàng), nên ở đây cộng MỌI loại
            # dòng. Nếu chỉ cộng 'thanh_toan' thì bảng nhóm bày ra một khoản lệch
            # 141 triệu không có thật, trong khi đối chiếu ở `checks` vẫn khớp 0đ —
            # kế toán sẽ đi truy một cảnh báo giả.
            "computed_payment": _sum(sub, "signed_amount"),
            "declared_payment": declared_by_doc.get(doc),
            # Tiền hàng hóa thuần của Clearing Doc. này, để người xem đối chiếu công nợ.
            "computed_goods": _sum(sub, "signed_amount", {"thanh_toan"}),
        })

    return {
        # Nhiều Clearing Doc. trong một file -> không có MỘT số chứng từ duy nhất.
        "advice_no": (list(declared_by_doc)[0] if len(declared_by_doc) == 1 else None),
        "payment_dates": dates,
        "declared_totals": {
            "total_payment": None if declared_overall is None else abs(declared_overall),
            "total_discount": None,
            "overall_result": declared_overall,
            "by_clearing_doc": declared_by_doc,
        },
        "rows": rows, "checks": checks, "groups": groups, "warnings": warnings,
    }


# ═══════════════════════════════════════════════════════════════════════════
# 4c. LOTTE  (.xls BIFF, MỌI ô đều là TEXT — kể cả tiền và ngày)
# ═══════════════════════════════════════════════════════════════════════════

_LT_COLS = ["NO", "Store CD", "Store Name", "Vendor CD", "Vendor Name", "Write Date",
            "Invoice Date", "Tax No", "Invoice No", "Deduct Name", "Deduct Cause",
            "Pay Amt", "Vat Amt", "Total Amt", "Payment Date"]
_LT_REQUIRED = ("Store CD", "Tax No", "Invoice No", "Deduct Name", "Deduct Cause",
                "Pay Amt", "Vat Amt", "Total Amt", "Payment Date")

# Chỉ BA giá trị này của 'Deduct Name' là DANH MỤC khoản trừ. Mọi giá trị khác của
# cột đó là SỐ CHỨNG TỪ (vd '260626-01006-1-0265' của hàng trả lại) — whitelist
# theo 3 giá trị này rồi bỏ phần còn lại sẽ mất 2 dòng ghi giảm -809.335đ.
_LT_DEDUCT_KIND = {
    "Basic discount - Auto": ("chiet_khau", "Chiết khấu cơ bản"),
    "Sale services fee - Auto": ("phi", "Phí bán hàng"),
    "Other services fee - Auto": ("phi", "Phí dịch vụ khác"),
}
# LOTTE có HAI mức dòng tổng: SUB SUM (mỗi siêu thị) và SUM (toàn file). Chỉ lọc
# 'SUB SUM' thì dòng SUM 229.719.804 lọt vào dữ liệu và tổng file bị NHÂN ĐÔI.
_LT_TOTAL_CAUSES = {"SUB SUM", "SUM"}


def parse_lotte(sheets):
    """Bảng kê LOTTE: 1 sheet 'Payment deduct detail(1)', 135 dòng, toàn ô TEXT.

    BẪY LỚN NHẤT: 'Deduct Name rỗng = hàng hóa' là SAI. 6 dòng 'NET OFF REGULAR'
    cũng có Deduct Name rỗng nhưng KHÔNG có hóa đơn, và chúng có cả số dương lẫn
    số âm (-5.529.739). Điều kiện đúng là: Tax No VÀ Invoice No đều khác rỗng.
    """
    warnings = []
    rows = []
    sum_parts = {"pay": [], "vat": [], "total": []}
    n_subsum = 0
    n_table = 0
    skipped_sheets = []
    n_bad_date = 0
    # Cộng Pay/Vat theo ĐÚNG DẤU GỐC đọc từ file, ngay trong vòng lặp.
    # VÌ SAO không tái tạo dấu từ cột Total ở cuối: `_line()` lưu Pay/Vat dưới dạng
    # TRỊ TUYỆT ĐỐI, nên muốn cộng lại phải đoán dấu theo dấu của Total. Ở file này
    # ba cột luôn cùng dấu nên đoán đúng, nhưng chỉ cần một dòng LOTTE có Pay và Vat
    # ngược dấu nhau (điều chỉnh thuế) là hai số kiểm tra SUM — Pay Amt / Vat Amt
    # báo lệch trong khi dữ liệu đọc hoàn toàn đúng — báo động giả cho kế toán.
    tot_pay = tot_vat = 0.0

    # §J.1 — quét MỌI sheet. Tên sheet của file thật là 'Payment deduct detail(1)':
    # hậu tố '(1)' cho thấy bản xuất có PHÂN TRANG, kỳ thanh toán lớn sẽ ra thêm
    # '(2)', '(3)'. Đọc mỗi sheets[0] là bỏ nửa số tiền trong khi dòng SUM nằm ở
    # sheet đầu vẫn khớp -> `reconciled=True` trên một bảng kê thiếu tiền.
    for name, grid in sheets:
        hdr_idx = None
        for r, row in enumerate(grid[:20]):
            cells = _row_texts(row)
            if "Tax No" in cells and "Invoice No" in cells and "Total Amt" in cells:
                hdr_idx = r
                break
        if hdr_idx is None:
            if _sheet_has_content(grid):
                skipped_sheets.append(name)
            continue
        n_table += 1

        idx = {}
        for c, label in enumerate(_row_texts(grid[hdr_idx])):
            if label and label not in idx:
                idx[label] = c
        missing = [c for c in _LT_REQUIRED if c not in idx]
        if missing:
            frappe.throw(_("Sheet {0}: file LOTTE thiếu cột bắt buộc: {1}")
                         .format(name, ", ".join(missing)))

        def cell(row, label, _idx=idx):
            i = _idx.get(label)
            return None if i is None else (row[i] if i < len(row) else None)

        for r in range(hdr_idx + 1, len(grid)):
            row = grid[r]
            if _is_blank(row):
                continue
            rno = r + 1
            cause = _txt(cell(row, "Deduct Cause"))
            dname = _txt(cell(row, "Deduct Name"))
            series = _txt(cell(row, "Tax No"))
            inv_no = _txt(cell(row, "Invoice No"))
            pay = to_number(cell(row, "Pay Amt"))
            vat = to_number(cell(row, "Vat Amt"))
            total = to_number(cell(row, "Total Amt"))

            if cause in _LT_TOTAL_CAUSES:
                if cause == "SUM":
                    sum_parts["pay"].append(pay)
                    sum_parts["vat"].append(vat)
                    sum_parts["total"].append(total)
                else:
                    n_subsum += 1
                continue

            # --- phân loại: TUYỆT ĐỐI không dùng dấu của số tiền ---
            needs_review = False
            doc_no = None
            if dname in _LT_DEDUCT_KIND:
                kind, label = _LT_DEDUCT_KIND[dname]
            elif dname:
                # Deduct Name có giá trị nhưng không thuộc 3 danh mục -> đó là SỐ CHỨNG TỪ
                # ghi giảm của LOTTE (quan sát thật: Deduct Cause='Hang tra lai').
                #
                # VÌ SAO nhánh này phải đứng TRƯỚC nhánh 'có Tax No + Invoice No':
                # trả hàng luôn quy về một hóa đơn cụ thể, nên LOTTE hoàn toàn có thể
                # điền Tax No/Invoice No cho dòng trả hàng. Xét hóa đơn trước thì dòng
                # đó thành 'thanh_toan' mang số ÂM, rồi `_paid_subquery` lấy SUM(ABS())
                # lật thành TIỀN ĐÃ THU — một khoản trả hàng 809.335đ biến thành
                # 809.335đ đã thu, sai 1.618.670đ trên đúng hóa đơn đó, và cả 3 số
                # kiểm tra SUM vẫn khớp 0đ nên không có cảnh báo nào.
                kind, label, doc_no = "ghi_giam", (cause or "Ghi giảm theo chứng từ"), dname
            elif inv_no and series:
                # THAM CHIẾU HÓA ĐƠN LÀ DẤU HIỆU MẠNH NHẤT của dòng thanh toán, nên
                # nhánh này phải đứng TRƯỚC nhánh 'chỉ có Deduct Cause'.
                #
                # Đặt sau là một lỗi câm: chỉ cần LOTTE điền thêm Deduct Cause cho
                # dòng hàng hóa (ví dụ 'THANH TOAN HD 202607') là cả 45 dòng thật
                # rơi hết sang ghi_giam — total_payment về 0, 45 hóa đơn đã thu tiền
                # vẫn nằm rổ "chưa thanh toán", công nợ LOTTE bị thổi lên
                # 276.933.600đ. Mà vì tiền chỉ đổi NHÓM chứ không đổi TỔNG nên cả ba
                # số kiểm tra SUM vẫn khớp 0đ và `reconciled` vẫn báo True.
                kind, label = "thanh_toan", "Hóa đơn hàng hóa"
                if cause:
                    # Tổ hợp này CHƯA từng gặp trong dữ liệu thật (45/45 dòng hàng
                    # hóa đều trống Deduct Cause). Không tự quyết hộ: vẫn tính là
                    # thanh toán nhưng bắt người xem lại.
                    needs_review = True
                    label = cause
                    warnings.append(
                        "Sheet %s dòng %d: dòng có hóa đơn %s|%s NHƯNG kèm Deduct Cause %r "
                        "— chưa gặp bao giờ, đã xếp là thanh toán, cần người xác nhận"
                        % (name, rno, series, inv_no, cause))
            elif cause:
                # Còn Deduct Cause mà KHÔNG có hóa đơn thì đây là khoản cấn trừ có
                # lý do — không thể là dòng thanh toán vì không biết trả cho hóa đơn nào.
                kind, label, needs_review = "ghi_giam", cause, True
            else:
                # Không hóa đơn, không Deduct Name, không Deduct Cause, không phải dòng
                # tổng. Không có hóa đơn ⇒ không thể tự khớp, bắt buộc con người quyết.
                kind, label, needs_review = "ghi_giam", "Không rõ", True

            # Bất biến nội tại: Pay + Vat phải bằng Total. Lệch = đọc sai cột.
            # Không raise cả file vì một dòng lạ; nhưng phải nêu ra và đánh dấu dòng đó,
            # và dù sao dòng SUM ở dưới cũng sẽ tố cáo nếu cả cột bị đọc lệch.
            if pay is not None and vat is not None and total is not None and abs(pay + vat - total) > MONEY_EPS:
                needs_review = True
                warnings.append("Sheet %s dòng %d: Pay+Vat != Total (%s + %s != %s)"
                                % (name, rno, pay, vat, total))

            # Payment Date đọc không ra = SAI KỲ CÔNG NỢ. Phải đếm và báo, vì dòng
            # payment_date=NULL sẽ vô hình trên mọi màn hình (mọi truy vấn đều lọc
            # BETWEEN, mà NULL không bao giờ thỏa BETWEEN).
            pay_date_raw = cell(row, "Payment Date")
            pay_date = to_date(pay_date_raw)
            if pay_date is None:
                n_bad_date += 1
                if n_bad_date <= 5:
                    warnings.append("Sheet %s dòng %d: không đọc được Payment Date %r."
                                    % (name, rno, _txt(pay_date_raw)))

            tot_pay += pay or 0.0
            tot_vat += vat or 0.0
            rows.append(_line(
                row_kind=kind, row_subtype=label,
                inv_series=series, inv_no=inv_no,
                inv_date=to_date(cell(row, "Invoice Date")),
                # Store CD là TEXT '01019' — ép số là mất số 0 đầu và hỏng khớp siêu thị.
                store_code=_txt(cell(row, "Store CD")),
                store_name=_txt(cell(row, "Store Name")),
                doc_no=doc_no,
                amount_before_vat=pay, vat_amount=vat, signed_amount=total,
                payment_date=pay_date,
                description=(cause or label) if kind != "thanh_toan" else label,
                needs_review=needs_review,
                source_sheet=name, source_row=rno,
            ))

    if not n_table:
        frappe.throw(_("Không tìm thấy dòng tiêu đề LOTTE (cần Tax No/Invoice No/Total Amt)"))
    if skipped_sheets:
        frappe.throw(_("File LOTTE còn sheet có dữ liệu mà không nhận ra bảng: {0}. "
                       "Tầng đọc chưa xác minh khuôn này — KHÔNG ghi nhận để tránh bỏ sót tiền.")
                     .format(", ".join(skipped_sheets)))
    if n_bad_date > 5:
        warnings.append("... và %d dòng nữa không đọc được Payment Date." % (n_bad_date - 5))

    declared_sum = {k: _declared_sum(v) for k, v in sum_parts.items()}

    # Cộng theo DẤU GỐC của file để so với dòng SUM (LOTTE in cả 3 cột ở dòng SUM).
    tot_total = _sum(rows, "signed_amount")
    computed = {"pay": tot_pay, "vat": tot_vat, "total": tot_total}

    checks = [
        _check("SUM — Total Amt", declared_sum.get("total"), tot_total),
        _check("SUM — Pay Amt", declared_sum.get("pay"), tot_pay),
        _check("SUM — Vat Amt", declared_sum.get("vat"), tot_vat),
    ]

    dates = sorted({r["payment_date"] for r in rows if r["payment_date"]})
    if rows and not dates:
        # BÁO ĐỘNG, không được lặng lẽ trả reconciled=True. Không có ngày nào thì
        # `groups` rỗng -> `mt._split_advices` gộp CẢ FILE (2 kỳ) vào MỘT bản ghi
        # payment_date=NULL; bản ghi đó vô hình trên mọi màn hình MT (mọi truy vấn
        # lọc `payment_date BETWEEN`, NULL không bao giờ thỏa) nên kế toán tưởng
        # chưa nạp và nạp lại -> nhân đôi tiền. Dừng ở đây rẻ hơn nhiều.
        frappe.throw(_("Không đọc được Payment Date nào của LOTTE ({0} dòng) — KHÔNG ghi nhận. "
                       "Kiểm tra lại định dạng cột 'Payment Date' trong file gốc.")
                     .format(len(rows)))
    if len(dates) > 1:
        # §H: cùng file có 20260710 và 20260730 — không được coi là MỘT lần thanh toán.
        warnings.append("File có %d ngày thanh toán (%s) — phải tách thành nhiều lần thanh toán."
                        % (len(dates), ", ".join(dates)))
    if n_subsum:
        # SUB SUM TRỘN cả 2 payment date ở 11/17 nhóm -> không dùng làm tiền một lần trả.
        warnings.append("Đã bỏ %d dòng 'SUB SUM' (tổng theo siêu thị, trộn nhiều ngày thanh toán)."
                        % n_subsum)

    groups = []
    for d in dates:
        sub = [r for r in rows if r["payment_date"] == d]
        groups.append({
            "key": d, "advice_no": None, "payment_date": d, "n_rows": len(sub),
            "computed_payment": _sum(sub, "signed_amount"),
            "declared_payment": None,
        })

    return {
        "advice_no": None,   # LOTTE không in số chứng từ thanh toán trong bảng kê
        "payment_dates": dates,
        "declared_totals": {
            "total_payment": declared_sum.get("total"),
            "total_discount": None,
            "sum_pay_amt": declared_sum.get("pay"),
            "sum_vat_amt": declared_sum.get("vat"),
        },
        "rows": rows, "checks": checks, "groups": groups, "warnings": warnings,
        "_computed_extra": computed,
    }


# ═══════════════════════════════════════════════════════════════════════════
# 4d. Emart  (.xls BIFF, 1 sheet, bảng lệch 2 cột trống sang phải)
# ═══════════════════════════════════════════════════════════════════════════

# Document Type -> loại dòng. Emart để hàng hóa ÂM, chiết khấu/phí DƯƠNG — ngược
# hẳn LOTTE. Lấy dấu làm căn cứ là ghi ngược chiều tiền.
_EM_DOCTYPE_KIND = {"RE": "thanh_toan", "I0": "chiet_khau", "I1": "phi"}

_EM_H_DOC_NO = "Document Number"
_EM_H_DOC_TYPE = "Document Type"
_EM_H_INV_NO = "Invoice no"
_EM_H_DOC_DATE = "Document Date"
_EM_H_AMOUNT = "Amount in doc. curr."
_EM_H_TEXT = "Text"


def parse_emart(sheets):
    """Bảng kê Emart: 1 sheet, khối header nhà cung cấp ở trên, bảng bắt đầu r9.

    BẪY CHÍ MẠNG: nhãn dòng cộng 'phải trả tiền mua hàng' TRÙNG NGUYÊN VĂN với nội
    dung cột Text của 25/26 dòng hàng hóa thật. Lọc dòng rác bằng cách quét chuỗi
    trên toàn dòng sẽ XÓA SẠCH 26 dòng hàng hóa. Nhãn tổng nằm ở cột Document
    Number (khác cột với Text) và dòng tổng để TRỐNG Document Type -> chỉ nhận dòng
    có Document Type thuộc danh sách đã biết.
    """
    warnings = []
    rows = []
    declared_parts = {}
    pay_date, vendor_code, vendor_name = None, None, None
    n_table = 0
    skipped_sheets = []

    # §J.1 — quét MỌI sheet. Emart in 'Page: 1/1' ở khối header: kỳ nhiều dòng sẽ
    # thành '1/2', '2/2' và bản xuất tách thành nhiều sheet. Đọc mỗi sheets[0] là
    # bỏ nửa số tiền, mà 4 dòng cộng của Emart nằm ở sheet đầu nên vẫn khớp 0đ.
    for name, grid in sheets:
        hrow, cols = None, {}
        for r, row in enumerate(grid):
            c = {t: j for j, t in enumerate(_row_texts(row)) if t}
            if _EM_H_DOC_NO in c and _EM_H_DOC_TYPE in c and _EM_H_AMOUNT in c:
                hrow, cols = r, c
                break
        if hrow is None:
            if _sheet_has_content(grid):
                skipped_sheets.append(name)
            continue
        n_table += 1

        # --- ngày thanh toán nằm ở KHỐI HEADER, không nằm trong bảng ---
        # BẪY: cột 'Net due date' của các dòng RE tình cờ trùng ngày thanh toán
        # (15-09-2025) nhưng các dòng I0/I1 lại là 31-08-2025. Lấy 'Net due date' làm
        # ngày thanh toán là gán sai kỳ cho toàn bộ chiết khấu và phí.
        for row in grid[:hrow]:
            texts = _row_texts(row)
            for j, t in enumerate(texts):
                up = t.upper()
                nxt = next((x for x in texts[j + 1:] if x), None)
                if up.startswith("PAYMENT DATE") and pay_date is None:
                    pay_date = to_date(nxt)
                elif up.startswith("VENDOR CODE") and vendor_code is None:
                    vendor_code = nxt
                elif up.startswith("VENDOR NAME") and vendor_name is None:
                    vendor_name = nxt

        def cell(row, key, _cols=cols):
            j = _cols.get(key)
            return row[j] if j is not None and j < len(row) else None

        for r in range(hrow + 1, len(grid)):
            row = grid[r]
            doc_type = _txt(cell(row, _EM_H_DOC_TYPE)).upper()
            amount = to_number(cell(row, _EM_H_AMOUNT))
            label = _txt(cell(row, _EM_H_DOC_NO))

            if not doc_type:
                # Dòng cộng nhóm ('chiết khấu', 'phí hỗ trợ', 'phải trả tiền mua hàng')
                # và dòng 'TOTAL'. File có BỐN dòng tổng chứ không phải hai — bỏ sót
                # dòng 'phải trả tiền mua hàng' là cộng dư -191.554.740 (nhân đôi hàng hóa).
                if label and amount is not None:
                    declared_parts.setdefault(label, []).append(amount)
                continue

            kind = _EM_DOCTYPE_KIND.get(doc_type)
            if kind is None:
                kind = "khac"
                warnings.append("Sheet %s dòng %d: Document Type lạ %r, đã xếp vào 'Khác' để người xử lý."
                                % (name, r + 1, doc_type))
            if amount is None:
                warnings.append("Sheet %s dòng %d: có Document Type %r nhưng không đọc được số tiền."
                                % (name, r + 1, doc_type))
                continue

            inv_no = _txt(cell(row, _EM_H_INV_NO))
            rows.append(_line(
                row_kind=kind, row_subtype=doc_type,
                # Emart KHÔNG CẤP KÝ HIỆU hóa đơn -> để rỗng, và LUÔN cần review khi
                # khớp (khớp bằng số + ngày + tiền theo §F). Đoán ký hiệu = gán nhầm HĐ.
                inv_series="", inv_no=inv_no,
                # 'Document Date' là ngày hóa đơn NCC; 'Posting Date' là ngày Emart ghi
                # sổ — hai ngày khác nhau ở hầu hết dòng RE.
                inv_date=to_date(cell(row, _EM_H_DOC_DATE)),
                # VENDOR CODE là mã NHÀ CUNG CẤP trong hệ thống Emart, KHÔNG phải mã
                # cửa hàng. Nhồi vào store_code là sai nghĩa. File không tách theo siêu thị.
                store_code=None, store_name=None,
                doc_no=_txt(cell(row, _EM_H_DOC_NO)),
                amount_before_vat=None, vat_amount=None,   # file chỉ có MỘT cột tiền
                signed_amount=amount,
                payment_date=pay_date,
                description=_txt(cell(row, _EM_H_TEXT)) or None,
                needs_review=True,
                source_sheet=name, source_row=r + 1,
            ))

    if not n_table:
        frappe.throw(_("Không tìm thấy dòng tiêu đề Emart (cần Document Number/Type/Amount)"))
    if skipped_sheets:
        frappe.throw(_("File Emart còn sheet có dữ liệu mà không nhận ra bảng: {0}. "
                       "Tầng đọc chưa xác minh khuôn này — KHÔNG ghi nhận để tránh bỏ sót tiền.")
                     .format(", ".join(skipped_sheets)))
    if not pay_date:
        # THROW chứ không warning: `payment_date=None` chảy thẳng vào MT Payment Advice
        # (field không reqd), và MỌI màn hình MT lọc `payment_date BETWEEN` — NULL không
        # bao giờ thỏa BETWEEN, nên cả phiếu lẫn 48 dòng tiền BIẾN MẤT khỏi giao diện.
        # Kế toán tưởng chưa nạp, nạp lại -> nhân đôi tiền. Không có ngày thì không ghi.
        frappe.throw(_("Không đọc được 'PAYMENT DATE' ở khối header Emart — KHÔNG ghi nhận. "
                       "Kiểm tra lại khối tiêu đề của file gốc."))
    # Gán lại cho MỌI dòng: nếu file bị cắt thành nhiều sheet, khối tiêu đề chỉ in
    # ở sheet có PAYMENT DATE, các sheet còn lại sẽ nhận None ngay lúc dựng dòng.
    # Emart chỉ có MỘT ngày thanh toán cho cả bảng kê nên gán đồng loạt là đúng;
    # để sót dòng payment_date=None là để dòng đó vô hình trên mọi màn hình MT.
    for r in rows:
        r["payment_date"] = pay_date

    # Gộp các dòng cộng của mọi sheet. `_declared_sum` trả None nếu có mảnh đọc
    # không ra, để check báo lệch thay vì coi mảnh đó bằng 0.
    declared_groups = {k: _declared_sum(v) for k, v in declared_parts.items()}

    # Đối chiếu với các dòng cộng do chính Emart in ra. Ghép nhãn -> loại dòng
    # bằng chuỗi đã bỏ dấu để không phụ thuộc 'phí' hay 'phi'.
    net = _sum(rows, "signed_amount")
    checks = []
    for lbl, declared in declared_groups.items():
        s = strip_tones(lbl)
        if s == "total":
            checks.append(_check("TOTAL", declared, net))
        elif "chiet khau" in s:
            checks.append(_check("Cộng chiết khấu", declared, _sum(rows, "signed_amount", {"chiet_khau"})))
        elif "phi ho tro" in s or "phi" in s:
            checks.append(_check("Cộng phí", declared, _sum(rows, "signed_amount", {"phi"})))
        elif "phai tra tien mua hang" in s:
            checks.append(_check("Cộng phải trả tiền mua hàng", declared,
                                 _sum(rows, "signed_amount", {"thanh_toan"})))

    total_row = declared_groups.get("TOTAL")
    # Dòng cộng 'phải trả tiền mua hàng' chính là tổng các dòng RE do Emart in ra —
    # dùng làm số kiểm tra của nhóm thay vì để None. Ghép nhãn bằng chuỗi đã bỏ dấu
    # vì Emart viết lẫn có dấu / không dấu.
    declared_goods = next((v for k, v in declared_groups.items()
                           if "phai tra tien mua hang" in strip_tones(k)), None)
    groups = [{
        "key": pay_date, "advice_no": None, "payment_date": pay_date, "n_rows": len(rows),
        "computed_payment": _sum(rows, "signed_amount", {"thanh_toan"}),
        "declared_payment": declared_goods,
        "computed_net": _sum(rows, "signed_amount"),
        "declared_net": total_row,
    }]

    return {
        "advice_no": None,   # Emart không in số chứng từ thanh toán
        "payment_dates": [pay_date] if pay_date else [],
        "declared_totals": {
            "total_payment": None if total_row is None else abs(total_row),
            "total_discount": None,
            "by_label": declared_groups,
            "vendor_code": vendor_code, "vendor_name": vendor_name,
        },
        "rows": rows, "checks": checks, "groups": groups, "warnings": warnings,
    }


# ═══════════════════════════════════════════════════════════════════════════
# 4e. Saigon Co.op  (.xlsx, MỖI SHEET LÀ MỘT LẦN THANH TOÁN RIÊNG)
# ═══════════════════════════════════════════════════════════════════════════

# Ký hiệu hóa đơn điện tử VN: [số mẫu]{C|K}{2 số năm}T{2-3 chữ}.
# C = hóa đơn bán hàng, K = hóa đơn điều chỉnh/trả hàng.
# Lookaround để KHÔNG nuốt nhầm khi ký hiệu dính liền chữ khác ('1K26THLTPCN').
# \d? optional vì có dòng thiếu hẳn số mẫu đầu ('K26TEK-80') — §E.
_COOP_SERIES_RE = re.compile(r"(?<![0-9A-Z])(\d?[CK]\d{2}T[A-Z]{2,3})(?![A-Z])")
_COOP_PAYDATE_RE = re.compile(r"Ngày thanh toán\s*:\s*(\d{2})/(\d{2})/(\d{4})")
_COOP_DOCNO_RE = re.compile(r"Số chứng từ\s*:\s*(\S+)")
_COOP_STORE_RE = re.compile(r"^(\d+)-SIPI")

# A B C D E F G H I L (1-based) — vị trí cố định, nhưng khuôn header được KIỂM TRA
# trước khi đọc nên nếu Co.op đổi mẫu thì dừng chứ không đọc sai âm thầm.
_COOP_COL = {"stt": 1, "doc_no": 2, "inv_no": 3, "inv_date": 4, "desc": 5,
             "value": 6, "disc_rate": 7, "disc_amt": 8, "pay": 9, "store": 12}


def _coop_clean_inv_no(raw):
    """`'P0007272` -> `7272`.

    Co.op ghi số hóa đơn NCC kèm DẤU NHÁY ĐƠN LÀ KÝ TỰ THẬT trong ô text (569/569
    ô), kèm tiền tố chữ của quyển (P/A/B/Z/DCG..., có cả CHỮ THƯỜNG 'a0006911') và
    số 0 độn. Bỏ nháy -> bỏ mọi ký tự đầu không phải chữ số -> bỏ số 0 thừa.

    TUYỆT ĐỐI không lấy số trong DIỄN GIẢI: 12/569 dòng con số ở đó là SỐ PHIẾU RTV,
    không phải số hóa đơn (vd cột C='A00000860' nhưng DIỄN GIẢI ghi 'RTV2153190').
    """
    s = _txt(raw).lstrip("'").strip()
    s = re.sub(r"^[^0-9]+", "", s).lstrip("0")
    return s or None


def parse_coop(sheets):
    """Bảng kê Saigon Co.op: 8 sheet dữ liệu, MỖI SHEET = MỘT lần thanh toán riêng.

    Ba bẫy tiền nặng nhất, đều đã đo trên file thật:
      1. Đọc mỗi sheet đầu là bỏ sót 469/569 dòng (~4,83 tỷ đồng).
      2. Cột THANH TOÁN/TIỀN chỉ điền ở DÒNG ĐẦU mỗi nhóm siêu thị và là tổng CẢ
         NHÓM. Forward-fill xuống mọi dòng rồi cộng ⇒ 13.843.267.903đ thay vì
         6.200.078.656đ — THỔI PHỒNG 2,23 LẦN.
      3. Không tự tính lại tiền: Co.op làm tròn chiết khấu ở cấp DÒNG và tiền thanh
         toán ở cấp NHÓM độc lập nhau (17/374 nhóm lệch ±1đ). Số Co.op ghi là chân lý.
    """
    warnings = []
    rows, groups, checks = [], [], []
    declared_gross = declared_disc = declared_pay = 0.0
    n_sheet = 0

    for name, grid in sheets:
        # Sheet mồi (Sheet3 rỗng). Không duyệt theo TÊN sheet: thứ tự vật lý trong
        # workbook là Sheet1, Sheet3, Sheet2, Sheet4... — tên không theo thời gian.
        if len(grid) < 20:
            continue

        # Khuôn header 2 tầng r18/r19. Sai khuôn thì DỪNG, không đọc sai âm thầm
        # rồi hạch toán nhầm tiền.
        if (_txt(_g(grid, 18, 1)) != "STT" or _txt(_g(grid, 19, 5)) != "DIỄN GIẢI"
                or _txt(_g(grid, 19, 6)) != "TRỊ GIÁ" or _txt(_g(grid, 19, 9)) != "TIỀN"):
            frappe.throw(_("Sheet {0}: header r18/r19 không đúng khuôn Co.op").format(name))
        n_sheet += 1

        # --- meta của LẦN THANH TOÁN này (mỗi sheet một bộ riêng) ---
        m = _COOP_PAYDATE_RE.search(_txt(_g(grid, 7, 9)))
        pay_date = "%s-%s-%s" % (m.group(3), m.group(2), m.group(1)) if m else None
        m = _COOP_DOCNO_RE.search(_txt(_g(grid, 6, 9)))
        payment_doc = m.group(1) if m else None
        if not pay_date:
            warnings.append("Sheet %s: không đọc được 'Ngày thanh toán' ở ô I7." % name)

        chk_gross = to_number(_g(grid, 13, 13))   # Tổng Tiền
        chk_disc = to_number(_g(grid, 14, 13))    # Tổng Giá Trị Chiết Khấu
        chk_pay = to_number(_g(grid, 15, 13))     # Tổng Tiền Thanh Toán

        # --- dòng "Tổng: ..." đóng vùng dữ liệu ---
        end = None
        for r in range(20, len(grid) + 1):
            v = _txt(_g(grid, r, _COOP_COL["inv_no"]))
            if v.startswith("Tổng"):
                end = r
                break
        if end is None:
            frappe.throw(_("Sheet {0}: không tìm thấy dòng Tổng").format(name))

        g_doc = g_store = g_code = None
        g_pay_sum = 0.0
        g_key = 0
        n_before = len(rows)

        for r in range(20, end):
            value = to_number(_g(grid, r, _COOP_COL["value"]))
            disc = to_number(_g(grid, r, _COOP_COL["disc_amt"]))
            pay = to_number(_g(grid, r, _COOP_COL["pay"]))
            desc = _txt(_g(grid, r, _COOP_COL["desc"]))
            raw_inv = _g(grid, r, _COOP_COL["inv_no"])
            raw_doc = _g(grid, r, _COOP_COL["doc_no"])

            # Ô cột TIỀN có giá trị = MỞ NHÓM MỚI (ô merge cấp nhóm). Cột SỐ HÓA ĐƠN
            # LH và COOP cũng chỉ điền ở dòng đầu nhóm -> forward-fill. Nhưng ĐỪNG
            # forward-fill cột TIỀN rồi cộng (xem docstring, bẫy #2).
            if pay is not None:
                g_key += 1
                g_pay_sum += pay
                g_doc = _txt(raw_doc) or None
                g_store = _txt(_g(grid, r, _COOP_COL["store"])) or None
                mm = _COOP_STORE_RE.match(_txt(raw_doc))
                # Tiền tố số của 'SỐ HÓA ĐƠN LH' ánh xạ 1:1 sang tên siêu thị ở cột
                # COOP (120 tiền tố, không cái nào ứng 2 tên) nên dùng làm store_code.
                # Ý nghĩa nghiệp vụ của mã này VẪN CHƯA được kế toán Co.op xác nhận (§I).
                g_code = mm.group(1) if mm else None

            m = _COOP_SERIES_RE.search(desc)
            series = m.group(1) if m else None
            # Phân loại theo KÝ HIỆU, không theo dấu tiền (§B). K = trả hàng/điều chỉnh.
            if series and "K" in series[:3]:
                kind, needs_review = "ghi_giam", False
            elif series:
                kind, needs_review = "thanh_toan", False
            else:
                # 17 dòng DIỄN GIẢI chỉ ghi số phiếu RTV / ghi chú điều chỉnh, không
                # có ký hiệu. Đều là ghi giảm, nhưng KHÔNG đoán ký hiệu — để người quyết.
                kind, needs_review = "ghi_giam", True

            inv_no = _coop_clean_inv_no(raw_inv)
            if inv_no is None:
                # 7/569 dòng cột HÓA ĐƠN NCC rỗng (chỉ còn dấu nháy) — số hóa đơn
                # chuyển sang cột SỐ HÓA ĐƠN LH của CHÍNH DÒNG đó.
                inv_no = _coop_clean_inv_no(raw_doc)
                needs_review = True
            if inv_no and not inv_no.isdigit():
                # Sau khi bóc nháy + tiền tố quyển, số hóa đơn NCC của Co.op phải là
                # SỐ THUẦN. Còn ký tự lạ nghĩa là ô ghi kiểu khác — có thật: Sheet6
                # r139 ô C = "'26-A532" (1.541.700đ), r96 ra '1K26TCH-106' vốn là KÝ
                # HIỆU chứ không phải số. `_coop_clean_inv_no` chỉ bóc tiền tố KHÔNG
                # PHẢI CHỮ SỐ ở đầu nên chuỗi bắt đầu bằng chữ số đi qua nguyên vẹn,
                # rồi `by_key` không tra được và dòng tiền im lặng không khớp. Bật cờ
                # để nó vào danh sách 'n dòng cần người xem lại' thay vì biến mất.
                needs_review = True

            base = dict(
                inv_series=series or "", inv_no=inv_no or "",
                inv_date=to_date(_g(grid, r, _COOP_COL["inv_date"])),
                store_code=g_code, store_name=g_store, doc_no=g_doc,
                payment_date=pay_date, source_sheet=name, source_row=r,
            )
            rows.append(_line(row_kind=kind, row_subtype="hang_hoa" if kind == "thanh_toan" else "dieu_chinh",
                              signed_amount=value, description=desc or None,
                              needs_review=needs_review, **base))
            if disc:
                # Chiết khấu Co.op là CỘT, không phải dòng — tách thành dòng riêng để
                # tổng chiết khấu của phiếu cộng được, KHÔNG tự tính lại từ tỉ lệ
                # ('17.75%-VAT 8' chỉ là nhãn text; 68/569 dòng lệch ±1đ so với round()).
                rows.append(_line(row_kind="chiet_khau", row_subtype="chiet_khau_cot_rieng",
                                  signed_amount=disc,
                                  description="Chiết khấu %s" % _txt(_g(grid, r, _COOP_COL["disc_rate"])),
                                  needs_review=False, **base))

        sub = rows[n_before:]
        got_gross = _sum(sub, "signed_amount", {"thanh_toan", "ghi_giam"})
        got_disc = _sum(sub, "signed_amount", {"chiet_khau"})
        checks.append(_check("%s — Tổng Tiền" % name, chk_gross, got_gross))
        checks.append(_check("%s — Tổng Giá Trị Chiết Khấu" % name, chk_disc, got_disc))
        checks.append(_check("%s — Tổng Tiền Thanh Toán" % name, chk_pay, g_pay_sum))
        # KHÔNG kiểm 'Tổng Tiền − Chiết khấu == Thanh toán': Co.op làm tròn hai cấp
        # độc lập nên quan hệ này lệch ±1..3đ mỗi sheet mà số liệu vẫn ĐÚNG.

        for k, v in (("gross", chk_gross), ("disc", chk_disc), ("pay", chk_pay)):
            if v is None:
                warnings.append("Sheet %s: thiếu số kiểm tra %s." % (name, k))
        declared_gross += chk_gross or 0.0
        declared_disc += chk_disc or 0.0
        declared_pay += chk_pay or 0.0

        groups.append({
            "key": name, "advice_no": payment_doc, "payment_date": pay_date,
            "n_rows": len(sub), "n_store_groups": g_key,
            # `computed_payment` = tổng ô TIỀN ở DÒNG ĐẦU mỗi nhóm siêu thị — cùng
            # nghĩa với `declared_payment` (ô M15 'Tổng Tiền Thanh Toán') nên so được.
            # KHÔNG lấy `got_gross` (tổng TRỊ GIÁ) đặt cạnh chk_pay: hai số đó lệch
            # nhau đúng bằng chiết khấu, bày ra như một sai lệch là báo động giả.
            "computed_payment": g_pay_sum,
            # Số Co.op THỰC TRẢ cho nhóm — đọc từ file, không tự tính lại.
            "declared_payment": chk_pay,
            "computed_gross": got_gross, "computed_discount": got_disc,
            "declared_gross": chk_gross, "declared_discount": chk_disc,
        })

    # --- Co.op ĐÒI LẠI tiền của một hóa đơn đã trả ---
    # Có thật trong file mẫu: Sheet6 r75 là dòng thanh toán +3.121.200đ cho hóa đơn
    # 3176 (kỳ 25/05), Sheet7 r21 là dòng ghi giảm −3.121.200đ CŨNG hóa đơn 3176
    # (kỳ 22/06, 'DCHINH DO NCC THAY BANG HD 3542'). Dòng ghi_giam KHÔNG được phép
    # nối Sales Invoice, nên cột `paid` chỉ thấy vế +3.121.200 và hóa đơn hiện 'đã
    # thanh toán đủ' dù tiền đã bị lấy lại — mà mọi số kiểm tra của Co.op vẫn khớp
    # 0đ nên không có tín hiệu nào. Ở đây chỉ ĐÁNH DẤU + BÀY RA cho người xử lý;
    # việc trừ ngược vào công nợ là quyết định của con người, không tự làm.
    paid_by_no = {}
    for r in rows:
        if r["row_kind"] == "thanh_toan" and r["inv_no_norm"]:
            paid_by_no.setdefault(r["inv_no_norm"], []).append(r)
    reversals = []
    for r in rows:
        if r["row_kind"] != "ghi_giam" or not r["inv_no_norm"]:
            continue
        # Chỉ đối chiếu khi dòng ghi giảm KHÔNG có ký hiệu riêng (số ở cột HÓA ĐƠN
        # NCC chính là hóa đơn gốc), hoặc trùng luôn ký hiệu. Dòng ghi giảm mang ký
        # hiệu 'K...' riêng là MỘT CHỨNG TỪ KHÁC có dãy số riêng — trùng số với một
        # hóa đơn bán là ngẫu nhiên, gắn cờ ở đó chỉ tạo báo động giả.
        hits = [p for p in paid_by_no.get(r["inv_no_norm"], [])
                if not r["inv_series_norm"] or p["inv_series_norm"] == r["inv_series_norm"]]
        if not hits:
            continue
        r["needs_review"] = True
        reversals.append("HĐ %s (%s dòng %s, %s: %s)" % (
            r["inv_no"], r["source_sheet"], r["source_row"], r["payment_date"] or "?",
            "{:,.0f}đ".format(r["signed_amount"] or 0)))
    if reversals:
        warnings.append(
            "%d khoản ghi giảm TRÙNG SỐ HÓA ĐƠN với dòng đã ghi nhận thanh toán trong "
            "cùng file — Co.op đòi lại tiền đã trả, hóa đơn đó vẫn đang hiện 'đã thanh "
            "toán đủ'. Phải xử lý tay: %s" % (len(reversals), "; ".join(reversals[:10])
                                              + (" ..." if len(reversals) > 10 else "")))

    dates = sorted({g["payment_date"] for g in groups if g["payment_date"]})
    # BẮT BUỘC nói rõ: Co.op ghi chiết khấu CÙNG DẤU với trị giá (cả hai đều dương ở
    # dòng bán), nên `computed_totals.net` của chuỗi này = trị giá + chiết khấu, KHÔNG
    # phải tiền thực trả. Số thực trả là 'Tổng Tiền Thanh Toán' do Co.op in ra. Không
    # cảnh báo thì rất dễ có người lấy nhầm 8,88 tỷ thay cho 6,20 tỷ.
    if declared_pay:
        warnings.append(
            "Với Co.op, tiền THỰC TRẢ là 'Tổng Tiền Thanh Toán' = {:,.0f}đ (đọc từ file). "
            "Tổng có dấu của các dòng KHÔNG phải tiền thực trả vì chiết khấu ghi cùng dấu "
            "với trị giá.".format(declared_pay))
    if n_sheet > 1:
        warnings.append("File Co.op có %d sheet = %d LẦN THANH TOÁN riêng (số chứng từ + ngày riêng). "
                        "Không gộp thành một phiếu." % (n_sheet, n_sheet))

    return {
        "advice_no": groups[0]["advice_no"] if len(groups) == 1 else None,
        "payment_dates": dates,
        "declared_totals": {
            "total_payment": declared_pay or None,
            "total_discount": declared_disc or None,
            "total_gross": declared_gross or None,
        },
        "rows": rows, "checks": checks, "groups": groups, "warnings": warnings,
    }


# ═══════════════════════════════════════════════════════════════════════════
# 5. Đầu vào chung
# ═══════════════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════════════
# 4f. AEON  (.xls BIFF, 6 sheet — khối header LẶP y hệt ở mọi sheet)
# ═══════════════════════════════════════════════════════════════════════════

# SLIP TYPE -> loại dòng. Phân loại theo CỘT, tuyệt đối không theo dấu tiền:
# đo trên file thật, dòng 312 (hàng trả) lưu số DƯƠNG trong bảng chi tiết nhưng
# khối tổng lại in ÂM (-2.545.560). Lấy dấu làm căn cứ là ghi ngược chiều tiền.
_AE_SLIP_KIND = {"311": ("thanh_toan", "PO"), "312": ("ghi_giam", "GRN")}

_AE_H_DOC = "SLIP TYPE"          # nhãn header bảng chi tiết sheet Doc
_AE_H_COST = "DEDUCTION CODE"    # nhãn header bảng chi tiết sheet Costdet


def _ae_norm(label) -> str:
    """Nhãn cột -> khóa so khớp: bỏ dấu, bỏ mọi ký tự không phải chữ/số.

    VÌ SAO không so nguyên văn: AEON in 'SUPPLIER INVOICE / CN NO.' (CÓ dấu chấm
    cuối, CÓ dấu gạch chéo) ở sheet Doc nhưng 'SLIP NO' (KHÔNG chấm) ở Costdet —
    cùng một file, hai quy ước. So nguyên văn là hụt đúng cột số hóa đơn và cột
    tiền, parser gãy ngay trên file mẫu thật.
    """
    return re.sub(r"[^a-z0-9]+", "", strip_tones(label))


def _ae_sheet(sheets, prefix):
    """Sheet theo TIỀN TỐ tên: tên sheet mang mã lần thanh toán `Doc(00E30_265294)`.

    So khớp đầy đủ tên là hỏng ngay kỳ sau (mã đổi); so tiền tố là đúng bản chất.
    """
    for name, grid in sheets:
        if _ae_norm(name).startswith(_ae_norm(prefix)):
            return name, grid
    return None, None


def _ae_header_meta(grid):
    """Đọc khối header r1–r11. Khối này LẶP y hệt ở cả 6 sheet nên đọc ở đâu cũng được.

    Ngoại lệ: sheet Summary KHÔNG có 'PAYMENT NO:' / 'CREDIT TERM:' — vì vậy phải
    đọc meta trên sheet Doc, đừng đọc trên Summary rồi tưởng file thiếu số bảng kê.
    """
    meta = {}
    for r in range(1, min(12, len(grid) + 1)):
        texts = _row_texts(grid[r - 1])
        for i, t in enumerate(texts):
            key = _ae_norm(t.rstrip(":"))
            if not key:
                continue
            val = next((texts[j] for j in range(i + 1, len(texts)) if texts[j]), "")
            if not val:
                continue
            if key == "paymentdate":
                meta["payment_date"] = to_date(val)
            elif key == "paymentno":
                meta["advice_no"] = val
            elif key == "supplier":          # 'SUPPLIER NAME' chuẩn hóa ra khóa KHÁC
                meta["vendor_code"] = val
            elif key == "creditterm":
                meta["credit_term"] = val
    return meta


def _ae_find_header(grid, label, start=1):
    """Dòng header theo NHÃN, không hardcode chỉ số cột. Trả (r 1-based, {khóa: cột}).

    `start` để tìm LẦN XUẤT HIỆN THỨ HAI của cùng một nhãn — sheet Doc dùng đúng
    nhãn 'Slip Type' cho cả bảng chi tiết (r13) lẫn khối tổng cuối bảng (r46).
    """
    tgt = _ae_norm(label)
    for r in range(start, len(grid) + 1):
        texts = _row_texts(grid[r - 1])
        if any(_ae_norm(t) == tgt for t in texts):
            return r, {_ae_norm(t): c for c, t in enumerate(texts, start=1) if t}
    return None, {}


def _ae_col(cols, *names):
    """Cột đầu tiên tìm được trong số các tên nhãn ứng viên. Không có -> None."""
    for n in names:
        c = cols.get(_ae_norm(n))
        if c:
            return c
    return None


def _gv(grid, r, c):
    """`_g` nhưng chấp nhận cột None (nhãn không có trong file) -> trả None.

    `_g` so sánh `c < 1` nên cột None làm nổ TypeError. Cột thiếu là chuyện BÌNH
    THƯỜNG (mẫu file đổi, bớt một cột mô tả) và phải suy biến thành "không có dữ
    liệu", chứ không phải làm hỏng cả lần nạp.
    """
    return _g(grid, r, c) if c else None


def _ae_split_invoice(raw):
    """'1-C26THG-00004246' -> ('C26THG', '00004246').

    Bỏ cụm ĐẦU khi nó là số mẫu hóa đơn (đúng 1 chữ số). KHÔNG dùng
    `split_invoice_ref`: chuỗi này có HAI dấu '-' chứ không phải một, mà hàm kia
    cắt tại dấu ĐẦU TIÊN nên sẽ trả ký hiệu '1' và số 'C26THG-00004246'.
    Không đúng khuôn -> trả ký hiệu RỖNG và đưa cả chuỗi vào số: thà không khớp
    còn hơn bịa ký hiệu rồi ghi tiền vào hóa đơn của người khác.
    """
    t = _txt(raw)
    if not t:
        return "", ""
    parts = [p.strip() for p in t.split("-") if p.strip()]
    if len(parts) == 3 and parts[0].isdigit() and len(parts[0]) == 1:
        return parts[1], parts[2]
    if len(parts) == 2:
        return parts[0], parts[1]
    return "", t


def parse_aeon(sheets):
    """Bảng kê thanh toán AEON Việt Nam.

    Sáu sheet: Summary · Doc · Costsumm · Rebsumm · Costdet · DcCharges.
    Chỉ `Doc` (hàng bán + hàng trả) và `Costdet` (khoản trừ) SINH DÒNG TIỀN.

    BẪY 1 — KHỐI TỔNG NẰM TRONG CHÍNH BẢNG. Cuối sheet Doc có khối tổng mở đầu
    bằng ĐÚNG nhãn 'Slip Type' và các dòng của nó cũng mang mã 311/312. Đọc tiếp
    qua đó là cộng đôi: đo thật ra 123.768.000 thay vì 61.884.000.

    BẪY 2 — CỘNG TRÙNG DcCharges. `DcCharges` KHÔNG phải khoản trừ riêng, nó là
    CHI TIẾT của đúng một dòng mã 'DC' trong `Costdet`; tổng hai bên bằng nhau
    tuyệt đối (2.512.222). Sinh dòng tiền từ cả hai sheet là cộng trùng đúng số
    đó mà 'Net Payment' VẪN khớp, vì Net Payment chỉ đối chiếu với Costdet. Vì
    vậy DcCharges chỉ dùng làm CHỐT ĐỐI CHIẾU cho riêng dòng 'DC'.

    BẪY 3 — KHOẢN TRỪ CÓ DÒNG ÂM. Trong Costdet, mã RBGPA/RBGPD/RBGPOS/RBPS có
    cả dòng âm (-8.683) lẫn dương, và Sub-Total là tổng ĐẠI SỐ của chúng. Lấy
    `-abs(amt)` là đảo dấu 8 dòng hoàn tiền thành 8 dòng bị trừ — lệch đúng hai
    lần số đó. Phải lấy `-amt`, giữ nguyên độ lớn lẫn chiều.

    `Costsumm`/`Rebsumm` là DANH MỤC mã khoản trừ (gần như toàn 0) — không sinh
    dòng, chỉ dùng soát: mã có giá trị khác 0 mà không thấy trong Costdet -> cảnh báo.
    """
    warnings, rows, checks = [], [], []

    doc_name, doc = _ae_sheet(sheets, "Doc")
    if not doc:
        frappe.throw(_("File AEON thiếu sheet 'Doc' — không đọc được hóa đơn nào"))

    meta = _ae_header_meta(doc)
    pay_date = meta.get("payment_date")
    advice_no = meta.get("advice_no")
    if not pay_date:
        warnings.append("Không đọc được 'PAYMENT DATE' ở khối header — mọi dòng sẽ "
                        "thiếu ngày thanh toán. Kế toán phải điền tay trước khi nạp.")

    # ── Sheet Doc: hàng bán (311) + hàng trả (312) ──────────────────────────
    hr, hcols = _ae_find_header(doc, _AE_H_DOC)
    if not hr:
        frappe.throw(_("Sheet Doc của AEON không có dòng nhãn '{0}' — mẫu file đã đổi")
                     .format(_AE_H_DOC))
    c_slip = _ae_col(hcols, _AE_H_DOC)
    c_inv = _ae_col(hcols, "SUPPLIER INVOICE / CN NO.")
    c_amt = _ae_col(hcols, "AMOUNT")
    c_store = _ae_col(hcols, "STORE CODE")
    c_date = _ae_col(hcols, "DELIVERY / RETURN DATE")
    c_contract = _ae_col(hcols, "CONTRACT NO.")
    c_slipno = _ae_col(hcols, "SLIP NO.")
    c_remarks = _ae_col(hcols, "REMARKS")
    if not (c_slip and c_inv and c_amt):
        frappe.throw(_("Sheet Doc của AEON thiếu cột bắt buộc (SLIP TYPE / SUPPLIER "
                       "INVOICE / AMOUNT) — mẫu file đã đổi, KHÔNG đọc tiếp"))

    # BẪY 1: dừng ĐÚNG ở dòng nhãn 'Slip Type' lần thứ hai (đầu khối tổng).
    sr, scols = _ae_find_header(doc, _AE_H_DOC, start=hr + 1)
    end = sr or (len(doc) + 1)

    for r in range(hr + 1, end):
        slip = _txt(_gv(doc, r, c_slip)).split(".")[0]
        if slip not in _AE_SLIP_KIND:
            continue
        kind, subtype = _AE_SLIP_KIND[slip]
        amt = to_number(_gv(doc, r, c_amt))
        if amt is None:
            warnings.append("Sheet %s dòng %d: mã slip %s nhưng KHÔNG đọc được AMOUNT "
                            "— dòng bị bỏ, số kiểm tra bên dưới sẽ lệch." % (doc_name, r, slip))
            continue
        series, inv_no = _ae_split_invoice(_gv(doc, r, c_inv))
        # Hàng trả (312) lưu DƯƠNG trong bảng nhưng bản chất là GIẢM tiền về; đảo
        # dấu để `signed_amount` đúng chiều. Việc PHÂN LOẠI đã do SLIP TYPE quyết,
        # không do dấu — đảo dấu ở đây chỉ là chuẩn hóa chiều tiền của một loại
        # dòng đã biết chắc, không phải suy loại từ dấu.
        signed = -amt if kind == "ghi_giam" else amt
        desc = " · ".join(x for x in (
            "AEON slip %s (%s)" % (slip, subtype),
            _txt(_gv(doc, r, c_slipno)) and ("slip no %s" % _txt(_gv(doc, r, c_slipno))),
            _txt(_gv(doc, r, c_remarks)) if c_remarks else "",
        ) if x)
        rows.append(_line(
            row_kind=kind, row_subtype=subtype,
            inv_series=series, inv_no=inv_no,
            inv_date=to_date(_gv(doc, r, c_date)) if c_date else None,
            store_code=_txt(_gv(doc, r, c_store)) if c_store else "",
            doc_no=_txt(_gv(doc, r, c_contract)) if c_contract else "",
            description=desc,
            signed_amount=signed, payment_date=pay_date,
            # Không bóc được ký hiệu = không khớp được hóa đơn -> phải có người xem.
            needs_review=not series,
            source_sheet=doc_name, source_row=r))

    # ── Khối tổng sheet Doc -> số kiểm tra (tiền VÀ số lượng slip) ──────────
    #
    # VÌ SAO phải đọc theo CỘT của khối tổng chứ không "số cuối cùng trên dòng":
    # khối này có cột 'No of Slips' đứng SAU cột 'Amount'. Lấy số cuối dòng thì
    # 'Net Purchase' ra 28 (số slip) thay vì 59.338.440 — số kiểm tra biến thành
    # rác mà vẫn trông như một con số hợp lệ.
    declared, declared_n = {}, {}
    by_slip, by_slip_n = {}, {}
    if sr:
        s_slip = _ae_col(scols, _AE_H_DOC)
        s_desc = _ae_col(scols, "DESCRIPTION")
        s_amt = _ae_col(scols, "AMOUNT")
        s_n = _ae_col(scols, "NO OF SLIPS")
        for r in range(sr + 1, len(doc) + 1):
            amt = to_number(_gv(doc, r, s_amt)) if s_amt else None
            cnt = to_number(_gv(doc, r, s_n)) if s_n else None
            if amt is None:
                continue
            code = _txt(_gv(doc, r, s_slip)).split(".")[0] if s_slip else ""
            if code in _AE_SLIP_KIND:
                by_slip[code] = amt
                by_slip_n[code] = cnt
                continue
            key = _ae_norm(_txt(_gv(doc, r, s_desc)) if s_desc else "")
            if key in ("netpurchase", "deduction", "netpayment"):
                declared.setdefault(key, amt)
                declared_n.setdefault(key, cnt)
    else:
        warnings.append("Sheet Doc không có khối tổng cuối bảng — mất toàn bộ số "
                        "kiểm tra Net Purchase / Deduction / Net Payment.")

    # ── Sheet Costdet: khoản AEON trừ, kèm số hóa đơn AEON xuất cho MÌNH ────
    cost_name, cost = _ae_sheet(sheets, "Costdet")
    by_code, sub_declared, cd_total_declared = {}, {}, None
    n_cost_rows = 0
    if not cost:
        warnings.append("File thiếu sheet 'Costdet' — KHÔNG có dòng khoản trừ nào "
                        "được đọc, trong khi khối tổng vẫn khai có khoản trừ.")
    else:
        chr_, ccols = _ae_find_header(cost, _AE_H_COST)
        if not chr_:
            warnings.append("Sheet Costdet không có dòng nhãn 'DEDUCTION CODE' — "
                            "bỏ qua TOÀN BỘ khoản trừ, số kiểm tra sẽ báo lệch.")
        else:
            k_code = _ae_col(ccols, _AE_H_COST)
            k_desc = _ae_col(ccols, "DEDUCTION DESCRIPTION")
            k_tax = _ae_col(ccols, "TAX INVOICE")
            k_store = _ae_col(ccols, "STORE CODE")
            k_date = _ae_col(ccols, "SLIP DATE")
            k_amt = _ae_col(ccols, "AMOUNT")
            k_slip = _ae_col(ccols, "SLIP NO")
            k_rem = _ae_col(ccols, "REMARKS")
            if not (k_code and k_amt):
                frappe.throw(_("Sheet Costdet của AEON thiếu cột DEDUCTION CODE / AMOUNT"))

            last_code = None
            for r in range(chr_ + 1, len(cost) + 1):
                code = _txt(_gv(cost, r, k_code))
                amt = to_number(_gv(cost, r, k_amt))
                joined = strip_tones(" ".join(t for t in _row_texts(cost[r - 1]) if t))
                if not code:
                    # Dòng 'Sub-Total' của từng mã và dòng 'Total' cuối sheet là SỐ
                    # KIỂM TRA, không phải khoản trừ. Nhận diện bằng chính nhãn.
                    if amt is None:
                        continue
                    if "sub-total" in joined or "subtotal" in joined:
                        if last_code:
                            sub_declared[last_code] = amt
                    elif joined.startswith("total"):
                        cd_total_declared = amt
                    continue
                if amt is None:
                    warnings.append("Sheet %s dòng %d: mã trừ %s nhưng KHÔNG đọc được "
                                    "AMOUNT — dòng bị bỏ." % (cost_name, r, code))
                    continue
                last_code = code
                n_cost_rows += 1
                by_code[code] = by_code.get(code, 0.0) + amt
                # `TAX INVOICE` là hóa đơn AEON xuất cho MÌNH (mẫu 1-K26TBE-...),
                # KHÔNG phải hóa đơn mình bán ra -> vào `doc_no`, tuyệt đối không
                # vào `inv_no`: vào inv_no là đi khớp nhầm với Sales Invoice.
                rows.append(_line(
                    row_kind="phi", row_subtype=code,
                    doc_no=_txt(_gv(cost, r, k_tax)) if k_tax else "",
                    store_code=_txt(_gv(cost, r, k_store)) if k_store else "",
                    inv_date=to_date(_gv(cost, r, k_date)) if k_date else None,
                    description=" · ".join(x for x in (
                        _txt(_gv(cost, r, k_desc)) if k_desc else "",
                        ("slip %s" % _txt(_gv(cost, r, k_slip))) if k_slip and _txt(_gv(cost, r, k_slip)) else "",
                        _txt(_gv(cost, r, k_rem)) if k_rem else "",
                    ) if x),
                    # BẪY 3: `-amt`, KHÔNG `-abs(amt)`. Dòng gốc âm là khoản hoàn
                    # lại, phải ra dương.
                    signed_amount=-amt, payment_date=pay_date,
                    source_sheet=cost_name, source_row=r))

    # ── Đối chiếu ──────────────────────────────────────────────────────────
    goods = _sum(rows, "signed_amount", {"thanh_toan"})
    returns = _sum(rows, "signed_amount", {"ghi_giam"})
    fees = _sum(rows, "signed_amount", {"phi"})
    n_goods = sum(1 for r in rows if r["row_kind"] == "thanh_toan")
    n_returns = sum(1 for r in rows if r["row_kind"] == "ghi_giam")
    # Từ đây trở xuống parser CHỈ được dựng check/warning, KHÔNG được thêm dòng
    # tiền nữa. `_assert_rows_frozen` ở cuối hàm canh đúng điều đó.
    n_frozen = len(rows)

    checks.append(_check("Tổng slip 311 (PO)", by_slip.get("311"), goods))
    checks.append(_check("Tổng slip 312 (GRN)", by_slip.get("312"), returns))
    checks.append(_check("Net Purchase (khối tổng sheet Doc)",
                         declared.get("netpurchase"), goods + returns))
    checks.append(_check("Deduction (khối tổng sheet Doc)",
                         declared.get("deduction"), fees))
    checks.append(_check("Net Payment (khối tổng sheet Doc)",
                         declared.get("netpayment"), goods + returns + fees))

    # SỐ LƯỢNG SLIP — chốt chống "mất dòng trong im lặng". Một dòng bị bỏ vì đọc
    # hụt cột tiền làm tổng lệch, nhưng một dòng bị bỏ vì trùng khóa/lọc sai có
    # thể vẫn khớp tổng nếu nó bằng 0. Đếm mới bắt được.
    checks.append(_check("Số slip 311 (PO)", by_slip_n.get("311"), float(n_goods)))
    checks.append(_check("Số slip 312 (GRN)", by_slip_n.get("312"), float(n_returns)))
    checks.append(_check("Số dòng khoản trừ (Deduction)",
                         declared_n.get("deduction"), float(n_cost_rows)))

    # Sub-Total từng mã khoản trừ — chốt mạnh nhất của sheet Costdet: nó bắt được
    # cả trường hợp một mã bị đọc nhầm sang mã khác (tổng chung vẫn khớp).
    for code in sorted(by_code):
        checks.append(_check("Costdet Sub-Total mã %s" % code,
                             sub_declared.get(code), by_code[code]))
    checks.append(_check("Costdet Total", cd_total_declared, sum(by_code.values())))

    # NET PAYMENT ở sheet Summary — nguồn ĐỘC LẬP với khối tổng sheet Doc.
    _sm_name, summary = _ae_sheet(sheets, "Summary")
    if summary:
        sm = None
        for r in range(1, len(summary) + 1):
            if "net payment" in strip_tones(" ".join(_row_texts(summary[r - 1]))):
                for rr in range(r + 1, min(r + 4, len(summary) + 1)):
                    v = next((to_number(c) for c in summary[rr - 1]
                              if to_number(c) is not None), None)
                    if v is not None:
                        sm = v
                        break
                break
        checks.append(_check("NET PAYMENT (sheet Summary)", sm, goods + returns + fees))

    # CHỐT CHỐNG CỘNG TRÙNG (BẪY 2): tổng sheet DcCharges phải bằng ĐÚNG dòng mã
    # 'DC' của Costdet. Lệch nghĩa là DcCharges KHÔNG còn là chi tiết của DC nữa,
    # lúc đó việc bỏ qua sheet này là SAI và phải đọc lại mẫu file trước khi nạp.
    _dc_name, dcs = _ae_sheet(sheets, "DcCharges")
    if dcs:
        dc_total = None
        for r in range(len(dcs), 0, -1):
            if "total" in strip_tones(" ".join(_row_texts(dcs[r - 1]))):
                dc_total = next((to_number(c) for c in reversed(dcs[r - 1])
                                 if to_number(c) is not None), None)
                if dc_total is not None:
                    break
        checks.append(_check("DcCharges = dòng mã DC của Costdet",
                             dc_total, by_code.get("DC", 0.0)))

    # Danh mục Costsumm/Rebsumm: mã có tiền mà KHÔNG thấy trong Costdet -> mất khoản trừ.
    for label, prefix in (("Costsumm", "Costsumm"), ("Rebsumm", "Rebsumm")):
        _n, grid = _ae_sheet(sheets, prefix)
        if not grid:
            continue
        gr, _gc = _ae_find_header(grid, _AE_H_COST)
        if not gr:
            continue
        # Bảng ĐÔI/BA khối cột song song: mỗi khối là (mã, mô tả, tiền) liền nhau.
        # Duyệt theo bước 3 cột thay vì đoán vị trí từng khối.
        for r in range(gr + 1, len(grid) + 1):
            texts = _row_texts(grid[r - 1])
            for c0 in range(0, len(texts) - 2, 3):
                code = texts[c0]
                amt = to_number(_g(grid, r, c0 + 3))
                if code and amt not in (None, 0.0) and code not in by_code:
                    warnings.append("Danh mục %s: mã %s có giá trị %s nhưng KHÔNG có "
                                    "dòng nào trong Costdet — nghi mất khoản trừ."
                                    % (label, code, f"{amt:,.0f}"))

    _assert_rows_frozen(rows, n_frozen, "parse_aeon")

    groups = [{
        "key": pay_date or advice_no or "aeon",
        "advice_no": advice_no, "payment_date": pay_date, "n_rows": len(rows),
        "declared_gross": declared.get("netpurchase"),
        "declared_payment": by_slip.get("311"),
        "computed_payment": goods,
        "computed_net": goods + returns + fees,
        "declared_net": declared.get("netpayment"),
    }]

    return {
        "advice_no": advice_no,
        "payment_dates": [pay_date] if pay_date else [],
        "declared_totals": {
            "total_payment": by_slip.get("311"),
            "total_discount": None,
            "net_payment": declared.get("netpayment"),
            "net_purchase": declared.get("netpurchase"),
            "vendor_code": meta.get("vendor_code"),
            "credit_term": meta.get("credit_term"),
        },
        "rows": rows, "checks": checks, "groups": groups, "warnings": warnings,
    }


# ═══════════════════════════════════════════════════════════════════════════
# 4g. Fuji Mart  (.xls BIFF, đuôi '.Xls' VIẾT HOA, 1 sheet, BỐN khối liên tiếp)
# ═══════════════════════════════════════════════════════════════════════════
#
# Bốn khối, nhận diện bằng nhãn header — KHÔNG bằng chỉ số dòng (file kỳ sau có
# số dòng khác hẳn):
#   K1 'THEO HĐTC'                -> hóa đơn <-> phiếu nhập kho + mã kho  (đối chiếu)
#   K2 'SỐ HÓA ĐƠN' + 'SỐ TIỀN'   -> tổng theo hóa đơn                   (SINH TIỀN)
#   K3 'GIÁ TRỊ TIỀN THEO PNK/XK' -> hàng trả, số ÂM sẵn                 (SINH TIỀN)
#   K4 'TÊN CHIẾT KHẤU'           -> chiết khấu / hỗ trợ, cặp 2 dòng     (SINH TIỀN)
#
# K1 và K2 là CÙNG MỘT SỐ TIỀN nhìn từ hai phía (90.010.980). Sinh dòng từ cả
# hai là nhân đôi doanh thu. Chỉ K2 sinh tiền; K1 dùng để (a) đối chiếu chéo và
# (b) bổ sung số phiếu nhập kho + mã kho cho từng hóa đơn.
_FJ_H_B1 = "THEO HĐTC"
_FJ_H_B2 = "SỐ HÓA ĐƠN"
_FJ_H_B3 = "GIÁ TRỊ TIỀN THEO PNK/XK"
_FJ_H_B4 = "TÊN CHIẾT KHẤU"

# Phân loại khoản trừ theo NHÃN (không theo dấu, không theo thứ tự dòng).
_FJ_DEDUCT_KIND = (
    ("chiet khau", "chiet_khau"),
    ("ho tro", "phi"),
)

# Sai số cho phép khi soát 'căn cứ × tỷ lệ = tiền'. Đây là kiểm HỢP LÝ, không
# phải số kiểm tra: Fuji làm tròn tới đồng (81.819.909 × 0,5% = 409.099,545 in
# thành 409.100), nên lệch dưới 1 đồng là chuyện bình thường. Quá 1 đồng mới là
# dấu hiệu đọc nhầm cột căn cứ hoặc cột tỷ lệ.
_FJ_RATE_EPS = 1.0


def _fj_find(grid, label, start=1):
    """Dòng header theo NHÃN. Trả (r 1-based, {khóa nhãn: cột})."""
    tgt = _ae_norm(label)
    for r in range(start, len(grid) + 1):
        texts = _row_texts(grid[r - 1])
        if any(_ae_norm(t) == tgt for t in texts):
            return r, {_ae_norm(t): c for c, t in enumerate(texts, start=1) if t}
    return None, {}


def _fj_sub(grid, top_row, top_col, next_top_col, label):
    """Cột của nhãn tầng 2 nằm DƯỚI một nhóm của tầng 1.

    VÌ SAO cần: K1 có nhãn 'NGÀY/THÁNG' xuất hiện HAI LẦN ở tầng 2 — một lần
    dưới nhóm 'THEO HĐTC' (ngày hóa đơn), một lần dưới nhóm 'THEO PHIẾU NK/XK'
    (ngày nhập kho). Tra phẳng theo nhãn sẽ lấy nhầm ngày, và ngày sai thì tầng
    khớp 'số + ngày + tiền' của hóa đơn không ký hiệu sẽ trượt sạch.
    """
    if not top_row or top_row + 1 > len(grid):
        return None
    tgt = _ae_norm(label)
    texts = _row_texts(grid[top_row])          # tầng 2 = dòng ngay dưới tầng 1
    hi = next_top_col or (len(texts) + 1)
    for c, t in enumerate(texts, start=1):
        if top_col <= c < hi and _ae_norm(t) == tgt:
            return c
    return None


def _fj_deduct_kind(name):
    """Nhãn khoản trừ -> row_kind. Không nhận ra thì 'khac' + để người xem lại."""
    key = strip_tones(name)
    for prefix, kind in _FJ_DEDUCT_KIND:
        if key.startswith(prefix):
            return kind, False
    return "khac", True


def parse_fuji(sheets):
    """Bảng kê thanh toán Fuji Mart.

    KHÔNG CÓ NGÀY THANH TOÁN VÀ KHÔNG CÓ SỐ BẢNG KÊ trong file — đã soi hết 65
    dòng x 14 cột của file mẫu, mười ba dòng đầu trống trơn. Không bịa: trả None
    và cảnh báo để kế toán điền tay.

    KHÔNG CÓ KÝ HIỆU HÓA ĐƠN — chỉ có số trần ('4409'). Giống Emart, nên tầng
    khớp phải đi nhánh 'số + ngày + tiền' và luôn để 'Cần review'
    (xem `_SERIESLESS_CHAINS` trong ketoan/api/mt.py).

    BẪY NHÂN ĐÔI DOANH THU: khối 1 và khối 2 là CÙNG một số tiền nhìn từ hai
    phía (phiếu nhập kho vs hóa đơn), cùng bằng 90.010.980 trên file mẫu. Sinh
    dòng tiền từ cả hai là nhân đôi. Chỉ khối 2 sinh tiền.
    """
    warnings, rows, checks = [], [], []

    # File một sheet, nhưng vẫn quét MỌI sheet để không bỏ sót (§J.1).
    grid, sheet_name = None, ""
    for name, g in sheets:
        if not _sheet_has_content(g):
            continue
        if _fj_find(g, _FJ_H_B2)[0] or _fj_find(g, _FJ_H_B1)[0]:
            grid, sheet_name = g, name
            break
    if grid is None:
        frappe.throw(_("Không tìm thấy bảng hóa đơn trong file Fuji (thiếu nhãn '{0}')")
                     .format(_FJ_H_B2))
    for name, g in sheets:
        if name != sheet_name and _sheet_has_content(g):
            warnings.append("Sheet '%s' còn dữ liệu nhưng tầng đọc Fuji chưa hiểu — "
                            "KIỂM TAY trước khi nạp, có thể đang bỏ sót tiền." % name)

    # ── Khối 1: hóa đơn <-> phiếu nhập kho (KHÔNG sinh tiền) ───────────────
    b1 = {}
    b1_total, n_b1 = None, 0
    r1, c1cols = _fj_find(grid, _FJ_H_B1)
    if not r1:
        warnings.append("Không thấy khối '%s' — mất đối chiếu chéo hóa đơn/phiếu nhập "
                        "kho và mọi dòng sẽ thiếu số PNK." % _FJ_H_B1)
    else:
        g_hdtc = _ae_col(c1cols, _FJ_H_B1)
        g_pnk = _ae_col(c1cols, "THEO PHIẾU NK/XK")
        c_amt1 = _ae_col(c1cols, "GIÁ TRỊ TIỀN THEO PNK")
        c_stt1 = _ae_col(c1cols, "STT")
        c_inv1 = _fj_sub(grid, r1, g_hdtc or 1, g_pnk, "SỐ HĐTC")
        c_idate = _fj_sub(grid, r1, g_hdtc or 1, g_pnk, "NGÀY/THÁNG")
        c_pnk = _fj_sub(grid, r1, g_pnk or 1, c_amt1, "SỐ PNK")
        c_kho = _fj_sub(grid, r1, g_pnk or 1, c_amt1, "MÃ KHO NHẬP")
        if not (c_inv1 and c_amt1 and c_stt1):
            warnings.append("Khối '%s' thiếu cột STT / SỐ HĐTC / GIÁ TRỊ TIỀN — bỏ qua "
                            "đối chiếu chéo." % _FJ_H_B1)
        else:
            tot = 0.0
            for r in range(r1 + 2, len(grid) + 1):     # +2: tầng 2 của header
                if to_number(_gv(grid, r, c_stt1)) is None:
                    break                               # hết khối (dòng trắng / header sau)
                inv = norm_inv_no(_txt(_gv(grid, r, c_inv1)))
                amt = to_number(_gv(grid, r, c_amt1))
                if not inv or amt is None:
                    continue
                n_b1 += 1
                tot += amt
                b1[inv] = {
                    "amount": amt,
                    "pnk": _txt(_gv(grid, r, c_pnk)) if c_pnk else "",
                    "kho": _txt(_gv(grid, r, c_kho)) if c_kho else "",
                    "inv_date": to_date(_gv(grid, r, c_idate)) if c_idate else None,
                    "row": r,
                }
            b1_total = tot

    # ── Khối 2: tổng theo hóa đơn (SINH TIỀN) ──────────────────────────────
    r2, c2cols = _fj_find(grid, _FJ_H_B2)
    if not r2:
        frappe.throw(_("File Fuji không có khối '{0}' — không đọc được hóa đơn nào")
                     .format(_FJ_H_B2))
    c_stt2 = _ae_col(c2cols, "STT")
    c_inv2 = _ae_col(c2cols, _FJ_H_B2)
    c_date2 = _ae_col(c2cols, "NGÀY HÓA ĐƠN")
    c_amt2 = _ae_col(c2cols, "SỐ TIỀN")
    if not (c_stt2 and c_inv2 and c_amt2):
        frappe.throw(_("Khối hóa đơn của Fuji thiếu cột STT / SỐ HÓA ĐƠN / SỐ TIỀN"))

    n_matched = 0
    for r in range(r2 + 1, len(grid) + 1):
        if to_number(_gv(grid, r, c_stt2)) is None:
            break
        raw_no = _txt(_gv(grid, r, c_inv2))     # có khoảng trắng thừa, _txt đã trim
        amt = to_number(_gv(grid, r, c_amt2))
        if not raw_no or amt is None:
            warnings.append("Khối hóa đơn dòng %d: thiếu số hóa đơn hoặc số tiền — bỏ qua." % r)
            continue
        key = norm_inv_no(raw_no)
        ref = b1.get(key) or {}
        if ref and abs(ref["amount"] - amt) <= MONEY_EPS:
            n_matched += 1
        rows.append(_line(
            row_kind="thanh_toan", row_subtype="HĐTC",
            # Fuji KHÔNG in ký hiệu -> để rỗng. Bịa 'C26THG' ở đây là ghi tiền vào
            # hóa đơn của chuỗi khác cùng dải số.
            inv_series="", inv_no=raw_no,
            inv_date=to_date(_gv(grid, r, c_date2)) if c_date2 else None,
            store_code=ref.get("kho") or "",
            doc_no=ref.get("pnk") or "",
            description="Hóa đơn %s%s" % (
                raw_no, (" · PNK %s" % ref["pnk"]) if ref.get("pnk") else ""),
            signed_amount=amt, payment_date=None,
            source_sheet=sheet_name, source_row=r))

    # ── Khối 3: hàng trả (SINH TIỀN, số đã ÂM sẵn trong file) ──────────────
    #
    # `n_b3_seen` được đếm ĐỘC LẬP với cột tiền, chỉ dựa vào cột STT, để làm số
    # kiểm tra cho chính việc đọc khối này. Xem chốt "Số dòng hàng trả" ở phần
    # đối chiếu — không có nó thì bỏ sót cả khối vẫn báo khớp.
    n_b3_seen = 0
    r3, c3cols = _fj_find(grid, _FJ_H_B3)
    if not r3:
        warnings.append("Không thấy khối '%s' (hàng trả). Nếu kỳ này thật sự không có "
                        "hàng trả thì bỏ qua; nếu có thì đang BỎ SÓT tiền ghi giảm."
                        % _FJ_H_B3)
    else:
        c_amt3 = _ae_col(c3cols, _FJ_H_B3)
        c_stt3 = _ae_col(c3cols, "STT")
        g_pnk3 = _ae_col(c3cols, "THEO PHIẾU NK/XK")
        c_pnk3 = _fj_sub(grid, r3, g_pnk3 or 1, c_amt3, "SỐ PNK/XK")
        c_date3 = _fj_sub(grid, r3, g_pnk3 or 1, c_amt3, "NGÀY/THÁNG")
        # Đếm dòng của khối TRƯỚC khi bàn tới cột tiền — đếm phải sống sót cả khi
        # cột tiền không dò ra, nếu không thì "hỏng cột tiền" = "khối không tồn
        # tại" và chốt bên dưới mất tác dụng.
        if c_stt3:
            for r in range(r3 + 2, len(grid) + 1):
                if to_number(_gv(grid, r, c_stt3)) is None:
                    break
                n_b3_seen += 1
        if not (c_amt3 and c_stt3):
            warnings.append("Khối hàng trả thiếu cột STT / GIÁ TRỊ — bỏ qua toàn khối.")
        else:
            for r in range(r3 + 2, len(grid) + 1):
                if to_number(_gv(grid, r, c_stt3)) is None:
                    break
                amt = to_number(_gv(grid, r, c_amt3))
                if amt is None:
                    continue
                pnk = _txt(_gv(grid, r, c_pnk3)) if c_pnk3 else ""
                # GIỮ NGUYÊN DẤU của file: khối này in sẵn số âm. `-abs()` ở đây thì
                # đúng trên file mẫu nhưng sẽ nuốt mất một dòng điều chỉnh dương nếu
                # kỳ sau có. Chiều tiền do file quyết, loại dòng do KHỐI quyết.
                rows.append(_line(
                    row_kind="ghi_giam", row_subtype="PNK/XK",
                    inv_series="", inv_no="",
                    inv_date=to_date(_gv(grid, r, c_date3)) if c_date3 else None,
                    doc_no=pnk,
                    description="Hàng trả theo phiếu %s" % (pnk or "(không số)"),
                    signed_amount=amt, payment_date=None,
                    # Không có số hóa đơn -> không tự nối được về hóa đơn nào.
                    needs_review=True,
                    source_sheet=sheet_name, source_row=r))

    # ── Khối 4: chiết khấu / hỗ trợ (SINH TIỀN, cặp 2 dòng mỗi mục) ─────────
    #
    # Cấu trúc thật: dòng TÊN mang nhãn + số tiền; dòng CHI TIẾT ngay dưới mang
    # STT + doanh số căn cứ + tỷ lệ + số tiền (LẶP LẠI). Đọc gộp cả hai dòng là
    # cộng đôi toàn bộ chiết khấu.
    items = []
    r4, c4cols = _fj_find(grid, _FJ_H_B4)
    if not r4:
        warnings.append("Không thấy khối '%s' — nếu kỳ này có chiết khấu/hỗ trợ thì "
                        "đang BỎ SÓT." % _FJ_H_B4)
    else:
        c_name = _ae_col(c4cols, _FJ_H_B4)
        c_stt4 = _ae_col(c4cols, "STT")
        c_base = _ae_col(c4cols, "DOANH SỐ BAS")
        c_rate = _ae_col(c4cols, "TỶ LỆ")
        c_amt4 = _ae_col(c4cols, "TIỀN CHIẾT KHẤU")
        if not (c_name and c_amt4):
            warnings.append("Khối chiết khấu thiếu cột TÊN / TIỀN — bỏ qua toàn khối.")
        else:
            cur = None
            for r in range(r4 + 1, len(grid) + 1):
                name = _txt(_gv(grid, r, c_name))
                stt = to_number(_gv(grid, r, c_stt4)) if c_stt4 else None
                amt = to_number(_gv(grid, r, c_amt4))
                if name:
                    if cur:
                        items.append(cur)   # mục trước thiếu dòng chi tiết
                    cur = {"name": name, "label_amount": amt, "row": r,
                           "base": None, "rate": None, "amount": None, "detail_row": None}
                elif stt is not None and cur is not None:
                    cur["base"] = to_number(_gv(grid, r, c_base)) if c_base else None
                    cur["rate"] = to_number(_gv(grid, r, c_rate)) if c_rate else None
                    cur["amount"] = amt
                    cur["detail_row"] = r
                    # Số thứ tự do FILE in ra -> số kiểm tra cho chính số MỤC đọc
                    # được. Xem chốt "Số mục chiết khấu/hỗ trợ" ở phần đối chiếu.
                    cur["stt"] = stt
                    items.append(cur)
                    cur = None
            if cur:
                items.append(cur)

    gross = _sum(rows, "signed_amount", {"thanh_toan"})
    returns = _sum(rows, "signed_amount", {"ghi_giam"})
    net_base = gross + returns
    bases = [it["base"] for it in items if it["base"] is not None]

    for it in items:
        amt = it["amount"] if it["amount"] is not None else it["label_amount"]
        if amt is None:
            warnings.append("Khoản '%s' (dòng %d): KHÔNG đọc được số tiền — bỏ qua, "
                            "số kiểm tra sẽ báo lệch." % (it["name"], it["row"]))
            continue
        kind, unknown = _fj_deduct_kind(it["name"])
        if unknown:
            warnings.append("Khoản '%s' không bắt đầu bằng 'Chiết khấu' hay 'Hỗ trợ' — "
                            "xếp tạm vào 'Khác', kế toán phải phân loại tay." % it["name"])
        base, rate = it["base"], it["rate"]
        # Căn cứ tính phải là MỘT trong hai đại lượng đọc được từ chính file:
        # tổng hóa đơn, hoặc tổng hóa đơn trừ hàng trả. Khác hai cái đó nghĩa là
        # Fuji dùng một doanh số mình không kiểm được -> bắt người xem.
        base_ok = base is not None and (abs(base - gross) <= MONEY_EPS
                                        or abs(base - net_base) <= MONEY_EPS)
        rate_ok = True
        if base is not None and rate is not None:
            rate_ok = abs(base * rate - amt) <= _FJ_RATE_EPS
            if not rate_ok:
                warnings.append("Khoản '%s': căn cứ %s × %s = %s nhưng file in %s — "
                                "nghi đọc nhầm cột." % (
                                    it["name"], f"{base:,.0f}", rate,
                                    f"{base * rate:,.0f}", f"{amt:,.0f}"))
        if base is not None and not base_ok:
            warnings.append("Khoản '%s': căn cứ %s không bằng tổng hóa đơn (%s) cũng "
                            "không bằng tổng sau trả hàng (%s) — không kiểm được, "
                            "kế toán xác nhận tay." % (
                                it["name"], f"{base:,.0f}", f"{gross:,.0f}",
                                f"{net_base:,.0f}"))
        desc = it["name"]
        if base is not None and rate is not None:
            desc += " · căn cứ %s × %s" % (f"{base:,.0f}", rate)
        rows.append(_line(
            row_kind=kind, row_subtype=it["name"][:60],
            inv_series="", inv_no="",
            description=desc,
            # Khoản trừ làm GIẢM tiền về -> âm. Loại dòng do NHÃN quyết, không do dấu.
            signed_amount=-abs(amt), payment_date=None,
            # Chỉ gắn cờ khi CHÍNH tầng đọc không chắc (nhãn lạ / căn cứ không kiểm
            # được / tỷ lệ không ra). Khoản kỳ nào cũng có và kiểm được thì không
            # gắn — gắn tất là làm cờ mất nghĩa.
            needs_review=bool(unknown or not base_ok or not rate_ok),
            source_sheet=sheet_name,
            source_row=it.get("detail_row") or it["row"]))

    # ── Đối chiếu ──────────────────────────────────────────────────────────
    #
    # File Fuji KHÔNG in tổng thanh toán ròng, nên số kiểm tra ở đây đều là
    # ĐẲNG THỨC GIỮA HAI SỐ CÙNG DO FILE IN RA — vẫn đủ mạnh: chúng bắt được đọc
    # nhầm cột, sót dòng và lệch khối.
    n_b2 = sum(1 for r in rows if r["row_kind"] == "thanh_toan")
    # Xem `_assert_rows_frozen`: dưới mốc này chỉ được dựng check/warning.
    n_frozen = len(rows)
    checks.append(_check("Tổng theo PNK (khối 1) = Tổng theo hóa đơn (khối 2)",
                         b1_total, gross))
    checks.append(_check("Số dòng khối 1 = số dòng khối 2", float(n_b1), float(n_b2)))
    checks.append(_check("Khớp từng hóa đơn giữa khối 1 và khối 2",
                         float(n_b2), float(n_matched)))

    # Đẳng thức chiết khấu: dòng TÊN và dòng CHI TIẾT in cùng một số tiền. Chỉ
    # thêm khi kỳ này CÓ khoản trừ — kỳ không có khoản nào mà vẫn dựng một mục
    # `declared=None` là tự làm `reconciled=False` cho một file hoàn toàn sạch.
    if items:
        # `_declared_sum` trả None nếu bất kỳ vế nào đọc không ra — đúng ý: thiếu
        # một số thì KHÔNG được coi là 0 rồi báo khớp.
        checks.append(_check(
            "Tiền chiết khấu/hỗ trợ: dòng tên = dòng chi tiết",
            _declared_sum([it["label_amount"] for it in items]),
            _declared_sum([it["amount"] for it in items]) or 0.0))

    # File không in tổng hàng trả, nhưng có in doanh số căn cứ tính chiết khấu,
    # và căn cứ đó chính bằng "tổng hóa đơn − hàng trả" -> đối chiếu chéo được.
    dg = next((b for b in bases if abs(b - gross) <= MONEY_EPS), None)
    dn = next((b for b in bases if abs(b - net_base) <= MONEY_EPS), None)
    checks.append(_check("Tổng hóa đơn = doanh số căn cứ in trong file", dg, gross))
    checks.append(_check("Tổng hóa đơn − hàng trả = doanh số căn cứ in trong file",
                         dn, net_base))

    # CHỐT CHỐNG CỘNG TRÙNG / BỎ SÓT MỤC CHIẾT KHẤU.
    #
    # VÌ SAO đẳng thức "dòng tên = dòng chi tiết" KHÔNG đủ: mỗi mục góp một số
    # vào CẢ HAI vế, nên nhân đôi (hay bỏ sót) cả mục thì hai vế vẫn bằng nhau.
    # Đã đo bằng đột biến: `items.append` hai lần làm chiết khấu tăng đúng gấp
    # đôi mà cả 7 số kiểm tra vẫn báo khớp. File KHÔNG in tổng khoản trừ, nhưng
    # CÓ in số thứ tự — và số thứ tự cuối cùng chính là số mục.
    stts = [it["stt"] for it in items if it.get("stt") is not None]
    if items:
        checks.append(_check("Số mục chiết khấu/hỗ trợ (theo cột STT)",
                             float(max(stts)) if stts else None, float(len(items))))

    # CHỐT CHỐNG BỎ SÓT CẢ KHỐI HÀNG TRẢ.
    #
    # VÌ SAO hai đẳng thức trên KHÔNG đủ: nếu khối 3 không được đọc thì `net_base`
    # tụt về đúng `gross`, mà `gross` LẠI LÀ một trong các doanh số căn cứ in
    # trong file — nên `dn` vẫn tìm thấy và phép kiểm khớp một cách GIẢ. Đã đo
    # bằng đột biến: ép `c_amt3 = None` làm mất sạch 8.191.071 tiền hàng trả mà
    # cả 6 số kiểm tra đều báo khớp.
    #
    # Chốt này đếm số dòng của khối bằng cột STT — độc lập hoàn toàn với cột
    # tiền, nên "hỏng cột tiền" không còn giả trang được thành "khối không có".
    n_b3_rows = sum(1 for r in rows if r["row_kind"] == "ghi_giam")
    checks.append(_check("Số dòng hàng trả đọc được", float(n_b3_seen), float(n_b3_rows)))

    warnings.append("File Fuji không in NGÀY THANH TOÁN và SỐ BẢNG KÊ — kế toán "
                    "phải điền ngày thanh toán tay sau khi nạp, nếu không bảng kê "
                    "sẽ không lên đúng kỳ trên màn hình Công nợ MT.")

    _assert_rows_frozen(rows, n_frozen, "parse_fuji")

    groups = [{
        "key": "fuji",
        "advice_no": None, "payment_date": None, "n_rows": len(rows),
        "declared_gross": b1_total,
        "declared_payment": b1_total,
        "computed_payment": gross,
        "computed_net": _sum(rows, "signed_amount"),
        "declared_net": None,
    }]

    return {
        "advice_no": None,
        "payment_dates": [],
        "declared_totals": {
            "total_payment": b1_total,
            "total_discount": None,
            "net_payment": None,
        },
        "rows": rows, "checks": checks, "groups": groups, "warnings": warnings,
    }


# ═══════════════════════════════════════════════════════════════════════════
# 4h. Mega Market  (.xls BIFF, 1 sheet, bảng PHẲNG nhãn tiếng Anh)
# ═══════════════════════════════════════════════════════════════════════════
#
# Đo trên file thật `docs/mt/samples/cttt_mega.xls`: 1 sheet 'Sheet1', header ở
# dòng 1, 18 dòng dữ liệu, 10 cột:
#
#   Store no | Supplier code | Supplier name | Description | Invoice no |
#   Amount | Invoice Date | GL date | Due date | Payment date
#
# `GL date` và `Due date` RỖNG toàn bộ. `Payment date` giống nhau ở mọi dòng
# (2026-07-10) — đó là ngày thanh toán. `Description` = "<Invoice no>,<Store no>",
# suy ra được từ hai cột kia nên không mang thêm thông tin, nhưng chính vì suy ra
# được nên nó là phép KIỂM CỘT: đọc lệch cột thì đẳng thức đó vỡ.
#
# ── Hai loại chứng từ, phân biệt bằng KÝ HIỆU chứ không bằng dấu tiền ──────
#
#   1C26THG_00004450   <- hóa đơn BÁN RA của mình  (8 dòng, tất cả DƯƠNG)
#   C26TAP 3269        <- chứng từ ký hiệu khác     (10 dòng, tất cả ÂM)
#
# Dấu và loại trùng khớp 18/18 dòng trên file mẫu, NHƯNG phân loại theo dấu là
# điều cấm của hợp đồng (§B) và đã có tiền lệ hỏng: khi chuỗi đổi quy ước dấu,
# một khoản ghi giảm lặng lẽ trôi thành TIỀN ĐÃ THU trên đúng hóa đơn đó, mà mọi
# số kiểm tra SUM vẫn khớp vì tổng không đổi. Nên phân loại bằng ký hiệu, còn dấu
# chỉ dùng để GẮN CỜ 'cần người xem lại'.
#
# ── Bằng chứng cho cách phân loại ──────────────────────────────────────────
#
# Đối chiếu với file công nợ Mega (`congno_mega_market.xlsx`): cả 8 dòng ký hiệu
# THG khớp ĐÚNG TỪNG ĐỒNG với cột TỔNG của hóa đơn tương ứng, và cả 8 hóa đơn đó
# đều đã về `Số còn nợ = 0`. Không dòng `C26TAP` nào khớp được như vậy.
#
# ⚠ Số của `C26TAP` ĐỤNG số hóa đơn của mình: `C26TAP 3264` và hóa đơn
# `00003264` (29/08/2024) cùng tồn tại. Đây đúng là ca MT2-G đã đo trên 7 chuỗi —
# khớp dòng ghi giảm sang Sales Invoice bằng số là vơ nhầm hóa đơn của mình.
# Dòng ghi giảm vẫn GIỮ số để kế toán nhìn, nhưng không đường nào cho nó nối
# Sales Invoice (chặn ở `_match_row`, ở `relink_line`, và ở `_paid_subquery`).
#
# ── KHÔNG CÓ SỐ KIỂM TRA TRONG FILE ───────────────────────────────────────
#
# Không dòng TỔNG CỘNG, không ô net payment, không số bảng kê. Nên `checks` để
# RỖNG và `reconciled` = False. "Không kiểm được" KHÔNG phải là "đã kiểm và
# đúng" — nhét phép kiểm cấu trúc vào `checks` để màn hình sáng xanh là nói dối
# về thứ chưa từng được đối chiếu.

# Ba ký tự cuối của ký hiệu hóa đơn là MÃ NGƯỜI BÁN (TT78: <mẫu số><C|K><năm>
# <3 ký tự>). Của mình là THG — Thực phẩm Hoàng Giang. Đổi pháp nhân thì đổi ở
# ĐÂY, không rải chuỗi 'THG' vào từng parser.
OUR_ISSUER_CODE = "THG"

_MG_H_INV = "Invoice no"
_MG_H_AMT = "Amount"
_MG_H_STORE = "Store no"
_MG_H_DESC = "Description"
_MG_H_IDATE = "Invoice Date"
_MG_H_PDATE = "Payment date"
_MG_H_SUPP = "Supplier code"

# Ký hiệu và số cách nhau bằng '_' (hóa đơn của mình) hoặc dấu cách (chứng từ
# ký hiệu khác) — CÙNG một file dùng cả hai.
_MG_SEPS = " _"


def is_our_series(series) -> bool:
    """Ký hiệu này có phải hóa đơn BÁN RA của mình không.

    Dùng mã người bán ở cuối ký hiệu, không dùng dấu tiền và không dùng độ dài.
    Ký hiệu rỗng -> False: không biết thì KHÔNG kết luận là của mình.
    """
    s = norm_series(series)
    return bool(s) and s.endswith(OUR_ISSUER_CODE)


def parse_mega(sheets):
    """Bảng kê thanh toán Mega Market.

    Xem khối chú thích ngay trên hàm này để biết vì sao phân loại bằng ký hiệu
    và vì sao file này không có số kiểm tra nào.
    """
    warnings = []
    rows = []

    hdr = None
    for name, grid in sheets:
        r, cols = _ae_find_header(grid, _MG_H_INV)
        if r and _ae_col(cols, _MG_H_AMT):
            hdr = (name, grid, r, cols)
            break
    if not hdr:
        frappe.throw(_("File Mega Market không có dòng tiêu đề mang đủ '{0}' và '{1}' — "
                       "không đọc được cột tiền nào.").format(_MG_H_INV, _MG_H_AMT))
    sheet_name, grid, hrow, cols = hdr

    c_inv = _ae_col(cols, _MG_H_INV)
    c_amt = _ae_col(cols, _MG_H_AMT)
    c_store = _ae_col(cols, _MG_H_STORE)
    c_desc = _ae_col(cols, _MG_H_DESC)
    c_idate = _ae_col(cols, _MG_H_IDATE)
    c_pdate = _ae_col(cols, _MG_H_PDATE)
    c_supp = _ae_col(cols, _MG_H_SUPP)

    if not c_pdate:
        warnings.append(
            "File không có cột '%s' — mọi dòng sẽ thiếu ngày thanh toán, kế toán phải "
            "điền tay trước khi sinh bút toán." % _MG_H_PDATE)

    n_desc_ok = n_desc_seen = 0
    pay_dates = set()
    skipped = 0

    for r in range(hrow + 1, len(grid) + 1):
        raw_inv = _txt(_gv(grid, r, c_inv))
        amt = to_number(_gv(grid, r, c_amt))
        if not raw_inv and amt is None:
            continue                       # dòng trắng giữa/cuối bảng
        if amt is None:
            warnings.append("Dòng %d: có tham chiếu '%s' nhưng KHÔNG đọc được số tiền — "
                            "bỏ qua." % (r, raw_inv))
            skipped += 1
            continue

        series, no = split_invoice_ref(raw_inv, _MG_SEPS)
        ours = is_our_series(series)
        store = _txt(_gv(grid, r, c_store)) if c_store else ""
        # Store no ra từ ô số nên mang đuôi '.0' — cắt đi, nếu không mã cửa hàng
        # '590072.0' không khớp mã nào trong master MT Store.
        if store.endswith(".0"):
            store = store[:-2]

        # KIỂM CỘT bằng chính sự thừa của file: Description = "<số>,<mã cửa hàng>".
        if c_desc:
            want = "%s,%s" % (raw_inv, store)
            got = _txt(_gv(grid, r, c_desc))
            if got:
                n_desc_seen += 1
                if got == want:
                    n_desc_ok += 1

        # KHÔNG có dấu phân cách -> không bóc được ký hiệu. Thà để rỗng và gắn cờ
        # còn hơn bịa ký hiệu rồi ghi tiền vào hóa đơn của người khác.
        needs_review = not series
        note = ""
        if not series:
            note = ("KHÔNG tách được ký hiệu khỏi '%s' (không có dấu '_' hay dấu cách)"
                    % raw_inv)
        # Dấu ngược với loại: KHÔNG đổi loại, chỉ bắt người nhìn.
        elif (ours and amt < 0) or (not ours and amt > 0):
            needs_review = True
            note = ("DẤU TIỀN NGƯỢC với loại chứng từ (ký hiệu %s, %s đ) — phân loại vẫn "
                    "theo KÝ HIỆU, xem tay trước khi khớp hóa đơn"
                    % (series, "{:+,.0f}".format(amt)))

        pdate = to_date(_gv(grid, r, c_pdate)) if c_pdate else None
        if pdate:
            pay_dates.add(pdate)

        rows.append(_line(
            row_kind="thanh_toan" if ours else "ghi_giam",
            row_subtype="Hóa đơn bán ra" if ours else "Chứng từ ký hiệu %s" % (series or "?"),
            # Dòng ghi giảm VẪN giữ ký hiệu/số để kế toán đối chiếu với chứng từ
            # Mega gửi. An toàn vì mọi đường nối Sales Invoice đều chặn dòng không
            # phải 'Thanh toán' (MT2-G).
            inv_series=series, inv_no=no,
            inv_date=to_date(_gv(grid, r, c_idate)) if c_idate else None,
            store_code=store or None,
            doc_no=_txt(_gv(grid, r, c_supp)) if c_supp else None,
            # `_line` không có ô ghi chú riêng — cảnh báo của DÒNG phải đi cùng
            # dòng, không thì nó chỉ còn ở `warnings` chung và mất dấu là dòng nào.
            description=((_txt(_gv(grid, r, c_desc)) if c_desc else raw_inv)
                         + (" · %s" % note if note else "")),
            signed_amount=amt,
            payment_date=pdate,
            needs_review=needs_review,
            source_sheet=sheet_name, source_row=r))

    if not rows:
        frappe.throw(_("Không đọc được dòng nào dưới tiêu đề của file Mega Market."))

    if n_desc_seen and n_desc_ok != n_desc_seen:
        warnings.append(
            "%d/%d dòng có cột '%s' KHÔNG bằng '<%s>,<%s>' — rất có thể đã đọc lệch cột. "
            "Đối chiếu lại bố cục file trước khi ghi nhận."
            % (n_desc_seen - n_desc_ok, n_desc_seen, _MG_H_DESC, _MG_H_INV, _MG_H_STORE))

    n_pay = sum(1 for r in rows if r["row_kind"] == "thanh_toan")
    net = round(sum(float(r["signed_amount"] or 0) for r in rows), 2)

    # File CẤN TRỪ HẾT: không có tiền mặt về. Nói thẳng ra — kế toán chờ tiền vào
    # tài khoản mà không thấy sẽ đi tìm nhầm chỗ.
    if abs(net) <= MONEY_EPS:
        warnings.append(
            "Bảng kê này CẤN TRỪ HẾT: %d hóa đơn bán ra (%s đ) trừ đúng bằng %d chứng "
            "từ ghi giảm — tiền thực nhận bằng 0. Không phải lỗi đọc file."
            % (n_pay,
               "{:,.0f}".format(sum(float(r["signed_amount"] or 0) for r in rows
                                    if r["row_kind"] == "thanh_toan")),
               len(rows) - n_pay))

    n_ded = len(rows) - n_pay
    if n_ded:
        warnings.append(
            "%d dòng ghi giảm của Mega KHÔNG có cột phân loại nào trong file — không "
            "tách được đâu là chiết khấu, đâu là phí, đâu là hàng trả lại. Tất cả vào "
            "một nhóm 'Ghi giảm' và sẽ hạch toán vào MỘT tài khoản theo `MT Account "
            "Map`. Cần tách thì sửa loại dòng trên chứng từ trước khi sinh bút toán."
            % n_ded)

    warnings.append(
        "File Mega Market KHÔNG in dòng tổng cộng, số bảng kê hay số tiền thanh toán "
        "nào — không có số kiểm tra để đối chiếu, nên bảng kê này luôn ở trạng thái "
        "'chưa khớp số kiểm tra'. Đối chiếu tổng %s đ với thông báo thanh toán Mega gửi "
        "rồi tự tick khi đã soi." % "{:,.0f}".format(net))
    if skipped:
        warnings.append("%d dòng bị bỏ qua vì không đọc được số tiền." % skipped)

    return {
        # Không có số bảng kê trong file — để None chứ không bịa từ tên file.
        "advice_no": None,
        "payment_dates": sorted(pay_dates),
        "declared_totals": {
            "total_payment": None,
            "total_discount": None,
            "net_payment": None,
        },
        # RỖNG có chủ đích — xem khối chú thích ở đầu mục 4h.
        "checks": [],
        "groups": [],
        "rows": rows,
        "warnings": warnings,
    }


PARSERS = {
    "wincommerce": parse_wincommerce,
    "central_retail": parse_central_retail,
    "lotte": parse_lotte,
    "emart": parse_emart,
    "coop": parse_coop,
    "aeon": parse_aeon,
    "fuji": parse_fuji,
    "mega_market": parse_mega,
}


def parse_sheets(sheets, chain_key):
    """Chạy parser của một chuỗi trên các sheet đã đọc và bổ sung phần tổng chung."""
    if chain_key not in PARSERS:
        # Phân biệt hai ca khác hẳn nhau: chuỗi lạ hoàn toàn, và chuỗi CÓ trong
        # danh sách nhưng CHƯA có tầng đọc. Hiện cả 8 chuỗi đều có parser, nhưng
        # nhánh này phải ở lại: thêm chuỗi vào `MT_CHAINS` mà chưa viết parser là
        # chuyện sẽ lặp lại, và khi đó báo sai lý do là bắt kế toán đi soi file.
        if chain_key in CHAIN_LABEL:
            frappe.throw(_(
                "Chưa có tầng đọc bảng kê cho chuỗi {0}. Gán khách vào chuỗi này thì "
                "được, nhưng nạp file thì chưa — cần một file bảng kê mẫu thật để "
                "viết parser. KHÔNG chọn tạm chuỗi khác: parser sai chuỗi đọc SAI "
                "CỘT TIỀN mà vẫn ra một con số trông hợp lý."
            ).format(CHAIN_LABEL[chain_key]))
        frappe.throw(_("Chuỗi không hợp lệ: {0}").format(chain_key))
    res = PARSERS[chain_key](sheets)
    rows = res["rows"]

    computed = {
        "thanh_toan": _sum(rows, "signed_amount", {"thanh_toan"}),
        "chiet_khau": _sum(rows, "signed_amount", {"chiet_khau"}),
        "phi": _sum(rows, "signed_amount", {"phi"}),
        "ghi_giam": _sum(rows, "signed_amount", {"ghi_giam"}),
        "khac": _sum(rows, "signed_amount", {"khac"}),
        "net": _sum(rows, "signed_amount"),
    }
    checks = res.get("checks") or []
    warnings = list(res.get("warnings") or [])
    # `reconciled` chỉ bật khi CÓ số kiểm tra và MỌI số đều khớp. Không có số kiểm
    # tra thì để False — "không kiểm được" không phải là "đã kiểm và đúng".
    reconciled = bool(checks) and all(c["ok"] for c in checks)
    if not reconciled:
        bad = [c["label"] for c in checks if not c["ok"]]
        warnings.append("CHƯA khớp số kiểm tra của file: %s. KHÔNG ghi nhận cho tới khi làm rõ."
                        % (", ".join(bad) if bad else "file không có số kiểm tra nào"))

    # LƯỚI AN TOÀN — BẢNG KÊ THANH TOÁN MÀ KHÔNG CÓ DÒNG THANH TOÁN NÀO.
    #
    # VÌ SAO cần riêng chốt này: khi một dòng bị xếp nhầm loại, tiền chỉ đổi NHÓM
    # chứ TỔNG không đổi, nên mọi số kiểm tra SUM vẫn khớp 0đ và `reconciled` vẫn
    # True. Đây là chiều hỏng câm nhất của cả module.
    #
    # Dấu hiệu chắc chắn: file CÓ dòng mang tham chiếu hóa đơn mà KHÔNG dòng nào
    # được xếp là thanh toán. Bảng kê nào cũng phải trả tiền cho ít nhất một hóa
    # đơn — không thì nó đâu còn là bảng kê thanh toán.
    has_inv_ref = any(r.get("inv_no") and r.get("inv_series") for r in rows)
    if has_inv_ref and not any(r["row_kind"] == "thanh_toan" for r in rows):
        n_ref = sum(1 for r in rows if r.get("inv_no") and r.get("inv_series"))
        warnings.append(
            "NGHI ĐỌC SAI LOẠI DÒNG: file có %d dòng mang số hóa đơn nhưng KHÔNG dòng nào "
            "được xếp là thanh toán. Rất có thể chuỗi đã đổi nhãn cột phân loại. "
            "Số kiểm tra vẫn khớp vì tiền chỉ đổi nhóm — ĐỪNG tin `reconciled` ở ca này."
            % n_ref)
        reconciled = False

    # LƯỚI AN TOÀN CHUNG CHO CẢ 5 CHUỖI — dòng thanh toán NGƯỢC DẤU với đa số.
    # VÌ SAO: mọi lá chắn phân loại đều dựa vào nhãn trong file (Doc.Type, ký hiệu
    # chứng từ, Deduct Name). Khi chuỗi đổi ký hiệu (ví dụ năm mới) hay đổi nhãn,
    # một dòng GHI GIẢM lặng lẽ trôi sang 'thanh_toan' mà MỌI số kiểm tra vẫn khớp
    # 0đ vì tổng NET không đổi. Dấu ngược chiều là dấu hiệu còn lại duy nhất, và
    # nó nguy hiểm gấp đôi vì `_paid_subquery` lấy SUM(ABS(...)) — một khoản trả
    # hàng bị lật thành TIỀN ĐÃ THU trên đúng hóa đơn đó.
    # Đây CHỈ là cờ 'cần người xem lại', KHÔNG tự đổi loại dòng — phân loại bằng
    # dấu là điều cấm của hợp đồng (§B). Trên cả 7 file thật, số dòng bị gắn cờ = 0.
    pay_rows = [r for r in rows if r["row_kind"] == "thanh_toan" and r.get("signed_amount")]
    n_pos = sum(1 for r in pay_rows if r["signed_amount"] > 0)
    n_neg = len(pay_rows) - n_pos
    if n_pos and n_neg:
        minority = 1 if n_pos < n_neg else (-1 if n_neg < n_pos else 0)
        odd = [r for r in pay_rows
               if minority and (1 if r["signed_amount"] > 0 else -1) == minority]
        for r in odd:
            r["needs_review"] = True
        warnings.append(
            "%d/%d dòng 'Thanh toán' NGƯỢC DẤU với đa số (%d dương / %d âm) — nghi bị "
            "xếp nhầm loại (trả hàng/điều chỉnh đọc thành tiền thu). Số kiểm tra KHÔNG "
            "phát hiện được lỗi này vì tổng không đổi. Phải xem tay trước khi khớp hóa đơn."
            % (len(odd) or min(n_pos, n_neg), len(pay_rows), n_pos, n_neg))

    n_review = sum(1 for r in rows if r["needs_review"])
    if n_review:
        warnings.append("%d/%d dòng cần người xem lại trước khi khớp hóa đơn."
                        % (n_review, len(rows)))

    return {
        "chain": CHAIN_LABEL[chain_key],
        "chain_key": chain_key,
        "advice_no": res.get("advice_no"),
        "payment_dates": res.get("payment_dates") or [],
        "declared_totals": res.get("declared_totals") or {},
        "computed_totals": computed,
        "checks": checks,
        "reconciled": reconciled,
        "groups": res.get("groups") or [],
        "rows": rows,
        "warnings": warnings,
    }


def read_payment_advice(content, chain=None):
    """Đọc bảng kê thanh toán MT từ base64 -> dict (xem docstring đầu module).

    `chain`: khóa ASCII ('lotte'...) hoặc nhãn ('LOTTE'...). Bỏ trống thì tự nhận;
    không nhận ra được thì THROW để người chọn tay, không đoán bừa — parser sai
    chuỗi vẫn ra một con số trông hợp lý nhưng đọc sai cột tiền.
    """
    sheets = read_sheets(content)
    if chain:
        key = _resolve_chain_key(chain)
        # Người đã CHỌN chuỗi mà báo "không nhận ra chuỗi từ file" là thông báo
        # sai địa chỉ — họ sẽ đi soi file trong khi lỗi nằm ở tên chuỗi gửi lên.
        if not key:
            frappe.throw(_("Không có chuỗi siêu thị tên '{0}'. Chọn lại trong danh sách.")
                         .format(chain))
    else:
        key = detect_chain(sheets)
        if not key:
            frappe.throw(_("Không nhận ra chuỗi siêu thị từ file. Hãy chọn chuỗi bằng tay."))
    out = parse_sheets(sheets, key)
    out["sheets"] = [{"name": n, "rows": len(g)} for n, g in sheets]
    if chain is None:
        out["chain_detected"] = True
    return out


def _resolve_chain_key(chain):
    """Nhãn ('AEON') hoặc khóa ASCII ('aeon') -> khóa ASCII. Không nhận ra -> None.

    Nhận diện theo `CHAIN_LABEL` chứ KHÔNG theo `PARSERS`: chuỗi đã có trong danh
    sách nhưng chưa có tầng đọc vẫn phải phân giải được, để `parse_sheets` báo
    đúng lý do "chưa có parser" thay vì "không nhận ra chuỗi".
    """
    c = _txt(chain)
    if c in CHAIN_LABEL:
        return c
    for k, label in CHAIN_LABEL.items():
        if strip_tones(label) == strip_tones(c):
            return k
    return None


# ═══════════════════════════════════════════════════════════════════════════
# 6. Whitelisted — CHỈ ĐỌC, không ghi gì
# ═══════════════════════════════════════════════════════════════════════════

MAX_PREVIEW = 30


@frappe.whitelist()
def preview(content, chain=None):
    """Đọc file bảng kê MT và trả bản xem trước để kế toán DUYỆT trước khi ghi nhận.

    Guard ở DÒNG ĐẦU theo quy ước dự án. Đây là kênh MT nên dùng guard_mt().
    Method này KHÔNG ghi database, KHÔNG tạo/sửa/hủy chứng từ kế toán nào — mọi
    quyết định hạch toán do con người làm ở bước sau.
    """
    guard_mt()
    res = read_payment_advice(content, chain=chain)
    # Không trả nguyên vài trăm dòng vào màn hình xem trước; số liệu tổng và số
    # kiểm tra mới là thứ kế toán cần nhìn để quyết định.
    res["total_rows"] = len(res["rows"])
    res["sample"] = res["rows"][:MAX_PREVIEW]
    del res["rows"]
    return res


@frappe.whitelist()
def detect(content):
    """Chỉ nhận diện chuỗi + liệt kê sheet, không bóc dòng. Dùng cho bước chọn chuỗi."""
    guard_mt()
    sheets = read_sheets(content)
    key = detect_chain(sheets)
    return {
        "chain_key": key,
        "chain": CHAIN_LABEL.get(key),
        "sheets": [{"name": n, "rows": len(g)} for n, g in sheets],
    }
