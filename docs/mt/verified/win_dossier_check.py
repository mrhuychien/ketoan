"""Kiểm HỒ SƠ THANH TOÁN WINMART (`ketoan/api/mt_win.py`).

    python3 docs/mt/verified/win_dossier_check.py

Win chỉ xử lý thanh toán khi nhận đủ bảng kê + file PDF hóa đơn ĐẶT ĐÚNG TÊN.
Sai một chữ là hồ sơ bị trả về và cả đợt trượt kỳ thanh toán. Phép kiểm vì vậy
so THẲNG với chính file mẫu Win gửi, không so với đặc tả viết lại:

  1. TÊN CỘT phải trùng file mẫu tới từng chữ, đúng thứ tự, header ở DÒNG 2.
  2. TÊN FILE PDF phải đúng khuôn `YYYYMMDD_<mã NCC>_<NN>_PF` — kể cả hậu tố
     `_PF` mà §2.2 SOP viết gọn đã bỏ mất.
  3. FILE XUẤT RA phải đọc lại được và ra đúng số tiền đã đưa vào.
  4. MỘT HÓA ĐƠN CHỈ NỘP MỘT LẦN — trùng trong hồ sơ, hoặc đã ở hồ sơ khác.
  5. STT do kế toán sửa thì KHÔNG bị đánh lại (mẫu thật đánh không theo thứ tự).

Chạy KHÔNG cần bench — stub frappe của `regression_check`.
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import regression_check as rc  # noqa: E402

SAMPLE = "Mẫu bảng kê ghi nhận hồ sơ thanh toán Winmart.xlsx"

# Đọc từ file mẫu thật: header nằm ở dòng 2, mười cột, đúng thứ tự này.
SAMPLE_HEADER_ROW = 2
SAMPLE_PDF_NAME = "20260817_2007766_01_PF"
SAMPLE_VENDOR = "2007766"


class _D(dict):
    def __getattr__(self, k):
        try:
            return self[k]
        except KeyError:
            raise AttributeError(k)

    def __setattr__(self, k, v):
        self[k] = v


def _sample_columns():
    import openpyxl

    wb = openpyxl.load_workbook(os.path.join(rc.SAMPLES, SAMPLE), data_only=True, read_only=True)
    sh = wb[wb.sheetnames[0]]
    rows = list(sh.iter_rows(min_row=1, max_row=4, values_only=True))
    wb.close()
    header = [str(v).strip() for v in rows[SAMPLE_HEADER_ROW - 1] if v not in (None, "")]
    first = rows[SAMPLE_HEADER_ROW]      # dòng dữ liệu đầu tiên
    return header, first


def _make_classes():
    """Dựng bản giả của hai controller: KẾ THỪA lớp thật, chỉ bỏ __init__ của Document.

    Kế thừa chứ không nhại: phép kiểm phải chạy ĐÚNG mã sẽ chạy thật, kể cả các
    hàm phụ (`_number_lines`, `_check_duplicate_invoices`…). Nhại lại là đo một
    thứ khác với thứ sẽ chạy.
    """
    from ketoan.mt.doctype.mt_win_dossier.mt_win_dossier import MTWinDossier
    from ketoan.mt.doctype.mt_win_dossier_line.mt_win_dossier_line import MTWinDossierLine

    class FakeLine(MTWinDossierLine):
        def __init__(self, **kw):
            self.__dict__.update(kw)

        def get(self, k, d=None):
            return self.__dict__.get(k, d)

    class FakeDossier(MTWinDossier):
        def __init__(self, **kw):
            self.__dict__.update(kw)

        def get(self, k, d=None):
            return self.__dict__.get(k, d)

    return FakeDossier, FakeLine


def _fake_doc(lines, vendor=SAMPLE_VENDOR, submit="2026-08-17", no=1, status="Nháp"):
    Dossier, Line = _make_classes()
    rows = []
    for i, x in enumerate(lines, start=1):
        kw = dict(x)
        kw.setdefault("idx", i)
        for f in ("stt", "po_vcm", "inv_series", "inv_no", "inv_date",
                  "amount_before_vat", "vat_amount", "total_amount",
                  "pdf_name", "sales_invoice"):
            kw.setdefault(f, None)
        rows.append(Line(**kw))
    return Dossier(name="MT-HSW-00001", company="HGC", customer="KH-WIN",
                   vendor_code=vendor, submit_date=submit, dossier_no=no,
                   status=status, file_prefix=None,
                   period_from="2026-08-01", period_to="2026-08-17",
                   total_before_vat=0, total_vat=0, total_amount=0, lines=rows)


def main():
    rc._stub_frappe()
    sys.path.insert(0, rc.REPO)
    import importlib

    import frappe

    frappe.db.table_exists = lambda dt: True
    frappe.db.sql = lambda *a, **k: []
    frappe.db.has_column = lambda dt, col: True
    # guard_mt() đọc vai trò; stub mặc định trả [] nên mọi method whitelisted
    # đều bị chặn. Gắn vai trò kế toán trưởng để kiểm được PHẦN THÂN hàm.
    frappe.get_roles = lambda *a, **kw: ["Ke Toan Truong"]

    mw = importlib.import_module("ketoan.api.mt_win")
    mwd = importlib.import_module("ketoan.mt.doctype.mt_win_dossier.mt_win_dossier")

    print("=" * 78)
    print("KIỂM HỒ SƠ THANH TOÁN WINMART")
    print("=" * 78)
    bad = 0

    # ── 1. Tên cột trùng file mẫu tới từng chữ ────────────────────────────
    header, first = _sample_columns()
    ok = list(mw.WIN_COLUMNS) == header
    print(f"  {'✅' if ok else '❌'} 10 tên cột trùng file mẫu, đúng thứ tự, header ở dòng {SAMPLE_HEADER_ROW}")
    if not ok:
        print(f"       └─ mẫu:  {header}")
        print(f"       └─ code: {list(mw.WIN_COLUMNS)}")
    bad += not ok

    # ── 2. Tên file PDF đúng khuôn, KỂ CẢ hậu tố _PF ──────────────────────
    print("-" * 78)
    got = mwd.build_prefix("2026-08-17", SAMPLE_VENDOR, 1)
    ok = got == SAMPLE_PDF_NAME
    print(f"  {'✅' if ok else '❌'} tên file: {got} (mẫu thật {SAMPLE_PDF_NAME})")
    bad += not ok

    ok = got.endswith("_PF")
    print(f"  {'✅' if ok else '❌'} có hậu tố `_PF` — §2.2 SOP viết gọn đã bỏ mất")
    bad += not ok

    got2 = mwd.build_prefix("2026-08-17", SAMPLE_VENDOR, 12)
    ok = got2 == "20260817_2007766_12_PF"
    print(f"  {'✅' if ok else '❌'} hồ sơ thứ 12 trong ngày -> {got2} (hai chữ số)")
    bad += not ok

    # Cột 'Tên File PDF' của mẫu giống nhau ở mọi dòng -> nó định danh HỒ SƠ.
    sample_pdf = str(first[9]).strip() if first and len(first) > 9 else ""
    ok = sample_pdf == SAMPLE_PDF_NAME
    print(f"  {'✅' if ok else '❌'} dòng đầu của mẫu mang đúng tên file {sample_pdf!r}")
    bad += not ok

    # ── 3. Xuất Excel rồi ĐỌC LẠI, phải ra đúng số ───────────────────────
    print("-" * 78)
    lines = [
        {"stt": 3, "po_vcm": "4193862223", "inv_series": "1C26THG", "inv_no": "6999",
         "inv_date": "2026-08-14", "amount_before_vat": 640000, "vat_amount": 51200,
         "pdf_name": SAMPLE_PDF_NAME, "sales_invoice": "SI-1"},
        {"stt": 4, "po_vcm": "4193876257", "inv_series": "1C26THG", "inv_no": "7006",
         "inv_date": "2026-08-14", "amount_before_vat": 10383500, "vat_amount": 830680,
         "pdf_name": SAMPLE_PDF_NAME, "sales_invoice": "SI-2"},
    ]
    doc = _fake_doc(lines)
    doc.validate()

    captured = {}
    frappe.local.response = _D()
    frappe.get_doc = lambda dt, name=None, **k: doc
    mw._company = lambda company=None: "HGC"
    mw.export_dossier("MT-HSW-00001")
    captured["name"] = frappe.local.response.get("filename")
    captured["content"] = frappe.local.response.get("filecontent")

    ok = captured["name"] == SAMPLE_PDF_NAME + ".xlsx"
    print(f"  {'✅' if ok else '❌'} tên file Excel xuất ra: {captured['name']}")
    bad += not ok

    import io

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(captured["content"]), data_only=True)
    sh = wb[wb.sheetnames[0]]
    out_header = [sh.cell(row=SAMPLE_HEADER_ROW, column=c).value for c in range(1, 11)]
    ok = out_header == list(mw.WIN_COLUMNS)
    print(f"  {'✅' if ok else '❌'} file xuất ra có header ở đúng dòng {SAMPLE_HEADER_ROW}, đủ 10 cột")
    bad += not ok

    got_rows = [[sh.cell(row=r, column=c).value for c in range(1, 11)] for r in (3, 4)]
    ok = (got_rows[0][0] == 3 and got_rows[0][1] == SAMPLE_VENDOR
          and got_rows[0][2] == "4193862223" and got_rows[0][4] == "6999"
          and got_rows[0][6] == 640000 and got_rows[0][7] == 51200
          and got_rows[0][8] == 691200 and got_rows[0][9] == SAMPLE_PDF_NAME)
    print(f"  {'✅' if ok else '❌'} dòng đầu đọc lại đúng: STT {got_rows[0][0]} · "
          f"HĐ {got_rows[0][4]} · {got_rows[0][8]:,} đ · {got_rows[0][9]}")
    if not ok:
        print(f"       └─ {got_rows[0]}")
    bad += not ok

    ok = (round(doc.total_before_vat) == 11023500 and round(doc.total_vat) == 881880
          and round(doc.total_amount) == 11905380)
    print(f"  {'✅' if ok else '❌'} tổng: {doc.total_before_vat:,.0f} + {doc.total_vat:,.0f} "
          f"= {doc.total_amount:,.0f}")
    bad += not ok

    # ── 4. Một hóa đơn chỉ nộp MỘT lần ───────────────────────────────────
    print("-" * 78)
    dup = _fake_doc(lines + [dict(lines[0])])
    try:
        dup.validate()
        print("  ❌ hóa đơn trùng TRONG hồ sơ -> KHÔNG dừng")
        bad += 1
    except Exception as e:  # noqa: BLE001
        ok = "hai lần" in str(e)
        print(f"  {'✅' if ok else '❌'} hóa đơn trùng trong cùng hồ sơ -> dừng")
        bad += not ok

    frappe.db.sql = lambda *a, **k: [_D(sales_invoice="SI-1", parent="MT-HSW-00009")]
    try:
        _fake_doc(lines).validate()
        print("  ❌ hóa đơn đã ở hồ sơ khác -> KHÔNG dừng")
        bad += 1
    except Exception as e:  # noqa: BLE001
        ok = "MT-HSW-00009" in str(e) and "MỘT lần" in str(e)
        print(f"  {'✅' if ok else '❌'} hóa đơn đã nằm ở hồ sơ khác -> dừng, nêu đích danh hồ sơ")
        bad += not ok
    frappe.db.sql = lambda *a, **k: []

    # ── 5. STT: sinh cho dòng trống, GIỮ dòng kế toán đã sửa ─────────────
    print("-" * 78)
    mixed = _fake_doc([
        {"stt": 9, "inv_no": "A", "amount_before_vat": 100, "vat_amount": 8},
        {"stt": 0, "inv_no": "B", "amount_before_vat": 100, "vat_amount": 8},
        {"stt": 3, "inv_no": "C", "amount_before_vat": 100, "vat_amount": 8},
        {"stt": 0, "inv_no": "D", "amount_before_vat": 100, "vat_amount": 8},
    ])
    mixed.validate()
    got = [l.stt for l in mixed.lines]
    ok = got == [9, 1, 3, 2]
    print(f"  {'✅' if ok else '❌'} STT: giữ 9 và 3 của kế toán, điền 1 và 2 vào chỗ trống "
          f"-> {got}")
    bad += not ok

    # ── 6. Hồ sơ ĐÃ NỘP thì không xóa được ───────────────────────────────
    sent = _fake_doc(lines, status="Đã nộp")
    try:
        sent.on_trash()
        print("  ❌ hồ sơ đã nộp -> vẫn xóa được")
        bad += 1
    except Exception as e:  # noqa: BLE001
        print(f"  ✅ hồ sơ đã nộp -> không xóa được ({str(e)[:44]}…)")

    # ── 7. Thiếu mã NCC / ngày nộp -> DỪNG (tên file cần cả hai) ─────────
    for kw, what in (({"vendor": ""}, "thiếu mã NCC"), ({"submit": None}, "thiếu ngày nộp")):
        try:
            _fake_doc(lines, **kw).validate()
            print(f"  ❌ {what} -> KHÔNG dừng")
            bad += 1
        except Exception:  # noqa: BLE001
            print(f"  ✅ {what} -> dừng (tên file PDF cần thông tin này)")

    print("=" * 78)
    print("KẾT QUẢ:", "ĐẠT — file xuất ra đúng khuôn Win, tên file đúng chuẩn, không nộp trùng"
          if not bad else f"HỎNG {bad} mục")
    return 1 if bad else 0


if __name__ == "__main__":
    sys.exit(main())
