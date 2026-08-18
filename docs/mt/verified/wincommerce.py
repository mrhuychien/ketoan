"""Parser ĐÃ CHẠY THẬT trên file mẫu của chuỗi wincommerce — bản tham chiếu.

verified = True
sheet = 'Table 5 + Table 7 + Table 8 (BA sheet dữ liệu, không phải một). Table 1-4 = khối thông tin, Table 6 = chân trang chuyển số, Table 9 = chứng từ thanh toán. Tổng 9 sheet — KHÔNG phải 7 như gợi ý.'
header_row = 1

columns:
  doc_no -> 'Số chứng từ'
  invoice_raw -> 'Số hóa đơn'
  inv_date -> 'Ngày hóa đơn'
  recon_no -> 'Số đối soát'
  discount -> 'Chiết khấu'
  total_amount -> 'Số tiền'
  payment_date -> "Table 2 khối text 'Ngày thanh toán' + Table 9 cột 'Ngày'"
  payment_doc -> "Table 9 cột 'Chứng từ thanh toán' (= 2000141337, trùng 'Số đối soát' mọi dòng)"
  amount_before_vat -> 'KHÔNG CÓ trong file'
  vat_amount -> 'KHÔNG CÓ trong file'
  store_code -> 'KHÔNG CÓ trong file'
  store_name -> 'KHÔNG CÓ trong file'

invoice_parse: Cột 'Số hóa đơn' có dạng '<ký hiệu>#<số>'. Tách tại ký tự '#' ĐẦU TIÊN: trái = inv_series, phải = inv_no. Ví dụ thật trong file: '1C26THG#1730' -> ('1C26THG','1730'); '1C26THG#2355' -> ('1C26THG','2355'); '1C26THG#2723' -> ('1C26THG','2723'). Không có '#' thì TRẢ inv_series='' và đưa cả chuỗi vào inv_no — tuyệt đối không đoán ký hiệu (đoán là gán nhầm hóa đơn). Lưu ý dùng '#' chứ không dùng regex 'ký tự không phải chữ số' chung với Central Retail ('|'). Ở TẦNG KHỚP (không phải tầng đọc) phải áp §E của hợp đồng: bỏ chữ số dạng hóa đơn ở đầu ký hiệu, tức '1C26THG' và 'C26THG' là MỘT. File này chỉ chứa '1C26THG' nên bài học đó chưa lộ ra ở đây.
date_parse: Cột 'Ngày hóa đơn' đã là datetime.datetime thật của Excel (data_type='d') ở cả 36/36 dòng — dùng .date().isoformat(). KHÔNG có dòng nào ngày là chuỗi. Ngày thanh toán có HAI nguồn: (1) Table 2 r2, khối text nhiều dòng, nhãn 'Ngày thanh toán' nằm ở dòng trên và giá trị '25.06.2026' ở dòng NGAY DƯỚI (định dạng dd.mm.yyyy, dấu chấm) — phải split('\n') rồi lấy dòng kế; (2) Table 9 r2 cột 'Ngày' là datetime thật 2026-06-25. Hai nguồn khớp nhau. Parser ưu tiên nguồn (2) vì là kiểu ngày thật, không phụ thuộc parse chuỗi.
amount_parse: Cột 'Số tiền' và 'Chiết khấu' trên các dòng dữ liệu là số nguyên Excel thật (int), không phải chuỗi — dùng trực tiếp. Toàn bộ DƯƠNG, không có số âm nào trong file. File KHÔNG tách VAT: chỉ một cột tiền duy nhất (đã gồm thuế) => amount_before_vat=None và vat_amount=None, tuyệt đối không tự chia 1.1/1.08 để suy ra. Riêng ô tổng ở Table 9 r2 là CHUỖI '******245.795.904*': phải strip('*') rồi bỏ dấu '.' (phân cách nghìn kiểu VN) và đổi ',' thành '.' (thập phân) trước khi ép float. Chuỗi trong Table 6 cũng vậy: '70.880.508'.
payment_dates: ['2026-06-25']

row_types:
  {"kind": "thanh_toan", "label": "Dòng hóa đơn được thanh toán", "count": 36, "sign": "dương", "rule": "Sau dòng header, có 'Số hóa đơn' dạng <ký hiệu>#<số> VÀ 'Số tiền' ép được sang số. Phân bố: Table 5 = 11 dòng (r2,4,...,22), Table 7 = 21 dòng (r2,4,...,42), Table 8 = 4 dòng (r2,4,6,8)."}
  {"kind": "bo_qua", "label": "Dòng phân cách '**********'", "count": 36, "sign": "không có tiền", "rule": "Ô cột 'Ngày hóa đơn' là chuỗi TOÀN dấu sao. Đúng 1 dòng sao sau MỖI dòng dữ liệu (11+21+4). Chỉ là đường kẻ của bản in PDF."}
  {"kind": "bo_qua", "label": "Dòng tiêu đề cột của bảng dữ liệu", "count": 3, "sign": "không có tiền", "rule": "Table 5 r1, Table 7 r1, Table 8 r1 — chứa đủ 'Số hóa đơn' và 'Số tiền'."}
  {"kind": "bo_qua", "label": "Khối thông tin bên mua/bên bán/thanh toán/lời chào", "count": 5, "sign": "không có tiền", "rule": "Table 1 r1, Table 2 r1, Table 2 r2, Table 3 r1, Table 4 r1 — mỗi ô là một khối text nhiều dòng ngăn bởi \\n. Table 2 r2 chứa 'Ngày thanh toán\\n25.06.2026'."}
  {"kind": "bo_qua", "label": "Chân trang 'Số dư mang sang trang sau' (SỐ KIỂM TRA, không phải tiền mới)", "count": 1, "sign": "dương nhưng KHÔNG được cộng", "rule": "Table 6 r1, ô text chứa 'Số dư mang sang trang sau ... 0 ... 70.880.508'. Bằng ĐÚNG tổng Table 5. Cộng vào là nhân đôi tiền trang 1."}
  {"kind": "bo_qua", "label": "Dòng trống", "count": 1, "sign": "không có tiền", "rule": "Table 6 r2 — mọi ô None."}
  {"kind": "bo_qua", "label": "Mảnh tiêu đề chân trang lặp lại", "count": 1, "sign": "không có tiền", "rule": "Table 7 r44: chỉ có 'Chiết khấu' (cột E) và 'Số tiền' (cột G), không có số chứng từ, không có tiền. Là header của trang in kế tiếp bị cắt rơi lại."}
  {"kind": "bo_qua", "label": "Dòng 'Tổng cộng' (SỐ KIỂM TRA)", "count": 1, "sign": "dương nhưng KHÔNG được cộng", "rule": "Table 8 r10: cột A = 'Tổng cộng', cột 'Chiết khấu' = 0, cột 'Số tiền' = 245.795.904."}
  {"kind": "bo_qua", "label": "Bảng chứng từ thanh toán (Table 9)", "count": 2, "sign": "dương nhưng KHÔNG được cộng", "rule": "r1 header 'Chứng từ thanh toán | Ngày | Đơn vị tiền tệ | Số tiền'; r2 = 2000141337 | 2026-06-25 | VND | '******245.795.904*'. Là tổng lần thanh toán, số dạng CHUỖI có sao bao quanh."}
  {"kind": "chiet_khau", "label": "Chiết khấu (cột riêng trên chính dòng hóa đơn)", "count": 0, "sign": "dương, ghi giảm", "rule": "Cột 'Chiết khấu' TỒN TẠI trên mọi dòng dữ liệu nhưng trong file này = 0 ở cả 36/36 dòng. Parser đã có nhánh tách thành dòng riêng khi khác 0, nhưng nhánh đó CHƯA được xác minh bằng dữ liệu thật."}

totals:
  tong_dong_vat_ly_trong_file = 86
  tong_dong_parser_tra_ve = 85
  chenh_lech_dong = 1 — dòng trống Table 6 r2 bị bỏ, không emit. Cố ý.
  thanh_toan_Table5 = 70880508
  thanh_toan_Table7 = 134593596
  thanh_toan_Table8 = 40321800
  thanh_toan_TONG = 245795904
  so_dong_thanh_toan = 36
  chiet_khau_TONG = 0
  kiem_tra_1_carry_forward_Table6 = 70880508
  kiem_tra_1_khop = ĐÚNG — bằng chính xác tổng Table 5, xác nhận Table 6 là chân trang chứ không phải tiêu đề mục Chiết khấu
  kiem_tra_2_Tong_cong_Table8_r10 = 245795904
  kiem_tra_2_khop = ĐÚNG — bằng T5+T7+T8
  kiem_tra_3_Chung_tu_thanh_toan_Table9 = 245795904
  kiem_tra_3_khop = ĐÚNG — chuỗi '******245.795.904*' sau khi bóc sao và bỏ dấu chấm nghìn
  doc_no_duy_nhat = 36/36 — không có Số chứng từ trùng
  inv_no_duy_nhat = 36/36 — không có số hóa đơn trùng
  ky_hieu_hoa_don = chỉ một giá trị '1C26THG' trên cả 36 dòng
  ngay_hoa_don_phan_bo = ['2026-02-27', '2026-03-22', '2026-04-04', '2026-04-07']
  ngay_thanh_toan = 2026-06-25 (duy nhất, hai nguồn Table 2 và Table 9 khớp nhau)
  ket_luan = KHỚP TUYỆT ĐỐI ở cả 3 điểm kiểm tra độc lập, lệch 0 đồng.

BẪY:
  - BẪY LỚN NHẤT — GỢI Ý BAN ĐẦU SAI: file có 9 sheet chứ không phải 7, và Table 6 KHÔNG phải tiêu đề mục 'Chiết khấu'. Table 6 là CHÂN TRANG của bản in: 'Chiết khấu   Số tiền / Số dư mang sang trang sau   0   70.880.508'. Hai chữ 'Chiết khấu' ở đó là nhãn CỘT, không phải nhãn MỤC. Bằng chứng số học: 70.880.508 = đúng tổng Table 5. Suy ra Table 7 (và Table 8) là PHẦN TIẾP THEO của cùng bảng thanh toán, KHÔNG phải danh sách chiết khấu. Nếu tin gợi ý mà gán Table 7 = chiet_khau thì 134.593.596đ bị ghi nhận sai loại, và 40.321.800đ của Table 8 bị bỏ quên hoàn toàn.
  - BỎ SÓT SHEET = BỎ SÓT TIỀN: chỉ đọc Table 5 là mất 174.915.396đ (71% số tiền). Đây là file PDF convert sang Excel, MỖI TRANG IN thành một sheet, nên bảng dữ liệu bị cắt rời. Phải quét MỌI sheet và dò header, không được hardcode tên sheet.
  - LỆCH CỘT GIỮA CÁC SHEET: Table 7 có THÊM một cột A rỗng (merged A1:A42) nên mọi cột dịch phải 1 ô so với Table 5 và Table 8 (Số tiền ở index 6 thay vì 5). Hardcode chỉ số cột là đọc lệch cột tiền của 21/36 dòng. Bắt buộc dò dòng header theo NHÃN rồi map cột.
  - HAI SỐ KIỂM TRA DỄ BỊ CỘNG NHẦM THÀNH TIỀN: 'Số dư mang sang trang sau 70.880.508' (Table 6) và 'Tổng cộng 245.795.904' (Table 8 r10). Cộng cả hai vào là ra 561.472.316đ thay vì 245.795.904đ. Phải nhận diện và loại, rồi DÙNG chúng làm chốt đối chiếu.
  - DÒNG SAO KHÔNG PHẢI LÚC NÀO CŨNG 10 SAO: dòng phân cách trong bảng là '**********' (10 sao) nhưng Table 4 r1 chứa '*****************************************' (41 sao) và Table 9 r2 chứa '******245.795.904*'. So sánh cứng chuỗi 10 sao là mong manh; kiểm tra 'tập ký tự == {*}' thì Table 9 lại KHÔNG dính (vì có chữ số) — đúng như mong muốn.
  - SỐ TIỀN TỔNG NẰM TRONG CHUỖI, KHÔNG PHẢI SỐ: Table 9 r2 cột 'Số tiền' là chuỗi '******245.795.904*'. Ép float thẳng là ValueError. Và dấu '.' ở đây là phân cách NGHÌN kiểu VN, không phải thập phân — bỏ nhầm quy ước là sai 1000 lần.
  - MẢNH HEADER MỒ CÔI: Table 7 r44 chỉ có 'Chiết khấu' (cột E) và 'Số tiền' (cột G), không số chứng từ, không tiền. Nếu chỉ lọc 'dòng nào có chữ Chiết khấu thì là mục chiết khấu' sẽ tạo ra một dòng chiết khấu ma.
  - KHÔNG CÓ VAT, KHÔNG CÓ CỬA HÀNG: file chỉ có MỘT cột tiền, không tách trước/sau thuế; và không có mã/tên siêu thị thành viên nào. amount_before_vat, vat_amount, store_code, store_name buộc phải để None/'' — tự suy ra là bịa số.
  - 'Số đối soát' (2000141337) GIỐNG HỆT NHAU ở cả 36 dòng và TRÙNG với 'Chứng từ thanh toán' ở Table 9. Nó là mã lần thanh toán, KHÔNG phải khóa dòng. Dùng nó làm khóa khớp hóa đơn là gộp nhầm 36 dòng làm một.
  - CỘT 'Chiết khấu' TỒN TẠI TRÊN MỌI DÒNG nhưng bằng 0 hết. Đừng kết luận 'WinCommerce không có chiết khấu' — chỉ là kỳ thanh toán này không có. Cấu trúc cột riêng đúng như §C hợp đồng đã ghi.
  - SỐ DÒNG TRONG HỢP ĐỒNG (§A: '23 + 44') THIẾU: thực tế 23 + 44 + 10 (Table 8) + 2 (Table 9) + 5 (Table 1-4, 6) = 86 dòng vật lý. Cần sửa lại hợp đồng.
  - Ngày thanh toán ở Table 2 nằm TRONG một ô text nhiều dòng, nhãn và giá trị ở hai dòng \n khác nhau, định dạng dd.mm.yyyy. Không có cột riêng để đọc.

CHƯA XÁC MINH:
  - Nhánh chiet_khau CHƯA chạy được lần nào: cột 'Chiết khấu' bằng 0 ở cả 36/36 dòng của file này. Chưa biết chiết khấu khác 0 hiển thị thế nào (dương hay âm, có kèm dòng mô tả riêng không). Cần thêm một file WinCommerce có chiết khấu thật.
  - Chưa biết WinCommerce có dòng PHÍ / GHI GIẢM riêng hay không — file này không có dòng nào ngoài hóa đơn hàng hóa.
  - Chưa xác minh file WinCommerce nhiều kỳ thanh toán: file này chỉ có MỘT ngày thanh toán (25.06.2026). §H hợp đồng cảnh báo LOTTE có nhiều Payment Date — chưa biết WinCommerce có bao giờ gộp nhiều kỳ vào một file không.
  - Chưa xác minh số sheet biến thiên: file này 9 sheet. Số sheet phụ thuộc số TRANG IN, nên file nhiều hóa đơn hơn sẽ có nhiều sheet hơn (Table 10, 11...). Parser dò header nên tự thích nghi, nhưng chưa được thử trên file có số sheet khác.
  - Chưa xác minh ký hiệu không nhất quán trong CHÍNH file WinCommerce: ở đây chỉ có '1C26THG'. §E hợp đồng ghi nhận Central Retail và LOTTE có lẫn 'C26THG'/'1C26THG'. Chưa biết WinCommerce có lẫn không.
  - Chưa xác minh số tiền âm: toàn bộ 36 dòng đều dương. _num() có xử lý dấu ngoặc (1.234) và dấu trừ nhưng chưa có dữ liệu thật để kiểm.
  - 'Số đối soát' 2000141337 giống hệt 'Chứng từ thanh toán' ở Table 9 — nhiều khả năng là cùng một mã lần thanh toán, nhưng CHƯA có xác nhận từ kế toán WinCommerce (đúng như §I hợp đồng đã treo).
  - 'Mã công nợ của quý công ty tại hệ thống chúng tôi: / 2007766 /' (Table 2) — chưa rõ ánh xạ sang Customer nào của ERPNext.
  - Không có mã/tên siêu thị thành viên trong file => không thể đối chiếu theo cửa hàng. Chưa biết WinCommerce có bao giờ xuất bản kê chi tiết theo siêu thị không.
  - Chưa xác minh việc file có thể có ô công thức: đã đọc bằng data_only=True nên nếu file do máy tạo mà chưa cache giá trị thì sẽ ra None. File này là convert từ PDF nên toàn giá trị tĩnh, không rủi ro — nhưng file kỳ khác thì chưa biết.
"""

# -*- coding: utf-8 -*-
"""Đọc bảng kê thanh toán WinCommerce (.xlsx, nhiều sheet 'Table N').

VÌ SAO viết riêng cho WinCommerce: file này là bản in PDF được convert sang
Excel, mỗi TRANG IN thành một sheet. Bảng dữ liệu bị CẮT RỜI ra nhiều sheet
(Table 5, 7, 8) xen giữa là sheet chân trang (Table 6). Đọc mỗi Table 5 là
mất 2/3 số tiền.
"""
import datetime
import re

import openpyxl

# Nhãn cột thật đọc từ file (không suy đoán). Fieldname ASCII, label tiếng Việt.
COL_MAP = {
    "Số chứng từ": "doc_no",
    "Số hóa đơn": "invoice_raw",
    "Ngày hóa đơn": "inv_date",
    "Số đối soát": "recon_no",
    "Chiết khấu": "discount",
    "Số tiền": "amount",
}
# Bắt buộc phải thấy đủ 2 nhãn này mới coi là dòng header của bảng dữ liệu.
HEADER_MARKERS = ("Số hóa đơn", "Số tiền")

KEYS = [
    "row_kind", "inv_series", "inv_no", "inv_date", "store_code", "store_name",
    "doc_no", "amount_before_vat", "vat_amount", "total_amount",
    "payment_date", "description",
]


def _s(v):
    """Chuẩn hoá về str đã strip; None -> ''. Giữ nguyên số để caller tự ép."""
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def _is_star(v):
    """Dòng '**********' xen giữa mỗi dòng dữ liệu — chỉ là đường kẻ của bản in.

    VÌ SAO kiểm tra 'toàn dấu sao' chứ không so bằng chuỗi 10 sao: bản in còn
    có dòng 41 sao ở Table 4. Đếm cứng số sao là bỏ sót.
    """
    t = _s(v)
    return len(t) >= 3 and set(t) == {"*"}


def _num(v):
    """Ép số tiền. Chuỗi kiểu '******245.795.904*' dùng '.' làm phân cách nghìn."""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    t = _s(v)
    if not t:
        return None
    neg = t.startswith("(") and t.endswith(")")
    t = t.strip("*").strip("()").replace(" ", "")
    # VÌ SAO bỏ dấu chấm: định dạng VN dùng '.' cho hàng nghìn, ',' cho thập phân.
    t = t.replace(".", "").replace(",", ".")
    if not re.fullmatch(r"-?\d+(\.\d+)?", t or ""):
        return None
    x = float(t)
    return -x if neg else x


def _date(v):
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    t = _s(v)
    m = re.fullmatch(r"(\d{1,2})[./-](\d{1,2})[./-](\d{4})", t)
    if m:
        d, mo, y = (int(x) for x in m.groups())
        return datetime.date(y, mo, d).isoformat()
    return None


def split_invoice(raw):
    """'1C26THG#1730' -> ('1C26THG', '1730').

    VÌ SAO tách bằng '#': WinCommerce dùng '#', Central Retail dùng '|'. Dùng
    chung một regex 'ký tự không phải chữ/số' sẽ nuốt nhầm dấu khác. Không có
    '#' thì TRẢ VỀ ký hiệu rỗng chứ không đoán — đoán ký hiệu là gán nhầm HĐ.
    """
    t = _s(raw)
    if not t:
        return "", ""
    if "#" in t:
        a, b = t.split("#", 1)
        return a.strip(), b.strip()
    return "", t


def _blank(row):
    return all(_s(c) == "" for c in row)


def _find_header(rows):
    """Trả (chỉ số dòng header, dict {tên_chuẩn: chỉ số cột}) hoặc (None, None).

    VÌ SAO dò header thay vì hardcode cột: Table 7 có THÊM một cột A rỗng
    (merged A1:A42) nên mọi cột lệch sang phải 1 ô so với Table 5 / Table 8.
    Hardcode chỉ số cột là đọc lệch cột tiền của 21/36 dòng.
    """
    for i, row in enumerate(rows):
        labels = {_s(c): j for j, c in enumerate(row) if _s(c)}
        if all(m in labels for m in HEADER_MARKERS):
            return i, {COL_MAP[k]: j for k, j in labels.items() if k in COL_MAP}
    return None, None


def _payment_meta(wb):
    """Ngày thanh toán + số chứng từ thanh toán, lấy từ 2 nguồn rồi đối chiếu."""
    pay_date, pay_doc, pay_amount = None, "", None
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for c in row:
                t = _s(c)
                if "Ngày thanh toán" in t:
                    # Khối text nhiều dòng: nhãn ở dòng trên, giá trị ở dòng dưới.
                    lines = [x.strip() for x in t.split("\n")]
                    for k, ln in enumerate(lines):
                        if ln.startswith("Ngày thanh toán") and k + 1 < len(lines):
                            pay_date = pay_date or _date(lines[k + 1])
    # Table 9: bảng 'Chứng từ thanh toán | Ngày | Đơn vị tiền tệ | Số tiền'
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        for i, row in enumerate(rows):
            if _s(row[0] if row else "") == "Chứng từ thanh toán" and i + 1 < len(rows):
                nxt = rows[i + 1]
                pay_doc = _s(nxt[0])
                pay_date = _date(nxt[1]) or pay_date
                pay_amount = _num(nxt[3]) if len(nxt) > 3 else None
    return pay_date, pay_doc, pay_amount


def parse_wincommerce(path):
    """Đọc file -> list[dict] theo khoá chuẩn. Không bỏ sót sheet nào."""
    wb = openpyxl.load_workbook(path, data_only=True)
    pay_date, pay_doc, pay_amount = _payment_meta(wb)
    out = []

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        hi, cols = _find_header(rows)

        if hi is None:
            # Sheet không có bảng dữ liệu: khối thông tin bên mua/bên bán,
            # thông tin thanh toán, lời cảm ơn, hoặc chân trang chuyển số.
            for j, row in enumerate(rows):
                if _blank(row):
                    continue
                txt = " / ".join(_s(c) for c in row if _s(c))
                kind = "bo_qua"
                amt = None
                if "Số dư mang sang trang sau" in txt:
                    # SỐ KIỂM TRA: cộng dồn hết trang trước. KHÔNG phải dòng tiền
                    # mới — cộng vào tổng là nhân đôi tiền của Table 5.
                    m = re.search(r"Số dư mang sang trang sau\s+([\d.,]+)\s+([\d.,]+)", txt)
                    if m:
                        amt = _num(m.group(2))
                    txt = "CARRY_FORWARD: " + txt
                out.append(dict(
                    row_kind=kind, inv_series="", inv_no="", inv_date=None,
                    store_code="", store_name="", doc_no="",
                    amount_before_vat=None, vat_amount=None, total_amount=amt,
                    payment_date=pay_date,
                    description="[%s r%d] %s" % (ws.title, j + 1, txt),
                ))
            continue

        # Dòng tiêu đề của bảng
        out.append(dict(
            row_kind="bo_qua", inv_series="", inv_no="", inv_date=None,
            store_code="", store_name="", doc_no="",
            amount_before_vat=None, vat_amount=None, total_amount=None,
            payment_date=pay_date,
            description="[%s r%d] HEADER" % (ws.title, hi + 1),
        ))

        def g(row, key):
            j = cols.get(key)
            return row[j] if j is not None and j < len(row) else None

        for j in range(hi + 1, len(rows)):
            row = rows[j]
            tag = "[%s r%d] " % (ws.title, j + 1)
            base = dict(
                row_kind="bo_qua", inv_series="", inv_no="", inv_date=None,
                store_code="", store_name="", doc_no="",
                amount_before_vat=None, vat_amount=None, total_amount=None,
                payment_date=pay_date, description=tag,
            )
            if _blank(row):
                base["description"] = tag + "EMPTY"
                out.append(base)
                continue
            if _is_star(g(row, "inv_date")) or any(_is_star(c) for c in row):
                base["description"] = tag + "SEPARATOR ****"
                out.append(base)
                continue
            first = _s(row[cols.get("doc_no", 0)] if cols.get("doc_no") is not None else row[0])
            if first.lower().startswith("tổng cộng"):
                base["total_amount"] = _num(g(row, "amount"))
                base["description"] = tag + "TOTAL_CHECK (Tổng cộng)"
                out.append(base)
                continue
            inv_raw = _s(g(row, "invoice_raw"))
            amount = _num(g(row, "amount"))
            if not inv_raw or amount is None:
                # Ví dụ Table 7 r44: mảnh tiêu đề chân trang lặp lại, không có số.
                base["description"] = tag + "STRAY: " + " / ".join(_s(c) for c in row if _s(c))
                out.append(base)
                continue

            series, no = split_invoice(inv_raw)
            disc = _num(g(row, "discount")) or 0.0
            base.update(
                row_kind="thanh_toan",
                inv_series=series, inv_no=no,
                inv_date=_date(g(row, "inv_date")),
                doc_no=_s(g(row, "doc_no")),
                # File KHÔNG tách VAT: chỉ có một cột 'Số tiền' đã gồm thuế.
                amount_before_vat=None, vat_amount=None,
                total_amount=amount,
                description=tag + "recon=%s discount=%s pay_doc=%s"
                            % (_s(g(row, "recon_no")), disc, pay_doc),
            )
            out.append(base)

            # Chiết khấu ghi ở CỘT RIÊNG trên cùng dòng. Khác 0 thì tách ra
            # thành một dòng ghi giảm — file 25.06.2026 toàn 0 nên nhánh này
            # CHƯA được xác minh trên dữ liệu thật.
            if disc:
                d = dict(base)
                d.update(row_kind="chiet_khau", total_amount=disc,
                         description=tag + "CHIET_KHAU (cột riêng)")
                out.append(d)

    return out


# ---- KẾT QUẢ CHẠY THẬT trên
# /root/.claude/uploads/559420b3-493b-5f30-8e64-4157466e05ec/
#   f5e40ba4-Payment_Advice_from_Wincommerce_25.06.2026.xlsx
# Tổng dòng trả về: 85 (86 dòng vật lý - 1 dòng trống Table 6 r2)
# row_kind: {'thanh_toan': 36, 'bo_qua': 49}
# Table 5 = 70.880.508 | Table 7 = 134.593.596 | Table 8 = 40.321.800
# Cộng = 245.795.904
#   == 'Tổng cộng' Table 8 r10        -> KHỚP
#   == Table 9 '******245.795.904*'   -> KHỚP
# Table 5 == 'Số dư mang sang trang sau' Table 6 (70.880.508) -> KHỚP
# chiet_khau = 0 (cột tồn tại nhưng bằng 0 ở 36/36 dòng)
