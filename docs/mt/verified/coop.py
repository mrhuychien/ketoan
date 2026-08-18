"""Parser ĐÃ CHẠY THẬT trên file mẫu của chuỗi coop — bản tham chiếu.

verified = True
sheet = "8 SHEET DỮ LIỆU, không phải 1: Sheet1, Sheet2, Sheet4, Sheet5, Sheet6, Sheet7, Sheet8, Sheet9 (+ Sheet3 RỖNG 1 dòng). Thứ tự vật lý trong workbook là Sheet1, Sheet3, Sheet2, Sheet4..Sheet9 — tên sheet KHÔNG theo thời gian. MỖI SHEET = MỘT LẦN THANH TOÁN RIÊNG (Số chứng từ + Ngày thanh toán + 3 số kiểm tra riêng). Phải duyệt wb.worksheets và bỏ sheet có max_row < 20; tuyệt đối không hardcode 'Sheet1'."
header_row = 18

columns:
  __ghi_chu__ -> 'Header 2 TẦNG: r18 = nhóm (merge ngang), r19 = cột con. Dữ liệu bắt đầu r20. Khuôn giống hệt nhau ở cả 8 sheet (đã kiểm r18/r19 từng sheet).'
  stt -> "A (r18='STT') — STT reset và CÓ TRÙNG (vd Sheet1 r88/r89 cùng =69). KHÔNG dùng làm khóa."
  doc_no -> "B (r18='SỐ HÓA ĐƠN LH') — chứng từ bên Co.op, vd '197-SIPI-122025-1090629'. CHỈ điền ở dòng đầu mỗi nhóm (ô merge) → thuộc tính CẤP NHÓM, phải forward-fill."
  inv_no_raw -> 'C (r18=\'HÓA ĐƠN NCC\' merge C18:F18 / r19=\'HÓA ĐƠN\') — vd "\'P0007272"'
  inv_date -> "D (r19='NGÀY') — datetime thật, không phải chuỗi"
  description -> "E (r19='DIỄN GIẢI') — chứa KÝ HIỆU hóa đơn, vd '1C25THG|BANHDAUXANHHUONGV'"
  total_amount -> "F (r19='TRỊ GIÁ') — tiền THEO TỪNG DÒNG, giữ nguyên dấu"
  discount_rate -> "G (r18='CHIẾT KHẤU' merge G18:H18 / r19='TỈ LỆ') — hằng '17.75%-VAT 8' ở cả 569/569 dòng"
  discount_amount -> "H (r19='THÀNH TIỀN') — chiết khấu THEO TỪNG DÒNG, cùng dấu với TRỊ GIÁ"
  group_payment_amount -> "I (r18='THANH TOÁN' merge I18:K18 / r19='TIỀN', ô dữ liệu merge I:K) — TỔNG CẢ NHÓM, chỉ điền ở dòng đầu nhóm"
  store_name -> "L (r18='COOP', ô dữ liệu merge L:M) — tên siêu thị thành viên, CẤP NHÓM, forward-fill"
  J_K_M -> 'trống hoàn toàn ở vùng dữ liệu (chỉ là phần merge của I và L)'
  meta_batch -> "I5 = 'Tên batch : ...'"
  meta_payment_doc_no -> "I6 = 'Số chứng từ : 26303039'"
  meta_payment_date -> "I7 = 'Ngày thanh toán : 20/01/2026'"
  meta_supplier_code -> "A10 = 'Mã cung cấp: 012556'"
  meta_supplier -> "A11 = 'Nhà cung cấp: 233-Cty CP Hoang Giang'"
  check_total_amount -> "J13='Tổng Tiền' → giá trị ở M13"
  check_total_discount -> "J14='Tổng Giá Trị Chiết Khấu' → M14"
  check_total_payment -> "J15='Tổng Tiền Thanh Toán' → M15"
  footer_total_row -> "dòng có C bắt đầu 'Tổng: ' (vd C120='Tổng: 2,712,443,706'), G=tổng chiết khấu, I=tổng thanh toán — LẶP LẠI 3 số kiểm tra ở r13-r15"
  KHONG_CO -> 'File KHÔNG có cột tiền chưa VAT và KHÔNG có cột thuế GTGT → amount_before_vat = vat_amount = None. Đừng bịa.'

invoice_parse: SỐ HÓA ĐƠN lấy từ cột C 'HÓA ĐƠN NCC', KÝ HIỆU lấy từ cột E 'DIỄN GIẢI' — hai cột khác nhau.

SỐ: bỏ dấu nháy đơn (là KÝ TỰ THẬT trong ô text, 569/569 ô đều bắt đầu bằng ') → bỏ mọi ký tự đầu không phải chữ số → bỏ số 0 độn.
  "'P0007272" → 7272 · "'A00006828" → 6828 · "'A00000860" → 860 · "'a0006911" → 6911 (chữ thường!) · "'6936" → 6936 · "'40" → 40 · "'R79" → 79 · "'DCGZ0003176" → 3176.
  7/569 ô C rỗng (chỉ còn dấu nháy "'") → số hóa đơn nằm ở cột B của CHÍNH DÒNG ĐÓ: 438546, 4673, 'A006996' (Sheet1 r89/113/119), 34752, 34883, 'A000614' (Sheet2 r20/21/58), 1 dòng Sheet6 r20. Fallback sang B rồi áp cùng luật bóc.

KÝ HIỆU: regex (?<![0-9A-Z])(\d?[CK]\d{2}T[A-Z]{2,3})(?![A-Z]) trên DIỄN GIẢI → bắt được 552/569 dòng, 91 ký hiệu khác nhau.
  '1C25THG|BANHDAUXANHHUONGV' → 1C25THG · 'BBTT-1C25THG|BANH' → 1C25THG (có tiền tố BBTT-) · 'RTV2170798-1K26TDL-403|B DXANH' → 1K26TDL (tiền tố RTV) · 'K26TEK-80|BANH' → K26TEK (THIẾU số mẫu đầu → \d? phải optional; theo §E khớp phải bỏ qua chữ số đầu, K26TEK ≡ 1K26TEK).
  Dấu phân cách sau ký hiệu KHÔNG chỉ có '|': | (494) · - (50) · ¦ U+00A6 broken bar (4) · _ (2) · / (1) · khoảng trắng (1). Đừng split('|').

CẢNH BÁO NẶNG — số trong DIỄN GIẢI KHÔNG phải số hóa đơn: 12/569 dòng có con số ngay sau ký hiệu nhưng đó là SỐ PHIẾU RTV.
  Sheet1 r116: "'A00000860" + '1K25TEQ-2153190|BANH DXANH - RTV2153190' → số đúng là 860, KHÔNG phải 2153190.
  Sheet4 r27: "'327" + '1K26TCP|2171247|BDX - RTV2171247' → 327, không phải 2171247.
  Sheet5 r95: "'209" + '1K26THQ-207|VAT8|...' → C và E mâu thuẫn thật (209 vs 207), lấy C, gắn cờ review.
  Sheet1 r68: "'A0006842" + '1C25THG-68742|TPCN' → 6842, không phải 68742.
  ⇒ Luôn lấy số từ cột C. Lấy từ DIỄN GIẢI là gán nhầm hóa đơn.

17/569 dòng KHÔNG có ký hiệu nào trong DIỄN GIẢI (đều là dòng trả hàng/điều chỉnh) → inv_series=None + needs_review=True, KHÔNG đoán.
date_parse: NGÀY (cột D) đã là datetime.datetime thật của openpyxl trên cả 569/569 dòng — chỉ cần .date(), KHÔNG parse chuỗi. Ngày thanh toán KHÔNG có trong bảng, phải lấy từ meta ô I7 mỗi sheet bằng regex r'Ngày thanh toán\s*:\s*(\d{2})/(\d{2})/(\d{4})' → dd/mm/yyyy. Mỗi sheet MỘT ngày thanh toán khác nhau (8 ngày, 20/01/2026 → 22/07/2026), tuyệt đối không coi cả file là một lần thanh toán. Ngày hóa đơn trải rộng 02/2025 → 03/2026 và KHÔNG nằm trong cùng kỳ với ngày thanh toán (vd Sheet1 trả 20/01/2026 nhưng có hóa đơn 19/02/2025) → không được lọc hóa đơn theo tháng thanh toán.
amount_parse: Tất cả tiền là int thuần của openpyxl (không có chuỗi, không dấu phẩy, không ngoặc đơn cho số âm) trên cả 569 dòng — chỉ cần đọc thẳng. Ba con số per-dòng: TRỊ GIÁ (F) và CHIẾT KHẤU/THÀNH TIỀN (H) là CẤP DÒNG, cộng bình thường. THANH TOÁN/TIỀN (I) là CẤP NHÓM, chỉ cộng ở dòng đầu nhóm. Dấu: dòng bán 1C luôn dương, dòng trả 1K luôn âm, chiết khấu cùng dấu với trị giá cùng dòng — nhưng PHẢI phân loại bằng ký hiệu, không bằng dấu. KHÔNG tự tính lại: chiết khấu ≠ round(trị giá×17.75%) ở 68/569 dòng (lệch ±1đ); thanh toán nhóm ≠ Σtrị giá−Σchiết khấu ở 17/374 nhóm (lệch ±1đ). Ba số kiểm tra ở M13/M14/M15 (nhãn ở J13/J14/J15) lặp lại y hệt trên dòng 'Tổng' (C=text 'Tổng: 2,712,443,706', G, I) — dùng M13/M14/M15 vì là số, còn C của dòng Tổng là chuỗi có dấu phẩy.
payment_dates: ["2026-01-20 (Sheet1, số chứng từ 26303039, batch 'Quick Payment: ID=230833284')", "2026-02-23 (Sheet2, số chứng từ 26303420, batch '100-hongphuc-23022026-hoanggiang')", "2026-03-24 (Sheet4, số chứng từ 26304108, batch '100-HONGPHUC-24032026-HOANGGIQNG')", "2026-04-23 (Sheet5, số chứng từ 26304598, batch '100-HONGPHUC-23042026-HOANGGIANG')", "2026-05-25 (Sheet6, số chứng từ 26305083, batch '100-hongphuc-25052026-hoanggiang')", "2026-06-22 (Sheet7, số chứng từ 26305591, batch 'Quick Payment: ID=239779307')", "2026-07-07 (Sheet8, số chứng từ 26305852, batch 'Quick Payment: ID=240368412')", "2026-07-22 (Sheet9, số chứng từ 26306169, batch 'Quick Payment: ID=241704308')"]

row_types:
  {"kind": "bo_qua", "label": "Tiêu đề + header 2 tầng (r1–r19 mỗi sheet)", "count": 152, "sign": "n/a", "rule": "r1–r17 là tiêu đề/meta/số kiểm tra; r18–r19 là header 2 tầng. 19 dòng × 8 sheet = 152. Kiểm bằng cell(18,1)=='STT' và cell(19,5)=='DIỄN GIẢI'."}
  {"kind": "thanh_toan", "label": "Dòng hóa đơn bán hàng (ký hiệu 1C..)", "count": 443, "sign": "dương (443/443 dòng đều > 0)", "rule": "DIỄN GIẢI khớp regex (?<![0-9A-Z])(\\d?[CK]\\d{2}T[A-Z]{2,3})(?![A-Z]) và chữ trong 3 ký tự đầu ký hiệu là 'C'. Ký hiệu chủ yếu 1C25THG (207) và 1C26THG (236). PHÂN LOẠI THEO KÝ HIỆU, KHÔNG theo dấu (ràng buộc §B)."}
  {"kind": "ghi_giam", "label": "Hóa đơn trả hàng / điều chỉnh giảm (ký hiệu 1K..)", "count": 109, "sign": "âm (109/109 dòng đều < 0)", "rule": "Ký hiệu bắt được và chữ trong 3 ký tự đầu là 'K'. 88 ký hiệu K khác nhau (1K25TDS, 1K26TAN, 1K26TBD...). Có trường hợp thiếu số mẫu ở đầu: 'K26TEK-80' (Sheet4 r38) — regex cho phép \\d? nên vẫn bắt được."}
  {"kind": "ghi_giam", "label": "Trả hàng/điều chỉnh KHÔNG có ký hiệu trong DIỄN GIẢI — CẦN NGƯỜI REVIEW", "count": 17, "sign": "âm (17/17 dòng đều < 0)", "rule": "DIỄN GIẢI chỉ ghi số phiếu RTV hoặc ghi chú, vd 'RTV 2174174 HD115 - RTV2174174', 'RTV2179667-188|TPCN', 'DCHINH DO NCC THAY BANG HD 3542 (05/05/2026)', '1K26THLTPCN - RTV2170648'. Vẫn là ghi giảm nhưng KHÔNG suy ra được ký hiệu → gắn needs_review=True, để người quyết định. Riêng '1K26THLTPCN' gần như chắc là ký hiệu 1K26THL viết dính chữ TPCN nhưng KHÔNG có bằng chứng trong file → không đoán."}
  {"kind": "bo_qua", "label": "Dòng 'Tổng: ...' đóng vùng dữ liệu (1 dòng/sheet)", "count": 8, "sign": "n/a", "rule": "Cột C bắt đầu bằng chuỗi 'Tổng'. Đây là DÒNG SỐ KIỂM TRA thứ hai (C=tổng trị giá dạng text có dấu phẩy, G=tổng chiết khấu, I=tổng thanh toán), trùng khớp 100% với M13/M14/M15. Dùng để dừng vòng lặp, KHÔNG được cộng vào dữ liệu."}
  {"kind": "bo_qua", "label": "Chân trang sau dòng Tổng ('Ngày ….. tháng …..', 'Người lập bảng', dòng rỗng, dòng \\xa0) + Sheet3 rỗng", "count": 89, "sign": "n/a", "rule": "88 dòng chân trang (11 dòng × 8 sheet) + 1 dòng của Sheet3 rỗng. Mọi dòng sau dòng 'Tổng' đều bỏ."}
  {"kind": "chiet_khau", "label": "CHIẾT KHẤU LÀ CỘT, KHÔNG PHẢI DÒNG", "count": 569, "sign": "cùng dấu với TRỊ GIÁ cùng dòng (dương ở dòng bán, âm ở dòng trả)", "rule": "KHÔNG cộng con số 569 này vào tổng số dòng (152+443+109+17+8+89 = 818 = đúng tổng dòng workbook). Chiết khấu nằm ở cột H trên CẢ 569 dòng dữ liệu, tỉ lệ G='17.75%-VAT 8' ở 569/569 dòng. Co.op KHÔNG phát sinh dòng chiết khấu riêng và file KHÔNG có dòng phí."}

totals:
  __ket_luan__ = KHỚP TUYỆT ĐỐI — 24/24 số kiểm tra (3 số × 8 sheet) đều lệch 0 đồng.
  toan_file = {'tong_tri_gia': 7538089592, 'tong_chiet_khau': 1338010941, 'tong_tien_thanh_toan': 6200078656}
  theo_loai_dong = {'thanh_toan_1C': {'n': 443, 'tri_gia': 8451787806, 'chiet_khau': 1500192379}, 'ghi_giam_1K': {'n': 109, 'tri_gia': -775160674, 'chiet_khau': -137591023}, 'ghi_giam_khong_ky_hieu': {'n': 17, 'tri_gia': -138537540, 'chiet_khau': -24590415}}
  doi_chieu_tung_sheet = [{'sheet': 'Sheet1', 'so_chung_tu': '26303039', 'ngay_thanh_toan': '2026-01-20', 'n_dong': 100, 'n_nhom': 56, 'tri_gia_doc': 2712443706, 'check_M13': 2712443706, 'chiet_khau_doc': 481458758, 'check_M14': 481458758, 'thanh_toan_doc': 2230984947, 'check_M15': 2230984947, 'khop': True}, {'sheet': 'Sheet2', 'so_chung_tu': '26303420', 'ngay_thanh_toan': '2026-02-23', 'n_dong': 39, 'n_nhom': 26, 'tri_gia_doc': 1000851688, 'check_M13': 1000851688, 'chiet_khau_doc': 177651174, 'check_M14': 177651174, 'thanh_toan_doc': 823200514, 'check_M15': 823200514, 'khop': True}, {'sheet': 'Sheet4', 'so_chung_tu': '26304108', 'ngay_thanh_toan': '2026-03-24', 'n_dong': 40, 'n_nhom': 29, 'tri_gia_doc': 556821486, 'check_M13': 556821486, 'chiet_khau_doc': 98835818, 'check_M14': 98835818, 'thanh_toan_doc': 457985668, 'check_M15': 457985668, 'khop': True}, {'sheet': 'Sheet5', 'so_chung_tu': '26304598', 'ngay_thanh_toan': '2026-04-23', 'n_dong': 91, 'n_nhom': 73, 'tri_gia_doc': 711756873, 'check_M13': 711756873, 'chiet_khau_doc': 126336849, 'check_M14': 126336849, 'thanh_toan_doc': 585420026, 'check_M15': 585420026, 'khop': True}, {'sheet': 'Sheet6', 'so_chung_tu': '26305083', 'ngay_thanh_toan': '2026-05-25', 'n_dong': 169, 'n_nhom': 119, 'tri_gia_doc': 504290072, 'check_M13': 504290072, 'chiet_khau_doc': 89511514, 'check_M14': 89511514, 'thanh_toan_doc': 414778561, 'check_M15': 414778561, 'khop': True}, {'sheet': 'Sheet7', 'so_chung_tu': '26305591', 'ngay_thanh_toan': '2026-06-22', 'n_dong': 33, 'n_nhom': 28, 'tri_gia_doc': 708869934, 'check_M13': 708869934, 'chiet_khau_doc': 125824414, 'check_M14': 125824414, 'thanh_toan_doc': 583045520, 'check_M15': 583045520, 'khop': True}, {'sheet': 'Sheet8', 'so_chung_tu': '26305852', 'ngay_thanh_toan': '2026-07-07', 'n_dong': 40, 'n_nhom': 19, 'tri_gia_doc': 635506398, 'check_M13': 635506398, 'chiet_khau_doc': 112802387, 'check_M14': 112802387, 'thanh_toan_doc': 522704012, 'check_M15': 522704012, 'khop': True}, {'sheet': 'Sheet9', 'so_chung_tu': '26306169', 'ngay_thanh_toan': '2026-07-22', 'n_dong': 57, 'n_nhom': 24, 'tri_gia_doc': 707549435, 'check_M13': 707549435, 'chiet_khau_doc': 125590027, 'check_M14': 125590027, 'thanh_toan_doc': 581959408, 'check_M15': 581959408, 'khop': True}]
  lech_lam_tron_da_truy_ra_nguyen_nhan = {'hien_tuong': 'ΣTRỊ GIÁ − ΣCHIẾT KHẤU ≠ ΣTHANH TOÁN ở 4/8 sheet: Sheet1 +1đ, Sheet5 −2đ, Sheet6 −3đ, Sheet8 −1đ (Sheet2/4/7/9 lệch 0).', 'nguyen_nhan': 'Đã soi tới cấp NHÓM: 17/374 nhóm có (Σ trị giá nhóm − Σ chiết khấu nhóm) lệch đúng ±1đ so với ô TIỀN của nhóm. Co.op làm tròn chiết khấu ở CẤP DÒNG rồi làm tròn số tiền thanh toán ở CẤP NHÓM một cách độc lập → hai phép làm tròn chênh nhau 1đ. Cộng dồn ra ±1..3đ mỗi sheet.', 'he_qua': 'Số kiểm tra M15/dòng Tổng vẫn khớp TUYỆT ĐỐI với Σ cột TIỀN (0đ lệch) → cột TIỀN là số Co.op thực trả, phải coi là NGUỒN SỰ THẬT. TUYỆT ĐỐI KHÔNG tự tính lại thanh toán = trị giá − chiết khấu, và không tự tính lại chiết khấu = trị giá × 17.75% (68/569 dòng lệch ±1đ so với round()).', 'sai_so_cho_phep_khi_doi_chieu': '±1đ mỗi nhóm, ±3đ mỗi sheet'}

BẪY:
  - FILE CÓ 9 SHEET, KHÔNG PHẢI 1 — hợp đồng hiện ghi 'xlsx 1 sheet, 131 dòng' là SAI và sẽ bỏ sót 469/569 dòng (~64% dữ liệu, 4.83 tỷ đồng). 8 sheet có dữ liệu + Sheet3 rỗng làm mồi. Phải sửa docs/mt/mt_payment_advice_contract.md dòng 14 và dòng 50.
  - MỖI SHEET LÀ MỘT LẦN THANH TOÁN RIÊNG — riêng Số chứng từ, riêng Ngày thanh toán, riêng 3 số kiểm tra. Gộp cả file thành một Payment Advice là sai bản chất. Đây là biểu hiện của bẫy §H (nhiều ngày thanh toán/1 file), nhưng ở Co.op nó tách theo SHEET chứ không theo cột như LOTTE.
  - THỨ TỰ SHEET KHÔNG THEO TÊN: workbook xếp Sheet1, Sheet3(rỗng), Sheet2, Sheet4, Sheet5... Phải duyệt wb.worksheets và lọc max_row < 20, không hardcode tên sheet, không sort tên.
  - BẪY TIỀN NHÓM (đã đo thật): cột THANH TOÁN/TIỀN chỉ điền ở dòng đầu mỗi nhóm siêu thị và là tổng CẢ NHÓM (569 dòng nhưng chỉ 374 nhóm). Nếu forward-fill cột này xuống mọi dòng rồi cộng ⇒ 13.843.267.903đ thay vì 6.200.078.656đ — THỔI PHỒNG 2,23 LẦN.
  - Cột SỐ HÓA ĐƠN LH (B) và COOP (L) cũng là ô merge cấp nhóm, chỉ điền ở dòng đầu nhóm (đúng 374/569 ô có giá trị = đúng số nhóm). Phải forward-fill; nhưng ĐỪNG forward-fill cột TIỀN (I) rồi cộng — xem bẫy trên.
  - Ranh giới nhóm = ô cột I có giá trị. Đã đối chiếu với merged_cells.ranges của cột I: khớp 374/374 nhóm, không lệch một dòng nào. Cả 8 sheet đều có I ở dòng 20 nên không có nhóm mồ côi.
  - STT (cột A) BỊ TRÙNG và RESET — Sheet1 r88 và r89 cùng STT=69; Sheet2 STT chạy 1,2 rồi reset về 1. Không dùng STT làm khóa hay để đếm dòng.
  - DẤU NHÁY ĐƠN LÀ KÝ TỰ THẬT trong ô text (569/569 ô C bắt đầu bằng '), không phải escape của Excel. Phải lstrip("'") thủ công.
  - SỐ TRONG DIỄN GIẢI LÀ SỐ PHIẾU RTV, KHÔNG PHẢI SỐ HÓA ĐƠN — 12/569 dòng. Lấy số từ DIỄN GIẢI là gán nhầm hóa đơn. Chỉ lấy từ cột HÓA ĐƠN NCC.
  - 7/569 dòng cột HÓA ĐƠN NCC rỗng (chỉ còn dấu nháy) — số hóa đơn chuyển sang cột SỐ HÓA ĐƠN LH của chính dòng đó. Các dòng này đều thuộc nhóm 'LIÊN HIỆP HỢP TÁC XÃ THƯƠNG MẠI TP.HỒ CHÍ MINH' (văn phòng, không phải siêu thị thành viên).
  - DẤU PHÂN CÁCH TRONG DIỄN GIẢI KHÔNG CHỈ LÀ '|' — có cả ¦ (U+00A6 broken bar, 4 dòng), '_', '-', '/', khoảng trắng, và cả \xa0 (non-breaking space) xen giữa (Sheet6 r140). split('|') sẽ hỏng.
  - KÝ HIỆU THIẾU SỐ MẪU ĐẦU: 'K26TEK-80' (Sheet4 r38) không có chữ số đầu, trong khi các dòng khác là '1K26...'. Đúng như §E của hợp đồng: khớp phải bỏ qua chữ số đầu.
  - KÝ HIỆU DÍNH LIỀN CHỮ KHÁC: '1K26THLTPCN - RTV2170648' (Sheet6 r164) — không tách được an toàn, regex có lookahead nên trả None và gắn needs_review thay vì đoán bừa.
  - 17/569 dòng KHÔNG có ký hiệu nào (chỉ ghi số RTV hoặc 'DCHINH DO NCC THAY BANG HD 3542'). Không được suy ký hiệu — gắn needs_review, để người quyết định.
  - CHIẾT KHẤU LÀ CỘT, KHÔNG PHẢI DÒNG — Co.op không phát sinh dòng chiết khấu riêng, cũng không có dòng phí. Đừng đi tìm dòng chiết khấu như Emart/WinCommerce.
  - KHÔNG TỰ TÍNH LẠI TIỀN: chiết khấu ≠ round(TRỊ GIÁ × 17,75%) ở 68/569 dòng (lệch ±1đ); tiền thanh toán nhóm ≠ Σtrị giá − Σchiết khấu ở 17/374 nhóm (lệch ±1đ). Co.op làm tròn ở cấp dòng và cấp nhóm độc lập nhau. Phải đọc số Co.op ghi, không tái tính.
  - FILE KHÔNG CÓ CỘT TIỀN CHƯA VAT VÀ KHÔNG CÓ CỘT THUẾ GTGT. Ô tỉ lệ '17.75%-VAT 8' chỉ là nhãn text (569/569 dòng giống hệt), KHÔNG dùng để suy ngược tiền thuế. amount_before_vat và vat_amount phải để None.
  - Chữ hoa/thường lẫn lộn trong số hóa đơn: "'a0006911", "'a0007482" (Sheet1 r22/r23) — regex bóc chữ đầu phải bắt cả chữ thường, hoặc dùng ^[^0-9]+.
  - Ngày hóa đơn KHÔNG cùng kỳ với ngày thanh toán — Sheet1 trả 20/01/2026 nhưng chứa hóa đơn từ 19/02/2025. Không lọc hóa đơn theo tháng của lần thanh toán.
  - Chân trang có dòng chỉ chứa '\xa0\xa0\xa0\xa0\xa0 ' ở cột A (cuối mỗi sheet) — dòng này KHÔNG rỗng theo openpyxl nhưng là rác. Dừng ở dòng 'Tổng' là an toàn nhất.
  - Số kiểm tra xuất hiện HAI LẦN mỗi sheet: M13/M14/M15 (số) và dòng 'Tổng' cuối bảng (C là TEXT có dấu phẩy 'Tổng: 2,712,443,706', G/I là số). Dùng M13/M14/M15; nếu dùng C của dòng Tổng phải strip 'Tổng:' và bỏ dấu phẩy.

CHƯA XÁC MINH:
  - TRỊ GIÁ (cột F) đã bao gồm VAT hay chưa — file không có cột nào tách thuế. Nhãn '17.75%-VAT 8' chỉ nói chiết khấu 17,75% và thuế suất 8%, không nói F là giá trị nào. PHẢI hỏi kế toán MT trước khi khớp với Sales Invoice (khớp nhầm base sẽ lệch 8%).
  - Ý nghĩa chính xác của 'SỐ HÓA ĐƠN LH' (vd '197-SIPI-122025-1090629') — vẫn là mục §I của hợp đồng, chưa xác minh. Nhưng ĐÃ xác minh được: tiền tố số đầu (197, 199, 200...) là MÃ SIÊU THỊ THÀNH VIÊN — 120 tiền tố khác nhau, ánh xạ 1:1 sang tên ở cột COOP, KHÔNG có tiền tố nào ứng với 2 tên. Có thể dùng làm store_code, nhưng chưa được kế toán xác nhận.
  - Ánh xạ 120 mã/tên siêu thị Co.op → Customer của ERPNext. Tên ở cột COOP viết cả có dấu lẫn không dấu, cả en-dash '–' lẫn hyphen '-' cho cùng một đơn vị (vd 'CN LIEN HIEP HTX TM TP.HCM–CO.OPMART BA RIA' vs '...-CO.OPMART...'), và 'CTY TNHH MTV TMDV SÀI GÒN - HÀ TĨNH' xuất hiện ở 3 nhóm khác nhau trong cùng Sheet1. KHÔNG khớp Customer bằng chuỗi tên.
  - Dòng Sheet7 r21: DIỄN GIẢI 'DCHINH DO NCC THAY BANG HD 3542 (05/05/2026)' với HÓA ĐƠN NCC = "'DCGZ0003176". Luật bóc số cho ra 3176 nhưng diễn giải nhắc HD 3542 — chưa rõ chứng từ nào là chứng từ cần ghi giảm. Phải để người xử lý.
  - Ký hiệu thật của dòng Sheet6 r164 ('1K26THLTPCN') — gần như chắc là 1K26THL nhưng không có bằng chứng trong file.
  - Có phải MỌI file Co.op đều nhiều sheet không, hay chỉ file này gộp 8 kỳ. Mới có 1 file Co.op thật. Parser đã viết theo hướng 'n sheet' nên an toàn cả hai chiều, nhưng cần file Co.op thứ hai để chốt.
  - Vị trí header r18-r19 và vị trí meta I5/I6/I7, M13/M14/M15 có cố định ở mọi file Co.op không. Trong file này giống hệt ở cả 8 sheet nên parser vẫn assert cứng — nếu Co.op đổi khuôn, parser raise ValueError chứ không đọc sai âm thầm.
  - Quan hệ giữa 'Mã cung cấp: 012556' / 'Nhà cung cấp: 233-Cty CP Hoang Giang' với Supplier/Customer của ERPNext.
  - Chưa xác minh khả năng đọc file này qua ketoan.misa_import._rows() — đây là .xlsx nên openpyxl đọc được, nhưng _rows() hiện chỉ trả sheet đầu tiên (theo mô tả hợp đồng); với file 8 sheet thì phải sửa _rows() hoặc viết reader riêng cho kênh MT.
"""

# -*- coding: utf-8 -*-
"""Đọc bảng kê thanh toán chuỗi Saigon Co.op (.xlsx).

ĐÃ CHẠY THẬT trên /root/.claude/uploads/559420b3-.../2780b47b-HOANGGIANG26_CO.OP.xlsx
Kết quả: 569 dòng dữ liệu + 8 dòng kiểm tra; 24/24 số kiểm tra khớp 0 đồng.
"""
import re, datetime
import openpyxl

# Ký hiệu hóa đơn điện tử VN: [số mẫu]{C|K}{2 số năm}T{2-3 chữ}
# C = hóa đơn bán hàng, K = hóa đơn điều chỉnh/trả hàng.
# Lookaround để không nuốt nhầm khi ký hiệu dính liền chữ khác (vd '1K26THLTPCN').
# \d? optional vì có dòng thiếu số mẫu đầu ('K26TEK-80') — hợp đồng §E: bỏ qua chữ số đầu.
_SERIES = re.compile(r'(?<![0-9A-Z])(\d?[CK]\d{2}T[A-Z]{2,3})(?![A-Z])')
_PAYDATE = re.compile(r'Ngày thanh toán\s*:\s*(\d{2})/(\d{2})/(\d{4})')
_DOCNO = re.compile(r'Số chứng từ\s*:\s*(\S+)')
_STORE = re.compile(r'^(\d+)-SIPI')

COL = {'stt': 1, 'doc_no': 2, 'inv_no': 3, 'inv_date': 4,
       'desc': 5, 'value': 6, 'disc_rate': 7, 'disc_amt': 8,
       'pay': 9, 'store': 12}          # A B C D E F G H I L


def _clean_inv_no(raw):
    """`'P0007272` -> `7272`.

    VÌ SAO: Co.op ghi số hóa đơn NCC kèm dấu nháy đơn THẬT trong ô text,
    kèm tiền tố chữ của quyển (P/A/B/Z/DCG..., có cả chữ thường 'a0006911')
    và số 0 độn. Bỏ nháy -> bỏ mọi ký tự đầu không phải chữ số -> bỏ số 0 thừa.
    KHÔNG lấy số trong DIỄN GIẢI: 12/569 dòng số đó là số phiếu RTV, không
    phải số hóa đơn -> lấy nhầm là gán sai hóa đơn.
    """
    if raw is None:
        return None
    s = str(raw).strip().lstrip("'").strip()
    s = re.sub(r'^[^0-9]+', '', s)
    s = s.lstrip('0')
    return s or None


def _series(desc):
    if not desc:
        return None
    m = _SERIES.search(str(desc))
    return m.group(1) if m else None


def parse_coop_payment_advice(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []

    for ws in wb.worksheets:
        # Sheet rỗng (Sheet3) là sheet mồi -> bỏ. Không được duyệt theo TÊN sheet:
        # thứ tự tên không theo thời gian và có sheet rỗng xen giữa.
        if ws.max_row < 20:
            continue

        # --- header 2 tầng r18-r19: kiểm tra thật, không tin vị trí cứng ---
        # Sai khuôn thì DỪNG, không đọc sai âm thầm rồi hạch toán nhầm tiền.
        if (ws.cell(18, 1).value != 'STT'
                or ws.cell(19, 5).value != 'DIỄN GIẢI'
                or ws.cell(19, 6).value != 'TRỊ GIÁ'
                or ws.cell(19, 9).value != 'TIỀN'):
            raise ValueError('Sheet %s: header r18/r19 không đúng khuôn Co.op' % ws.title)

        # --- meta của LẦN THANH TOÁN (mỗi sheet = MỘT lần thanh toán riêng) ---
        pay_date = None
        m = _PAYDATE.search(str(ws.cell(7, 9).value or ''))
        if m:
            pay_date = datetime.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        m = _DOCNO.search(str(ws.cell(6, 9).value or ''))
        payment_doc = m.group(1) if m else None

        chk_total = ws.cell(13, 13).value      # Tổng Tiền
        chk_disc = ws.cell(14, 13).value       # Tổng Giá Trị Chiết Khấu
        chk_pay = ws.cell(15, 13).value        # Tổng Tiền Thanh Toán

        # --- dòng "Tổng: ..." đóng vùng dữ liệu ---
        end = None
        for r in range(20, ws.max_row + 1):
            v = ws.cell(r, COL['inv_no']).value
            if isinstance(v, str) and v.strip().startswith('Tổng'):
                end = r
                break
        if end is None:
            raise ValueError('Sheet %s: không tìm thấy dòng Tổng' % ws.title)

        # --- dữ liệu ---
        g_doc = g_store = g_code = None
        g_pay = None
        g_key = 0
        for r in range(20, end):
            val = ws.cell(r, COL['value']).value
            disc = ws.cell(r, COL['disc_amt']).value
            pay = ws.cell(r, COL['pay']).value
            desc = ws.cell(r, COL['desc']).value
            raw_inv = ws.cell(r, COL['inv_no']).value
            raw_doc = ws.cell(r, COL['doc_no']).value

            # BẪY: cột THANH TOÁN/TIỀN chỉ điền ở dòng ĐẦU mỗi nhóm siêu thị và
            # là tổng CẢ NHÓM. Cột SỐ HÓA ĐƠN LH + COOP cũng chỉ điền ở dòng đầu
            # (ô merge). Ô có giá trị = mở nhóm mới; các dòng sau kế thừa.
            # Fill cột TIỀN xuống mọi dòng rồi cộng = 13.843.267.903 thay vì
            # 6.200.078.656 -> thổi phồng 2,23 lần. Đã đo thật.
            is_first = pay is not None
            if is_first:
                g_key += 1
                g_pay = pay
                g_doc = raw_doc
                g_store = ws.cell(r, COL['store']).value
                mm = _STORE.match(str(raw_doc)) if isinstance(raw_doc, str) else None
                g_code = mm.group(1) if mm else None

            series = _series(desc)
            # Phân loại theo KÝ HIỆU, không theo dấu tiền (ràng buộc hợp đồng §B).
            # Ký hiệu K = hóa đơn trả hàng/điều chỉnh giảm.
            if series and 'K' in series[:3]:
                kind = 'ghi_giam'
            elif series:
                kind = 'thanh_toan'
            else:
                # 17 dòng DIỄN GIẢI chỉ ghi số phiếu RTV / ghi chú điều chỉnh,
                # không có ký hiệu. Đều là trả hàng -> ghi_giam, nhưng phải
                # đánh dấu để người kiểm tra soi lại. KHÔNG đoán ký hiệu.
                kind = 'ghi_giam_khong_ky_hieu'

            # Số hóa đơn: ưu tiên cột HÓA ĐƠN NCC; 7 dòng cột đó rỗng (chỉ còn
            # dấu nháy) thì số nằm ở cột SỐ HÓA ĐƠN LH của chính dòng đó.
            inv_no = _clean_inv_no(raw_inv)
            inv_no_src = 'HOA_DON_NCC'
            if inv_no is None:
                inv_no = _clean_inv_no(raw_doc)
                inv_no_src = 'SO_HOA_DON_LH' if inv_no else None

            d = ws.cell(r, COL['inv_date']).value
            out.append({
                'row_kind': kind,
                'inv_series': series,
                'inv_no': inv_no,
                'inv_date': d.date() if isinstance(d, datetime.datetime) else d,
                'store_code': g_code,
                'store_name': g_store,
                'doc_no': g_doc,
                'amount_before_vat': None,   # file KHÔNG có cột chưa thuế
                'vat_amount': None,          # file KHÔNG có cột thuế GTGT
                'total_amount': val,         # cột TRỊ GIÁ, giữ nguyên dấu
                'payment_date': pay_date,
                'description': desc,
                # --- phụ trợ để đối chiếu, không thuộc 12 khóa chuẩn ---
                'sheet': ws.title,
                'excel_row': r,
                'discount_amount': disc,     # cột CHIẾT KHẤU/THÀNH TIỀN
                'discount_rate': ws.cell(r, COL['disc_rate']).value,
                'group_key': '%s#%d' % (ws.title, g_key),
                'is_group_first': is_first,
                # chỉ dòng đầu nhóm mới mang số tiền nhóm -> cộng an toàn
                'group_payment_amount': g_pay if is_first else None,
                'inv_no_source': inv_no_src,
                'payment_doc_no': payment_doc,
                'needs_review': series is None,
            })

        # dòng số kiểm tra của sheet -> trả về để đối chiếu ngay trong dữ liệu
        out.append({
            'row_kind': 'kiem_tra',
            'inv_series': None, 'inv_no': None, 'inv_date': None,
            'store_code': None, 'store_name': None,
            'doc_no': payment_doc,
            'amount_before_vat': None, 'vat_amount': None,
            'total_amount': chk_total,
            'payment_date': pay_date,
            'description': 'Tổng Tiền / Tổng Giá Trị Chiết Khấu / Tổng Tiền Thanh Toán',
            'sheet': ws.title, 'excel_row': end,
            'discount_amount': chk_disc, 'discount_rate': None,
            'group_key': None, 'is_group_first': None,
            'group_payment_amount': chk_pay,
            'inv_no_source': None, 'payment_doc_no': payment_doc,
            'needs_review': False,
        })
    return out


if __name__ == '__main__':
    import collections
    P = "/root/.claude/uploads/559420b3-493b-5f30-8e64-4157466e05ec/2780b47b-HOANGGIANG26_CO.OP.xlsx"
    rows = parse_coop_payment_advice(P)
    print('rows:', len(rows))
    print(collections.Counter(r['row_kind'] for r in rows))

    ok = True
    for sh in sorted({r['sheet'] for r in rows}):
        d = [r for r in rows if r['sheet'] == sh and r['row_kind'] != 'kiem_tra']
        k = [r for r in rows if r['sheet'] == sh and r['row_kind'] == 'kiem_tra'][0]
        sF = sum(r['total_amount'] for r in d)
        sH = sum(r['discount_amount'] for r in d)
        # CHỈ cộng ở dòng đầu nhóm -- đây là chỗ dễ nhân tiền lên nhất
        sI = sum(r['group_payment_amount'] for r in d if r['is_group_first'])
        good = (sF == k['total_amount'] and sH == k['discount_amount']
                and sI == k['group_payment_amount'])
        ok = ok and good
        print('%-7s n=%3d TRI GIA %13d/%13d  CK %12d/%12d  TT %13d/%13d  %s  (dF-H-I=%+d) pay=%s'
              % (sh, len(d), sF, k['total_amount'], sH, k['discount_amount'],
                 sI, k['group_payment_amount'], 'OK' if good else 'LECH',
                 sF - sH - sI, k['payment_date']))
    print('ALL CHECKS', 'PASS' if ok else 'FAIL')

# ---------------------------------------------------------------------------
# KẾT QUẢ CHẠY THẬT (đã copy nguyên từ stdout):
# rows: 577
# Counter({'thanh_toan': 443, 'ghi_giam': 109, 'ghi_giam_khong_ky_hieu': 17, 'kiem_tra': 8})
# Sheet1  n=100 TRI GIA 2712443706/2712443706 CK 481458758/481458758 TT 2230984947/2230984947 OK (dF-H-I=+1) pay=2026-01-20
# Sheet2  n= 39 TRI GIA 1000851688/1000851688 CK 177651174/177651174 TT  823200514/ 823200514 OK (dF-H-I=+0) pay=2026-02-23
# Sheet4  n= 40 TRI GIA  556821486/ 556821486 CK  98835818/ 98835818 TT  457985668/ 457985668 OK (dF-H-I=+0) pay=2026-03-24
# Sheet5  n= 91 TRI GIA  711756873/ 711756873 CK 126336849/126336849 TT  585420026/ 585420026 OK (dF-H-I=-2) pay=2026-04-23
# Sheet6  n=169 TRI GIA  504290072/ 504290072 CK  89511514/ 89511514 TT  414778561/ 414778561 OK (dF-H-I=-3) pay=2026-05-25
# Sheet7  n= 33 TRI GIA  708869934/ 708869934 CK 125824414/125824414 TT  583045520/ 583045520 OK (dF-H-I=+0) pay=2026-06-22
# Sheet8  n= 40 TRI GIA  635506398/ 635506398 CK 112802387/112802387 TT  522704012/ 522704012 OK (dF-H-I=-1) pay=2026-07-07
# Sheet9  n= 57 TRI GIA  707549435/ 707549435 CK 125590027/125590027 TT  581959408/ 581959408 OK (dF-H-I=+0) pay=2026-07-22
# ALL CHECKS PASS
#
# dF-H-I != 0 ở 4 sheet là do Co.op làm tròn chiết khấu cấp DÒNG và làm tròn
# tiền thanh toán cấp NHÓM độc lập nhau (17/374 nhóm lệch đúng +/-1 đồng).
# Số kiểm tra M15 vẫn khớp tuyệt đối với tổng cột TIỀN -> cột TIỀN là số
# Co.op thực trả, KHÔNG được tự tính lại.
